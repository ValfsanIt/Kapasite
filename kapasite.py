from datetime import datetime, timedelta
import plotly.graph_objs as go
import ast
from dash import no_update
from dash import dash_table
from dash import html
import dash_bootstrap_components as dbc
import pandas as pd
from dash import dcc, Input, Output, State, ctx , no_update
from dash_ag_grid import AgGrid  # type: ignore[reportMissingImports]
from dash.dash_table.Format import Format, Scheme
import os
import sys
import plotly.express as px
from app import app, cache, TIMEOUT
import kapasite_data
from decimal import Decimal
from dash.exceptions import PreventUpdate
import re
import time

_RUN_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "run")
if _RUN_DIR not in sys.path:
    sys.path.append(_RUN_DIR)
from agent import ag  # type: ignore[reportMissingImports]


def _export_styled_excel(df, filename, table_title=None):
    
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.formatting.rule import FormulaRule
    except ImportError:
        buffer = io.BytesIO()
        df.to_excel(buffer, index=False, engine='openpyxl')
        buffer.seek(0)
        return dict(content=buffer.read(), filename=filename, base64=False)

    wb = Workbook()
    ws = wb.active
    ws.title = "Veri"

    
    TITLE_FILL          = PatternFill("solid", fgColor="0D2B5E")
    HEADER_FILL         = PatternFill("solid", fgColor="1565C0")
    FIRST_COL_FILL      = PatternFill("solid", fgColor="2563A8")  
    FIRST_COL_FILL_ALT  = PatternFill("solid", fgColor="1E4F8C")  
    ALT_ROW_FILL        = PatternFill("solid", fgColor="D6E4F7")
    WHITE_FILL          = PatternFill("solid", fgColor="FFFFFF")
  
    CF_YELLOW_FILL      = PatternFill("solid", fgColor="FEF08A")

    TITLE_FONT      = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
    HEADER_FONT     = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    FIRST_COL_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    BODY_FONT       = Font(name="Calibri", size=10, color="0D1B2A")
    CF_FONT         = Font(bold=True, color="1A1A00", name="Calibri", size=10)

    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
    CELL_ALIGN_C = Alignment(horizontal="center", vertical="center")
    CELL_ALIGN_L = Alignment(horizontal="left",   vertical="center")
    CELL_ALIGN_R = Alignment(horizontal="right",  vertical="center")
    thin_side    = Side(style="thin", color="AACCE8")
    BORDER       = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    NUM_FORMAT     = '#,##0'
    NUM_FORMAT_DEC = '#,##0.0'

    num_cols = len(df.columns)

    # ── Tablo Başlık Satırı ─────────────────────────────────────
    data_start_row = 1
    if table_title:
        ws.row_dimensions[1].height = 36
        title_cell = ws.cell(row=1, column=1, value=table_title)
        title_cell.fill      = TITLE_FILL
        title_cell.font      = TITLE_FONT
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        if num_cols > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        ws.row_dimensions[2].height = 6
        data_start_row = 3

    header_row = data_start_row

    # ── Sütun Başlıkları ─────────────────────────────────────────────────────
    ws.row_dimensions[header_row].height = 28
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border    = BORDER

    # ── Veri Satırları ───────────────────────────────────────────────────────
    total_data_rows = len(df)
    for row_offset, row in enumerate(df.itertuples(index=False), start=1):
        r = header_row + row_offset
        ws.row_dimensions[r].height = 20
        is_alt = (row_offset % 2 == 0)
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col_idx, value=value)
            cell.border = BORDER
            if col_idx == 1:
                cell.fill      = FIRST_COL_FILL_ALT if is_alt else FIRST_COL_FILL
                cell.font      = FIRST_COL_FONT
                cell.alignment = CELL_ALIGN_L
            else:
                cell.fill = ALT_ROW_FILL if is_alt else WHITE_FILL
                cell.font = BODY_FONT
                if isinstance(value, (int, float)) and value == value:
                    cell.alignment = CELL_ALIGN_R
                    if isinstance(value, float) and value != int(value):
                        cell.number_format = NUM_FORMAT_DEC
                    else:
                        cell.number_format = NUM_FORMAT
                else:
                    cell.alignment = CELL_ALIGN_C

    # ── Sütun Genişlikleri ───────────────────────────────────────────────────
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        for val in df.iloc[:200, col_idx - 1]:
            try:
                cell_len = len(str(val)) if val is not None else 0
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(8, min(max_len + 3, 40))

    
    freeze_cell = f"B{header_row + 1}"
    ws.freeze_panes = freeze_cell

    # ── Excel Table  ────
    if total_data_rows > 0:
        try:
            table_end_col = get_column_letter(num_cols)
            table_end_row = header_row + total_data_rows
            table_ref     = f"A{header_row}:{table_end_col}{table_end_row}"
            
            tab_style = TableStyleInfo(
                name="TableStyleLight1",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=False,
                showColumnStripes=False,
            )
            excel_table = Table(
                displayName="Tablo1",
                ref=table_ref,
                tableStyleInfo=tab_style,
            )
            ws.add_table(excel_table)
        except Exception:
            ws.auto_filter.ref = ws.dimensions
    else:
        ws.auto_filter.ref = ws.dimensions

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    import base64 as _b64
    encoded = _b64.b64encode(buffer.read()).decode('ascii')
    return dict(
        content=encoded,
        filename=filename,
        base64=True,
        type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



def run_query_safe(sql, retries=3, backoff=0.5):
    for attempt in range(retries):
        try:
            return ag.run_query(sql)
        except Exception as e:
            msg = str(e).lower()
            
            if 'bağlantı başka bir komutun sonuçlarıyla meşgul' in msg or 'busy' in msg or 'already in use' in msg:
                wait = backoff * (attempt + 1)
                time.sleep(wait)
                continue
            # otherwise re-raise
            raise
    # final attempt
    try:
        return ag.run_query(sql)
    except Exception:
        return None


# Yerel run.agent kullanılıyor (bağımsız çalışma)

# Feature-detect AccordionHeader/AccordionBody availability (older dbc versions don't have them)
HAS_ACCORDION_HEADER = hasattr(dbc, "AccordionHeader") and hasattr(dbc, "AccordionBody")


def generate_monthly_columns(selected_year=None):
    """Aylık kolonlar (Öngörü).
    format1/format4 kapasite için; verimlilik sadece yük tablosunda kolon olarak.

    selected_year verilirse sadece seçilen yıla(lar)a ait ayları üretir.
    selected_year None ise mevcut yıl + bir sonraki yıl üretilir.
    """
    today = datetime.today()

    # Normalizasyon: None -> [today, today+1], tek -> [tek], liste -> [liste]
    if selected_year is None or selected_year == []:
        years_to_generate = [today.year, today.year + 1]
    elif isinstance(selected_year, (str, int)):
        years_to_generate = [int(selected_year)]
    elif isinstance(selected_year, list):
        years_to_generate = [int(y) for y in selected_year if y is not None and str(y).strip()]
    else:
        years_to_generate = [int(selected_year)]

    # Deterministik sıra
    years_to_generate = sorted(set(years_to_generate))

    # Format listeleri
    format1, format2, format3, format4 = [], [], [], []

    for year in years_to_generate:
        for month in range(1, 13):
            month_str = str(month).zfill(2)
            ym = f"{year}-{month_str}"

            format1.append(f"SUM(B.[{ym}]) AS [{ym}]")
            format2.append(f"SUM([{ym}]) AS [{ym}], SUM([{ym}A]) AS [{ym}A]")
            format3.append(f"{ym}")
            format3.append(f"{ym}A")
            format4.append(
                f"CAST(CEILING((CAST(SUM(A.[{ym}]) AS DECIMAL(18, 3))/"
                f"CASE WHEN CAST(SUM(B.[{ym}]) AS DECIMAL(18, 3)) = 0 "
                f"THEN 1 ELSE CAST(SUM(B.[{ym}]) AS DECIMAL(18, 3)) "
                f"END)*100) AS int) AS [{ym}]"
            )

    return {
        'format1': ", ".join(format1),
        'format2': ", ".join(format2),
        'format3': ", ".join(format3),
        'format4': ", ".join(format4)
    }


#@cache.memoize(timeout=TIMEOUT)
def generate_weekly_columns():
    
    start_date = datetime.now() - timedelta(weeks=1)
    format1 = []
    format2 = []
    format3 = []
    format4 = []

    for i in range(19):  # Bu hafta dahil olmak üzere toplamda 14 hafta
        week_start = start_date + timedelta(weeks=i)
        year, week_num, _ = week_start.isocalendar()
        wk = f"{year}-{str(week_num).zfill(2)}"

        if i == 0:
            format1.append(f"0 AS [{wk}]")
        else:
            format1.append(f"SUM(B.[{wk}]) AS [{wk}]")

        format2.append(f"SUM([{wk}]) AS [{wk}], SUM([{wk}A]) AS [{wk}A]")

        format3.append(f"{wk}, {wk}A")

        if i == 0:
            format4.append(f"0 AS [{wk}]")
        else:
            format4.append(f"CAST(CEILING((CAST(SUM(A.[{wk}]) AS DECIMAL(18, 3))/CAST(SUM(B.[{wk}]) AS DECIMAL(18, 3)))*100) AS int) AS [{wk}]")

    
    columns_str_dict = {
        'format1': ", ".join(format1),
        'format2': ", ".join(format2),
        'format3': ", ".join(format3),
        'format4': ", ".join(format4)
    }

    return columns_str_dict


def _select_sum_with_unit(kolon_list, selected_units):
    """Birim seçimine göre SELECT ifadesi: saat → SUM([col])/60, vardiya → SUM([col])/510, dakika → SUM([col]). Kolon listesi liste veya virgülle ayrılmış string olabilir."""
    if not kolon_list:
        return ""
    if isinstance(kolon_list, str):
        kolon_list = [c.strip() for c in kolon_list.split(",") if c.strip()]
    divisor = None
    if selected_units and "hours" in selected_units:
        divisor = 60
    elif selected_units and "shifts" in selected_units:
        divisor = 510
    parts = []
    for col in kolon_list:
        if not col or not str(col).strip():
            continue
        if divisor:
            parts.append(f"SUM([{col}])/{divisor} AS [{col}]")
        else:
            parts.append(f"SUM([{col}]) AS [{col}]")
    return ", ".join(parts)


def format_with_thousands_separator(df, kolonlar):
    
    if df is None or kolonlar is None:
        return df

    if isinstance(kolonlar, str):
        kolonlar = kolonlar.split(', ')

    kolonlar = [col for col in kolonlar if col in df.columns]

    for col in kolonlar:
        try:
            df[col] = df[col].apply(
                lambda x: f"{float(x):,.0f}".replace(",", ".") if pd.notnull(x) and is_number(x) else x
            )
        except Exception as e:
            print(f"Hata {col} kolonunda: {e}")
    return df


def _coerce_numeric_records(records, columns):
   
    if not records or not columns:
        return records
    for r in records:
        for c in columns:
            if c not in r:
                continue
            v = r.get(c)
            if v is None:
                continue
            # If already numeric, skip
            if isinstance(v, (int, float)):
                continue
            if isinstance(v, str):
                s = v.strip()
                if s == "" or s.lower() == "nan":
                    r[c] = None
                    continue
                # Remove thousand separators (dots) and convert comma decimal if any
                try:
                    s2 = s.replace(".", "").replace(",", ".")
                    # If value still contains non-numeric, skip
                    r[c] = float(s2)
                except Exception:
                    # leave original if conversion fails
                    pass
    return records


def _blank_zero_cells_in_records(records, columns):
    """UI'da 0 gösterilmesini engellemek için seçili hücreleri boşaltır."""
    if not records or not columns:
        return records
    for r in records:
        for c in columns:
            if c not in r:
                continue
            v = r.get(c)
            if v is None:
                continue
            if isinstance(v, (int, float)):
                if float(v) == 0.0:
                    r[c] = None
                continue
            if isinstance(v, str):
                s = v.strip()
                if s == "":
                    continue
                try:
                    s2 = s.replace(".", "").replace(",", ".")
                    if float(s2) == 0.0:
                        r[c] = None
                except Exception:
                    pass
                continue
            # numpy sayı tipleri vb. int/float dışı numeric değerleri de kapsa
            try:
                if float(v) == 0.0:
                    r[c] = None
            except Exception:
                pass
    return records


def is_number(x):
    try:
        float(x)
        return True
    except (ValueError, TypeError):
        return False


def _clean_df_column_names(df):
   
    if df is None:
        return df
    cols = list(df.columns)
    new_cols = []
    for i, c in enumerate(cols):
        if c is None:
            new = f"COL_{i}"
        else:
            new = str(c).strip()
            if new == "" or new.lower() == "nan":
                new = f"COL_{i}"
        new_cols.append(new)
    df.columns = new_cols
    return df


def _display_col_name(col_id, zero_first_state):
    """Dashboard'da ZIP görünümüyle uyum için ilk tarih kolonu başlığını '0' göster."""
    if not col_id:
        return "\u00a0"

    c = str(col_id).strip()
    if not c:
        return "\u00a0"

    # Örn: "2026-03" ve "2026-03A" için aynı base ayı yakalayalım.
    if zero_first_state.get("available"):
        m = re.match(r"^(?P<base>\d{4}-\d{2})(?P<a>A)?$", c)
        if m:
            base = m.group("base")
            # İlk karşılaştığımız base ayı için hem A'lı hem A'sız kolona "0" yaz.
            if not zero_first_state.get("first_done", False):
                zero_first_state["first_date_base"] = base
                zero_first_state["first_done"] = True

            if zero_first_state.get("first_date_base") == base:
                return "0"

    return col_id


def get_table_columns(table_name):
    """Return set of column names for a table or empty set on error."""
    if not table_name:
        return set()
    try:
        df = ag.run_query(f"SELECT TOP 1 * FROM {table_name}")
        if df is None or (hasattr(df, "empty") and df.empty):
            return set()
        return set(df.columns.tolist())
    except Exception:
        return set()


def filter_graph_cols_against_capacity(kolon_graph_str, capacity_table_name):
    
    if not kolon_graph_str or not capacity_table_name:
        return ""
    try:
        col_names = re.findall(r'\[([^\]]+)\]', kolon_graph_str)
        if not col_names:
            return ""
        actual = get_table_columns(capacity_table_name)
        allowed = [c for c in col_names if c in actual]
        if not allowed:
            return ""
        return ", ".join([f"SUM(B.[{c}]) AS [{c}]" for c in allowed])
    except Exception:
        return ""


def is_number(x):
    try:
        float(x)
        return True
    except (ValueError, TypeError):
        return False


def format_percentage_columns(df, columns):
    for col in columns:
        def _fmt(x):
            if pd.isna(x) or x is None:
                return x
            try:
                return f"{float(x):.1f}%"
            except (TypeError, ValueError):
                return x
        df[col] = df[col].apply(_fmt)
    return df


################################################################################
# ORTAK TABLO STİLLERİ
################################################################################
def _get_table_style() -> dict:
    """Tüm tablolar (Yük, Kapasite Süresi, Malzeme) aynı yükseklik alanını kullanır."""
    return {
        'overflowY': 'auto',
        'overflowX': 'auto',
        'width': '100%',
        'border': 'none',
        'borderCollapse': 'separate',
        'height': '100%',
        'minHeight': '400px',
        'maxHeight': 'none',
    }



_TABLE_STYLE = _get_table_style()
_CELL_STYLE  = {'textAlign': 'center', 'border': '1px solid rgba(30,41,59,0.45)',
                        'minWidth': '90px', 'backgroundColor': '#ffffff',
                        'color': '#0d1b2a', 'fontSize': '15px', 'fontFamily': 'Inter, Segoe UI, sans-serif',
                        'whiteSpace': 'nowrap', 'overflow': 'hidden', 'textOverflow': 'ellipsis'}
_HEADER_STYLE       = {'fontWeight': 'bold', 'minWidth': '90px',
                        'backgroundColor': 'rgba(48, 184, 180, 1)',  # Canlıyla birebir aynı turkuaz
                        'borderBottom': '1px solid black',
                        'color': 'white', 'fontSize': '14px',
                        'fontFamily': 'Inter, Segoe UI, sans-serif',
                        'minHeight': '44px', 'padding': '10px 8px',
                        'overflow': 'visible', 'lineHeight': '1.3', 'verticalAlign': 'middle'}

# Sabit (sol) kolonlar — canlı teal (CSS + zebra override ile uyumlu)
_FIXED_PIN_COL_BG = '#14b8a6'
_FIXED_PIN_COL_FG = '#ffffff'
_FIXED_PIN_HEADER_BG = '#0d9488'


def _fixed_pin_cell_rule(column_id, min_w, max_w):
    return {
        'if': {'column_id': column_id},
        'backgroundColor': _FIXED_PIN_COL_BG,
        'color': _FIXED_PIN_COL_FG,
        'fontWeight': '900',
        'fontSize': '13px',
        'minWidth': min_w,
        'maxWidth': max_w,
        'textAlign': 'left',
        'paddingLeft': '14px',
        'overflow': 'visible',
    }


def _fixed_pin_header_rule(column_id, min_w, max_w):
    return {
        'if': {'column_id': column_id},
        'backgroundColor': _FIXED_PIN_HEADER_BG,
        'color': '#ffffff',
        'fontWeight': '900',
        'fontSize': '14px',
        'minWidth': min_w,
        'maxWidth': max_w,
        'textAlign': 'left',
        'paddingLeft': '14px',
    }


def _fixed_pin_style_data_overrides(column_ids):
    """Zebra (odd/even) kurallarından sonra eklenmeli; sabit kolonları soluk bırakmaz."""
    return [
        {
            'if': {'column_id': cid},
            'backgroundColor': _FIXED_PIN_COL_BG,
            'color': _FIXED_PIN_COL_FG,
            'fontWeight': '900',
        }
        for cid in column_ids
    ]


def _section_badge(icon, title, subtitle=""):
    
    return html.Div([
        html.Span(icon, className="kap-badge-icon"),
        html.Div([
            html.Span(title, className="kap-badge-title"),
            html.Span(subtitle, className="kap-badge-sub") if subtitle else None,
        ], className="kap-badge-text"),
    ], className="kap-section-badge")


def _ctrl_btn(label, btn_id, icon="⬇", variant="primary"):
    return html.Button(
        [html.Span(icon, className="btn-icon"), html.Span(label)],
        id=btn_id, n_clicks=0,
        className=f"kap-btn kap-btn-{variant}"
    )


def _acc_header(num, title, subtitle, color_cls):
    
    
    if subtitle:
        return f"{num} {title} — {subtitle}"
    return f"{num} {title}"


def _acc_title_str(num, title, subtitle):
    """
    dbc.AccordionItem title prop'u düz string bekler.
    Görsel stil CSS ile uygulanır.
    """
    return f"{num}  {title}  —  {subtitle}"


def _fullscreen_panel(panel_id, children, close_button_id=None):
    close_id = close_button_id or f"close-{panel_id}"
    return html.Div([
        html.Div([
            html.Div([
                html.Span("▣", className="kap-topbar-icon"),
                html.Span("PANEL AÇIK", className="kap-topbar-label"),
                html.Span("ESC veya ✕ ile kapatabilirsiniz", className="kap-topbar-hint"),
            ], className="kap-topbar-left"),
            html.Button(
                "✕",
                id=close_id,
                n_clicks=0,
                **{'data-panel': panel_id},
                className="kap-topbar-close",
                title="Kapat (ESC)"
            ),
        ], className="kap-topbar"),
        html.Div(children, className="kap-fullscreen-content"),
    ], id=panel_id, className="kap-fullscreen-panel")


def _selection_summary_box(scope_id, total_id, avg_id):
    return html.Div(
        className="kap-selection-info",
        style={"marginTop": "10px", "padding": "8px 12px", "border": "1px solid rgba(71,85,105,.25)", "borderRadius": "10px"},
        children=[
            html.Div(id=scope_id, style={"fontWeight": "800", "marginBottom": "4px"}),
            html.Span("Toplam: ", style={"opacity": 0.85}),
            html.Span(id=total_id, style={"fontWeight": "700", "marginRight": "14px"}),
            html.Span("Ortalama: ", style={"opacity": 0.85}),
            html.Span(id=avg_id, style={"fontWeight": "700"}),
        ],
    )






layout = dbc.Container([
    # ── Hidden stores ──────────────────────────────────────────────
    dcc.Store(id='isfirst_trigger', data=0),
    dcc.Store(id='iscapacity_trigger', data=0),
    dcc.Store(id='table_name'),
    dcc.Store(id='capacity_table_name'),
    dcc.Store(id='kolon_sumb'),
    dcc.Store(id='kolon_sum'),
    dcc.Store(id='kolon_list'),
    dcc.Store(id='kolon_graph'),
    dcc.Store(id='filtered_kolon_list'),
    dcc.Store(id='filtered_kolon_sum'),
    dcc.Store(id='filtered_kolon_graph'),
    dcc.Store(id='filtered_kolon_sumb'),
    dcc.Store(id='costcenter-last-click', data=None),
    dcc.Store(id='workcenter-last-click', data=None),
    dcc.Store(id='capcost-last-click', data=None),
    dcc.Store(id='capwc-last-click', data=None),
    dcc.Store(id='material-last-click', data=None),
    # ── Toggle durumu store'ları (A kolonlarını göster/gizle) ──────
    dcc.Store(id='costcenter-col-toggle-state', data=False),       # False = başlangıçta A kolonları gizli
    dcc.Store(id='workcenter-col-toggle-state', data=False),
    dcc.Store(id='cap-costcenter-col-toggle-state', data=False),
    dcc.Store(id='cap-workcenter-col-toggle-state', data=False),
    dcc.Store(id='material-col-toggle-state', data=False),

    # ── PAGE HEADER ────────────────────────────────────────────────
    html.Div([
        html.Div(className="kap-header-glow"),
        html.Div(
            id='veri-tipi-loading-hint',
            className='kap-header-veri-tipi-loading',
            style={'display': 'none'},
            children=[
                html.Span(className='kap-veri-tipi-spinner kap-veri-tipi-spinner--header', **{'aria-hidden': 'true'}),
                html.Span('Veriler Güncelleniyor', className='kap-veri-tipi-loading-text kap-veri-tipi-loading-text--header'),
            ],
        ),
        html.Div([
            html.Span("◈", className="kap-header-icon"),
            html.Div([
                html.H1("KAPASİTE ANALİZ", className="kap-header-title"),
                html.P("Üretim kapasitesi & yük dengesi gerçek zamanlı takip sistemi",
                       className="kap-header-sub"),
            ]),
        ], className="kap-header-inner"),
        html.Div([
            html.Div(className="kap-pulse-dot"),
            html.Span("CANLI", className="kap-live-label"),
        ], className="kap-live-badge"),
    ], className="kap-page-header"),

    # ── KONTROL PANELİ ─────────────────────────────────────────────
    html.Div([
        dcc.Interval(id='interval-component', interval=1*1000, n_intervals=0, max_intervals=1),

        dbc.Row([
            # Veri Tipi | Costcenter | Yıl (Öngörü'de) | Zaman Birimi — yan yana
            dbc.Col([
                html.Label([html.Span("📊", style={'marginRight': '6px'}), "Veri Tipi"],
                           className="kap-label kap-label-control"),
                dcc.Dropdown(
                    id='data-type-dropdown',
                    options=["İhtiyaç Miktarı", "Sipariş Miktarı", "Öngörü Miktarı"],
                    value="İhtiyaç Miktarı",
                    className="kap-dropdown"
                ),
            ], md=3, className="mb-0"),
            dbc.Col([
                html.Label([html.Span("🏭", style={'marginRight': '6px'}), "Costcenter"],
                           className="kap-label kap-label-control"),
                dcc.Dropdown(
                    id='costcenter-dropdown',
                    options=[],
                    value=None,
                    placeholder="Costcenter seçiniz...",
                    className="kap-dropdown"
                ),
            ], md=4, className="mb-0"),
            # Yıl — Öngörü seçildiğinde görünür, Zaman Birimi'nin hemen önünde
            dbc.Col([
                html.Div(
                    [
                        html.Label([html.Span("📅", style={'marginRight': '6px'}), "Yıl"],
                                   className="kap-label kap-label-control"),
                        dcc.Dropdown(
                            id='year-selector',
                            options=[],
                            value=None,
                            multi=True,
                            style={'display': 'none'},
                            className="kap-dropdown kap-dropdown-year"
                        ),
                    ],
                    id='year-selector-wrap',
                    style={'display': 'none'},
                    className="kap-year-wrap",
                ),
            ], md=3, className="mb-0"),
            dbc.Col([
                html.Label([html.Span("🕐", style={'marginRight': '6px'}), "Zaman Birimi"],
                           className="kap-label kap-label-control"),
                html.Div(
                    dcc.Checklist(
                        id='unit-checkbox',
                        options=[
                            {'label': 'Dakika', 'value': 'minutes'},
                            {'label': 'Saat',   'value': 'hours'},
                            {'label': 'Vardiya','value': 'shifts'},
                        ],
                        value=['minutes'],
                        inline=True,
                        labelStyle={"display": "inline-flex", "marginRight": "20px", "alignItems": "center"},
                        inputStyle={"marginRight": "6px"},
                        className="kap-checklist",
                        labelClassName="kap-check-label",
                        inputClassName="kap-check-input",
                    ),
                    className="kap-checklist-wrap",
                ),
            ], md=4, className="mb-0 kap-col-zaman-birimi"),
        ], className="g-3 align-items-end"),
        dcc.Interval(
            id='veri-tipi-load-hide',
            interval=2000,
            n_intervals=0,
            disabled=True,
        ),

    ], className="kap-control-panel"),

    # ── AKORDİYON ─────────────────────────────────────────────────
    # active_item değişince Python callback panel'i açar/kapatır.
    html.Div([
        dbc.Accordion(
            id='kapasite-accordion',
            start_collapsed=True,
            always_open=False,
            active_item=None,
            className="kap-accordion",
            children=[

                # ─ 1. Costcenter Analizi ─
                dbc.AccordionItem(
                    item_id="acc-item-1",
                    className="kap-accordion-item kap-acc-blue",
                    title=_acc_title_str("01", "Costcenter Analizi", "Yük tablosu · Kapasite · Grafik"),
                    children=[html.Div()],
                ),

                # ─ 2. Workcenter Analizi ─
                dbc.AccordionItem(
                    item_id="acc-item-3",
                    className="kap-accordion-item kap-acc-green",
                    title=_acc_title_str("02", "Workcenter Analizi", "Makine bazlı yük · Kapasite · Malzeme · Grafik"),
                    children=[html.Div()],
                ),

            ]
        ),
    ], className="kap-accordion-wrap"),

    # ── Gizli kapat butonu (JS callback için — accordion'u kapatır) ────────
    html.Button(id='kapasite-close-all', style={'display': 'none'}, n_clicks=0),



    
    _fullscreen_panel(
        panel_id="panel-costcenter",
        close_button_id="close-panel-costcenter",
        children=[

            # ── SEKMELİ ÜST BAR ──
            html.Div([
                html.Button("📊  Yük Tablosu",    id="tab-btn-yuk",     n_clicks=0, className="kap-tab-btn kap-tab-active"),
                html.Button("⬛  Kapasite Süresi", id="tab-btn-kap",     n_clicks=0, className="kap-tab-btn"),
                html.Button("📈  Grafik",            id="tab-btn-grafik",  n_clicks=0, className="kap-tab-btn"),
                dcc.Store(id="active-tab-store", data="yuk"),
            ], className="kap-tab-bar"),

            # ── İÇERİK ALANI ──
            html.Div([

                # ══ SEKME 1: YÜK TABLOSU ══════════════════════════════════
                html.Div([
                    html.Div([
                        _section_badge("▦", "Costcenter Yük Tablosu", "Haftalık süre dağılımı — tıklayarak satır seçin"),
                        html.Div([
                            _ctrl_btn("Tabloyu İndir", "btn-download-costcenter", "⬇"),
                            dcc.Download(id='download-costcenter'),
                            html.Button(
                                [html.Span("⇄", className="btn-icon"), html.Span("Kolonları Göster/Gizle")],
                                id="toggle-columns-costcenter", n_clicks=0,
                                className="kap-btn kap-btn-ghost"
                            ),
                            html.Div(id="selection-info-costcenter-yuk", className="kap-selection-info"),
                            dcc.Store(id="costcenter-hidden-toggle", data=True),
                        ], className="kap-table-actions"),
                        dash_table.DataTable(
                            id='costcenter_table',
                            style_table=_TABLE_STYLE,
                            style_cell={**_CELL_STYLE, 'minWidth': '110px', 'maxWidth': '160px', 'fontSize': '14px'},
                            style_header={**_HEADER_STYLE, 'minWidth': '110px', 'maxWidth': '160px'},
                            style_data_conditional=[
                                {'if': {'row_index': 'odd'}, 'backgroundColor': '#eef4ff'},
                            ] + _fixed_pin_style_data_overrides(['STAND']),
                            fixed_rows={'headers': True},
                            fixed_columns={'headers': True, 'data': 1},
                            row_selectable="multi",
                            column_selectable="multi",
                            cell_selectable=True,
                            selected_rows=[],
                            selected_cells=[],
                            include_headers_on_copy_paste=True,
                            active_cell=None,
                            style_cell_conditional=[
                                _fixed_pin_cell_rule('STAND', '150px', '220px'),
                            ],
                            style_header_conditional=[
                                _fixed_pin_header_rule('STAND', '150px', '220px'),
                            ],
                        ),
                        _selection_summary_box(
                            "selection-agg-scope-costcenter",
                            "selection-agg-total-costcenter",
                            "selection-agg-avg-costcenter",
                        ),
                        dcc.Store(id='cc_yuk_multi_sel_flag', data=False),
                    ], className="kap-tab-panel-inner", style={'width': '100%', 'height': '100%', 'display': 'flex', 'flexDirection': 'column'}),
                ], id="tab-panel-yuk", style={
                    'display': 'flex', 'flexDirection': 'column', 'width': '100%', 'height': '100%',
                    'padding': '12px 24px 16px', 'boxSizing': 'border-box',
                }),

                # ══ SEKME 2: KAPASİTE SÜRESİ (COSTCENTER) ══════════════════
                html.Div([
                    html.Div([
                        _section_badge("⬛", "Kapasite Süresi (Costcenter)", "Workcenter bazlı vardiya kapasite süresi"),
                        html.Div([
                            _ctrl_btn("İndir", "btn-download-costcenter_kapasite", "⬇"),
                            dcc.Download(id='download-costcenter_kapasite'),
                            html.Button(
                                [html.Span("⇄", className="btn-icon"), html.Span("Kolonları Göster/Gizle")],
                                id="toggle-columns-capacity-costcenter", n_clicks=0,
                                className="kap-btn kap-btn-ghost"
                            ),
                            html.Div(id="selection-info-costcenter-kap", className="kap-selection-info"),
                            dcc.Store(id="capacity-costcenter-hidden-toggle", data=True),
                        ], className="kap-table-actions"),
                        dash_table.DataTable(
                            id='capasity_table_costcenter',
                            style_table=_TABLE_STYLE,
                            style_cell={**_CELL_STYLE, 'minWidth': '110px', 'maxWidth': '160px', 'fontSize': '14px'},
                            style_header={**_HEADER_STYLE, 'minWidth': '110px', 'maxWidth': '160px'},
                            style_data_conditional=[
                                {'if': {'row_index': 'odd'}, 'backgroundColor': '#eef4ff'},
                            ] + _fixed_pin_style_data_overrides(['STAT']),
                            fixed_rows={'headers': True},
                            fixed_columns={'headers': True, 'data': 1},
                            style_cell_conditional=[
                                _fixed_pin_cell_rule('STAT', '180px', '240px'),
                            ],
                            style_header_conditional=[
                                _fixed_pin_header_rule('STAT', '180px', '240px'),
                            ],
                            row_selectable="multi",
                            column_selectable="multi",
                            cell_selectable=True,
                            selected_rows=[],
                            selected_cells=[],
                            include_headers_on_copy_paste=True,
                            active_cell=None,
                        ),
                        _selection_summary_box(
                            "selection-agg-scope-capcost",
                            "selection-agg-total-capcost",
                            "selection-agg-avg-capcost",
                        ),
                        dcc.Store(id='cc_kap_multi_sel_flag', data=False),
                    ], className="kap-tab-panel-inner", style={'width': '100%', 'height': '100%', 'display': 'flex', 'flexDirection': 'column'}),
                ], id="tab-panel-kap", style={
                    'display': 'none', 'flexDirection': 'column', 'width': '100%', 'height': '100%',
                    'padding': '12px 24px 16px', 'boxSizing': 'border-box',
                }),

                # ══ SEKME 3: GRAFİK ══════════════════════════════════════════
                html.Div([
                    html.Div([
                        _section_badge("📈", "Costcenter Kapasite Grafiği", "Seçili costcenter için haftalık yük/kapasite karşılaştırması"),
                        html.Div(id="selection-info-costcenter-grafik", className="kap-selection-info", style={"marginBottom": "8px"}),
                        html.Div([
                            dcc.Graph(
                                id='fig',
                                config={'displayModeBar': True, 'responsive': True},
                                style={'height': '100%', 'width': '100%'},
                            ),
                        ], style={'flex': '1 1 0%', 'minHeight': '0'}),
                    ], className="kap-tab-panel-inner", style={'width': '100%', 'height': '100%', 'display': 'flex', 'flexDirection': 'column'}),
                ], id="tab-panel-grafik", style={
                    'display': 'none', 'flexDirection': 'column', 'width': '100%', 'height': '100%',
                    'padding': '12px 24px 16px', 'boxSizing': 'border-box',
                }),

            ], style={
                'flex': '1 1 0%',
                'minHeight': '0',
                'overflow': 'auto',
                'display': 'flex',
                'flexDirection': 'column',
                'background': '#ffffff',
            }),
        ]
    ),

    
    _fullscreen_panel(
        panel_id="panel-workcenter",
        close_button_id="close-panel-workcenter",
        children=[

            
            html.Div([
                html.Button("📈  Grafik",            id="wc-tab-btn-grafik",  n_clicks=0, className="kap-tab-btn kap-tab-active"),
                html.Button("▦  Yük Tablosu",       id="wc-tab-btn-yuk",     n_clicks=0, className="kap-tab-btn"),
                html.Button("⬛  Kapasite Süresi",   id="wc-tab-btn-kap",     n_clicks=0, className="kap-tab-btn"),
                html.Button("📦  Malzeme",           id="wc-tab-btn-malzeme", n_clicks=0, className="kap-tab-btn"),
                dcc.Store(id="wc-active-tab-store", data="grafik"),
            ], className="kap-tab-bar"),

            # ── İÇERİK ALANI ──
            html.Div([

                # ══ SEKME 1: GRAFİK ══════════════════════════════════════════
                html.Div([
                    html.Div([
                        _section_badge("📈", "Workcenter Grafik", "Makine bazlı kapasite yük oranı"),
                        html.Div(id="selection-info-workcenter-grafik", className="kap-selection-info", style={"marginBottom": "8px"}),
                        html.Div([
                            # 1) Kapasite Grubu
                            html.Div([
                                html.Label(
                                    [html.Span("⬛", style={'marginRight': '6px'}), "Kapasite Grubu"],
                                    className="kap-label kap-label-control",
                                ),
                                dcc.Dropdown(
                                    id='workcenter-capacity-dropdown',
                                    options=[],
                                    value='Kapasite Grubu',
                                    className="kap-dropdown",
                                ),
                            ], style={'flex': '1', 'minWidth': '0'}),

                            # 2) İş Merkezi
                            html.Div([
                                html.Label(
                                    [html.Span("🏭", style={'marginRight': '6px'}), "İş merkezi"],
                                    className="kap-label kap-label-control",
                                ),
                                dcc.Dropdown(
                                    id='workcenter-dropdown',
                                    options=[],
                                    value='Hepsi',
                                    placeholder="İş merkezi seçiniz...",
                                    className="kap-dropdown",
                                ),
                            ], style={'flex': '1', 'minWidth': '0'}),
                        ], style={'display': 'flex', 'gap': '12px', 'marginBottom': '14px'}),
                        html.Div([
                            dcc.Graph(
                                id='figx',
                                config={'displayModeBar': True, 'responsive': True},
                                style={'height': '100%', 'width': '100%'},
                            ),
                        ], style={'flex': '1 1 0%', 'minHeight': '0'}),
                    ], className="kap-tab-panel-inner", style={'width': '100%', 'height': '100%', 'display': 'flex', 'flexDirection': 'column'}),
                ], id="wc-tab-panel-grafik", style={
                    'display': 'flex', 'flexDirection': 'column', 'width': '100%', 'height': '100%',
                    'padding': '12px 24px 16px', 'boxSizing': 'border-box',
                }),

                # ══ SEKME 2: YÜK TABLOSU ══════════════════════════════════════
                html.Div([
                    html.Div([
                        _section_badge("▦", "Workcenter Yük Tablosu", "Makine bazlı haftalık yük dağılımı — tıklayarak satır seçin"),
                        html.Div([
                            _ctrl_btn("İndir", "btn-download-workcenter", "⬇"),
                            dcc.Download(id='download-workcenter'),
                            html.Button(
                                [html.Span("⇄", className="btn-icon"), html.Span("Kolonları Göster/Gizle")],
                                id="toggle-columns-workcenter", n_clicks=0,
                                className="kap-btn kap-btn-ghost"
                            ),
                            html.Div(id="selection-info-workcenter-yuk", className="kap-selection-info"),
                            dcc.Store(id="workcenter-hidden-toggle", data=True),
                        ], className="kap-table-actions"),
                        dash_table.DataTable(
                            id='workcenter_table',
                            style_table=_TABLE_STYLE,
                            style_cell={**_CELL_STYLE, 'minWidth': '110px', 'maxWidth': '180px', 'fontSize': '14px'},
                            style_header={**_HEADER_STYLE, 'minWidth': '110px', 'maxWidth': '180px'},
                            style_data_conditional=[
                                {'if': {'row_index': 'odd'}, 'backgroundColor': '#eef4ff'},
                            ] + _fixed_pin_style_data_overrides(['CAPWORK']),
                            fixed_rows={'headers': True},
                            fixed_columns={'headers': True, 'data': 1},
                            style_cell_conditional=[
                                _fixed_pin_cell_rule('CAPWORK', '150px', '220px'),
                            ],
                            style_header_conditional=[
                                _fixed_pin_header_rule('CAPWORK', '150px', '220px'),
                            ],
                            row_selectable="multi",
                            column_selectable="multi",
                            cell_selectable=True,
                            selected_rows=[],
                            selected_cells=[],
                            include_headers_on_copy_paste=True,
                            active_cell=None,
                        ),
                        _selection_summary_box(
                            "selection-agg-scope-workcenter",
                            "selection-agg-total-workcenter",
                            "selection-agg-avg-workcenter",
                        ),
                        dcc.Store(id='wc_yuk_multi_sel_flag', data=False),
                    ], className="kap-tab-panel-inner", style={'width': '100%', 'height': '100%', 'display': 'flex', 'flexDirection': 'column'}),
                ], id="wc-tab-panel-yuk", style={
                    'display': 'none', 'width': '100%', 'height': '100%',
                    'padding': '12px 24px 16px', 'boxSizing': 'border-box',
                    'flexDirection': 'column',
                }),

                # ══ SEKME 3: KAPASİTE SÜRESİ (WORKCENTER) ════════════════════
                html.Div([
                    html.Div([
                        _section_badge("⬛", "Kapasite Süresi (Workcenter)", "Makine bazlı vardiya kapasite süresi — tıklayarak satır seçin"),
                        html.Div([
                            _ctrl_btn("İndir", "btn-download-workcenter_kapasite", "⬇"),
                            dcc.Download(id='download-workcenter_kapasite'),
                            html.Button(
                                [html.Span("⇄", className="btn-icon"), html.Span("Kolonları Göster/Gizle")],
                                id="toggle-columns-capacity-workcenter", n_clicks=0,
                                className="kap-btn kap-btn-ghost"
                            ),
                            html.Div(id="selection-info-workcenter-kap", className="kap-selection-info"),
                            dcc.Store(id="capacity-workcenter-hidden-toggle", data=True),
                        ], className="kap-table-actions"),
                        dash_table.DataTable(
                            id='capasity_table_workcenter',
                            style_table=_TABLE_STYLE,
                            style_cell={**_CELL_STYLE, 'minWidth': '110px', 'maxWidth': '160px', 'fontSize': '14px'},
                            style_header={**_HEADER_STYLE, 'minWidth': '110px', 'maxWidth': '160px'},
                            style_data_conditional=[
                                {'if': {'row_index': 'odd'}, 'backgroundColor': '#eef4ff'},
                            ] + _fixed_pin_style_data_overrides(['STAT']),
                            fixed_rows={'headers': True},
                            fixed_columns={'headers': True, 'data': 1},
                            style_cell_conditional=[
                                _fixed_pin_cell_rule('STAT', '180px', '240px'),
                            ],
                            style_header_conditional=[
                                _fixed_pin_header_rule('STAT', '180px', '240px'),
                            ],
                            row_selectable="multi",
                            column_selectable="multi",
                            cell_selectable=True,
                            selected_rows=[],
                            selected_cells=[],
                            include_headers_on_copy_paste=True,
                            active_cell=None,
                        ),
                        _selection_summary_box(
                            "selection-agg-scope-capwork",
                            "selection-agg-total-capwork",
                            "selection-agg-avg-capwork",
                        ),
                        dcc.Store(id='wc_kap_multi_sel_flag', data=False),
                    ], className="kap-tab-panel-inner", style={'width': '100%', 'height': '100%', 'display': 'flex', 'flexDirection': 'column'}),
                ], id="wc-tab-panel-kap", style={
                    'display': 'none', 'width': '100%', 'height': '100%',
                    'padding': '12px 24px 16px', 'boxSizing': 'border-box',
                    'flexDirection': 'column',
                }),

                # ══ SEKME 4: MALZEME TABLOSU (Workcenter seçimine göre filtrelenir) ═══════════
                html.Div([
                    html.Div([
                        _section_badge("📦", "Malzeme Yük Tablosu", "Seçilen workcenter'a göre malzeme bazlı yük dağılımı"),
                        html.Div([
                            _ctrl_btn("İndir", "btn-download-malzeme", "⬇"),
                            dcc.Download(id='download-malzeme'),
                            html.Button(
                                [html.Span("⇄", className="btn-icon"), html.Span("Kolonları Göster/Gizle")],
                                id="toggle-columns-material", n_clicks=0,
                                className="kap-btn kap-btn-ghost"
                            ),
                            html.Div(id="selection-info-workcenter-malzeme", className="kap-selection-info"),
                            dcc.Store(id="material-hidden-toggle", data=True),
                        ], className="kap-table-actions"),
                        dash_table.DataTable(
                            id='material_table',
                            style_table=_TABLE_STYLE,
                            style_cell={**_CELL_STYLE, 'minWidth': '110px', 'maxWidth': '160px', 'fontSize': '14px'},
                            style_header={**_HEADER_STYLE, 'minWidth': '110px', 'maxWidth': '160px'},
                            style_data_conditional=[
                                {'if': {'row_index': 'odd'}, 'backgroundColor': '#eef4ff'},
                            ] + _fixed_pin_style_data_overrides(['MATERIAL', 'DRAWNUM']),
                            fixed_rows={'headers': True},
                            fixed_columns={'headers': True, 'data': 2},
                            style_cell_conditional=[
                                _fixed_pin_cell_rule('MATERIAL', '150px', '220px'),
                                _fixed_pin_cell_rule('DRAWNUM', '150px', '220px'),
                            ],
                            style_header_conditional=[
                                _fixed_pin_header_rule('MATERIAL', '150px', '220px'),
                                _fixed_pin_header_rule('DRAWNUM', '150px', '220px'),
                            ],
                            row_selectable="multi",
                            column_selectable="multi",
                            cell_selectable=True,
                            selected_rows=[],
                            selected_cells=[],
                            include_headers_on_copy_paste=True,
                            active_cell=None,
                        ),
                        _selection_summary_box(
                            "selection-agg-scope-material",
                            "selection-agg-total-material",
                            "selection-agg-avg-material",
                        ),
                        dcc.Store(id='material_multi_sel_flag', data=False),
                    ], className="kap-tab-panel-inner", style={'width': '100%', 'height': '100%', 'display': 'flex', 'flexDirection': 'column'}),
                ], id="wc-tab-panel-malzeme", style={
                    'display': 'none', 'width': '100%', 'height': '100%',
                    'padding': '12px 24px 16px', 'boxSizing': 'border-box',
                    'flexDirection': 'column',
                }),

            ], style={
                'flex': '1 1 0%',
                'minHeight': '0',
                'overflow': 'auto',
                'display': 'flex',
                'flexDirection': 'column',
                'background': '#ffffff',
            }),
        ]
    ),
], fluid=True, className="kap-page-container")


# ─────────────────────────────────────────────────────────────────────────────
# CALLBACKS
# ─────────────────────────────────────────────────────────────────────────────

@app.callback(
     Output("table_name", "data"),
     Output('unit-checkbox', 'value'),
     Output("kolon_sumb", "data"),
     Output("kolon_sum", "data"),
     Output("kolon_list", "data"),
     Output("kolon_graph", "data"),
     Output("capacity_table_name", "data"),
     Input("interval-component", "n_intervals"),
     Input('unit-checkbox', 'value'),
     Input('data-type-dropdown', 'value'),
     Input('year-selector', 'value'),
    State("table_name", "data")
)
def selected_table(n,selected_units,selected_type,selected_year,current_table_name):
    
    trigger_id = ctx.triggered_id if hasattr(ctx, "triggered_id") else None

    if 'İhtiyaç Miktarı' in selected_type:
        table_name = 'VLFCAPFINALPIVOT'
        capacity_table_name = 'VLFVARDIYASURE'
        weekly_columns = generate_weekly_columns()

        format1_sumb = weekly_columns['format1']
        format2_sum = weekly_columns['format2']
        format3_list = weekly_columns['format3']
        format4_graph = weekly_columns['format4']

    elif 'Sipariş Miktarı' in selected_type:
        table_name = 'VLFCAPFINALSIPARIS'
        capacity_table_name='VLFVARDIYASURE'
        weekly_columns = generate_weekly_columns()

        format1_sumb = weekly_columns['format1']
        format2_sum = weekly_columns['format2']
        format3_list = weekly_columns['format3']
        format4_graph = weekly_columns['format4']




    elif 'Öngörü Miktarı' in selected_type:
        table_name = 'VLFCAPFINALOY'
        capacity_table_name='VLFVARDIYASUREAY'
        # year-selector sadece Öngörü'de görünür; burada seçili yılı baz al.
        monthly_columns = generate_monthly_columns(selected_year=selected_year)

        format1_sumb = monthly_columns['format1']
        format2_sum = monthly_columns['format2']
        format3_list = monthly_columns['format3']
        format4_graph = monthly_columns['format4']

    else:
        raise PreventUpdate

    if len(selected_units) > 1:
        selected_units = [selected_units[-1]]

    
    if trigger_id in ('data-type-dropdown', 'interval-component'):
        table_output = table_name
    else:
        table_output = no_update

  
    if trigger_id == 'unit-checkbox':
        unit_output = selected_units
    else:
        unit_output = no_update

    return table_output, unit_output, format1_sumb, format2_sum, format3_list, format4_graph, capacity_table_name


def _selection_info_children(veri_tipi, costcenter_value):
    """Veri tipi ve costcenter için navbar'larda gösterilecek metin bileşenini döndürür."""
    vt = veri_tipi or "—"
    cc = costcenter_value if costcenter_value else "—"
    return html.Span(
        [html.Span("Veri tipi: ", style={"opacity": 0.85}), html.Span(vt), " · ",
         html.Span("Costcenter: ", style={"opacity": 0.85}), html.Span(cc)],
        className="kap-selection-info-text",
    )


@app.callback(
    Output("selection-info-costcenter-yuk", "children"),
    Output("selection-info-costcenter-kap", "children"),
    Output("selection-info-costcenter-grafik", "children"),
    Output("selection-info-workcenter-grafik", "children"),
    Output("selection-info-workcenter-yuk", "children"),
    Output("selection-info-workcenter-kap", "children"),
    Output("selection-info-workcenter-malzeme", "children"),
    Input("data-type-dropdown", "value"),
    Input("costcenter-dropdown", "value"),
)
def update_selection_info_labels(veri_tipi, costcenter_value):
    children = _selection_info_children(veri_tipi, costcenter_value)
    return (children,) * 7


_VERI_TIPI_HINT_VISIBLE = {'display': 'flex'}
_VERI_TIPI_HINT_HIDDEN = {'display': 'none'}


@app.callback(
    Output('veri-tipi-loading-hint', 'style'),
    Output('veri-tipi-load-hide', 'disabled'),
    Output('veri-tipi-load-hide', 'n_intervals'),
    Input('data-type-dropdown', 'value'),
    Input('costcenter-dropdown', 'value'),
    Input('workcenter-capacity-dropdown', 'value'),
    Input('workcenter-dropdown', 'value'),
    prevent_initial_call=True,
)
def veri_tipi_show_loading_hint(_value, _cc, _capgrp, _wc):
    return _VERI_TIPI_HINT_VISIBLE, False, 0


@app.callback(
    Output('veri-tipi-loading-hint', 'style', allow_duplicate=True),
    Output('veri-tipi-load-hide', 'disabled', allow_duplicate=True),
    Input('veri-tipi-load-hide', 'n_intervals'),
    prevent_initial_call=True,
)
def veri_tipi_hide_loading_hint(n):
    if n is None or n < 1:
        raise PreventUpdate
    return _VERI_TIPI_HINT_HIDDEN, True


@app.callback(
    Output('costcenter_table', 'selected_cells'),
    Output('costcenter_table', 'active_cell'),
    Output('cc_yuk_multi_sel_flag', 'data'),
    Output('capasity_table_costcenter', 'selected_cells'),
    Output('capasity_table_costcenter', 'active_cell'),
    Output('cc_kap_multi_sel_flag', 'data'),
    Output('workcenter_table', 'selected_cells'),
    Output('workcenter_table', 'active_cell'),
    Output('wc_yuk_multi_sel_flag', 'data'),
    Output('capasity_table_workcenter', 'selected_cells'),
    Output('capasity_table_workcenter', 'active_cell'),
    Output('wc_kap_multi_sel_flag', 'data'),
    Output('material_table', 'selected_cells'),
    Output('material_table', 'active_cell'),
    Output('material_multi_sel_flag', 'data'),
    Input('costcenter_table', 'selected_cells'),
    Input('costcenter_table', 'active_cell'),
    Input('capasity_table_costcenter', 'selected_cells'),
    Input('capasity_table_costcenter', 'active_cell'),
    Input('workcenter_table', 'selected_cells'),
    Input('workcenter_table', 'active_cell'),
    Input('capasity_table_workcenter', 'selected_cells'),
    Input('capasity_table_workcenter', 'active_cell'),
    Input('material_table', 'selected_cells'),
    Input('material_table', 'active_cell'),
    State('cc_yuk_multi_sel_flag', 'data'),
    State('cc_kap_multi_sel_flag', 'data'),
    State('wc_yuk_multi_sel_flag', 'data'),
    State('wc_kap_multi_sel_flag', 'data'),
    State('material_multi_sel_flag', 'data'),
    prevent_initial_call=True,
)
def _clear_multi_cell_highlight_on_next_click(
    cc_selected_cells,
    cc_active_cell,
    cc_kap_selected_cells,
    cc_kap_active_cell,
    wc_selected_cells,
    wc_active_cell,
    wc_kap_selected_cells,
    wc_kap_active_cell,
    material_selected_cells,
    material_active_cell,
    cc_yuk_multi_flag,
    cc_kap_multi_flag,
    wc_yuk_multi_flag,
    wc_kap_multi_flag,
    material_multi_flag,
):
    """
    Shift ile çoklu hücre seçildiğinde highlight göster.
    Kullanıcı bir sonraki tıklamada o tablonun active_cell'ını değiştirirse,
    çoklu highlight'ı tamamen kapat (selected_cells=[], active_cell=None).
    """
    # Varsayılan: hiçbir tabloyu değiştirme.
    out_cc_sel = no_update
    out_cc_active = no_update
    out_cc_flag = no_update

    out_cc_kap_sel = no_update
    out_cc_kap_active = no_update
    out_cc_kap_flag = no_update

    out_wc_sel = no_update
    out_wc_active = no_update
    out_wc_flag = no_update

    out_wc_kap_sel = no_update
    out_wc_kap_active = no_update
    out_wc_kap_flag = no_update

    out_mat_sel = no_update
    out_mat_active = no_update
    out_mat_flag = no_update

    def _sel_len(sel):
        return len(sel) if sel else 0

    prop_id = ctx.triggered[0]['prop_id'] if ctx.triggered else ''
    if '.' in prop_id:
        comp_id, prop = prop_id.rsplit('.', 1)
    else:
        comp_id, prop = prop_id, None

    # 1) Shift/çoklu seçim tamamlanırken (selected_cells değiştiğinde) flag set et.
    if comp_id == 'costcenter_table' and prop == 'selected_cells':
        out_cc_flag = (_sel_len(cc_selected_cells) > 1)
    if comp_id == 'capasity_table_costcenter' and prop == 'selected_cells':
        out_cc_kap_flag = (_sel_len(cc_kap_selected_cells) > 1)
    if comp_id == 'workcenter_table' and prop == 'selected_cells':
        out_wc_flag = (_sel_len(wc_selected_cells) > 1)
    if comp_id == 'capasity_table_workcenter' and prop == 'selected_cells':
        out_wc_kap_flag = (_sel_len(wc_kap_selected_cells) > 1)
    if comp_id == 'material_table' and prop == 'selected_cells':
        out_mat_flag = (_sel_len(material_selected_cells) > 1)

    # 2) Sonra kullanıcı active_cell'a tıklayınca (aktif hücre değişince) çoklu highlight'ı kaldır.
    if comp_id == 'costcenter_table' and prop == 'active_cell' and cc_yuk_multi_flag:
        out_cc_sel = []
        out_cc_active = None
        out_cc_flag = False

    if comp_id == 'capasity_table_costcenter' and prop == 'active_cell' and cc_kap_multi_flag:
        out_cc_kap_sel = []
        out_cc_kap_active = None
        out_cc_kap_flag = False

    if comp_id == 'workcenter_table' and prop == 'active_cell' and wc_yuk_multi_flag:
        out_wc_sel = []
        out_wc_active = None
        out_wc_flag = False

    if comp_id == 'capasity_table_workcenter' and prop == 'active_cell' and wc_kap_multi_flag:
        out_wc_kap_sel = []
        out_wc_kap_active = None
        out_wc_kap_flag = False

    if comp_id == 'material_table' and prop == 'active_cell' and material_multi_flag:
        out_mat_sel = []
        out_mat_active = None
        out_mat_flag = False

    return (
        out_cc_sel,
        out_cc_active,
        out_cc_flag,
        out_cc_kap_sel,
        out_cc_kap_active,
        out_cc_kap_flag,
        out_wc_sel,
        out_wc_active,
        out_wc_flag,
        out_wc_kap_sel,
        out_wc_kap_active,
        out_wc_kap_flag,
        out_mat_sel,
        out_mat_active,
        out_mat_flag,
    )


def _try_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        # bool'i sayısal kabul etmeyelim
        if isinstance(v, bool):
            return None
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return None
        # Binlik ayırıcıları temizle: 1.234,56 -> 1234.56
        s2 = s.replace(".", "").replace(",", ".")
        try:
            return float(s2)
        except Exception:
            return None
    return None


def _format_num(v):
    if v is None:
        return "—"
    try:
        # Excel mantığı: çok küçük hatalar için yuvarla
        if isinstance(v, float):
            return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return f"{v:,}".replace(",", ".")
    except Exception:
        return str(v)


@app.callback(
    Output("selection-agg-scope-costcenter", "children"),
    Output("selection-agg-total-costcenter", "children"),
    Output("selection-agg-avg-costcenter", "children"),
    Output("selection-agg-scope-capcost", "children"),
    Output("selection-agg-total-capcost", "children"),
    Output("selection-agg-avg-capcost", "children"),
    Output("selection-agg-scope-workcenter", "children"),
    Output("selection-agg-total-workcenter", "children"),
    Output("selection-agg-avg-workcenter", "children"),
    Output("selection-agg-scope-capwork", "children"),
    Output("selection-agg-total-capwork", "children"),
    Output("selection-agg-avg-capwork", "children"),
    Output("selection-agg-scope-material", "children"),
    Output("selection-agg-total-material", "children"),
    Output("selection-agg-avg-material", "children"),
    Input("costcenter_table", "selected_cells"),
    Input("costcenter_table", "active_cell"),
    Input("costcenter_table", "selected_rows"),
    Input("costcenter_table", "selected_columns"),
    Input("capasity_table_costcenter", "selected_cells"),
    Input("capasity_table_costcenter", "active_cell"),
    Input("capasity_table_costcenter", "selected_rows"),
    Input("capasity_table_costcenter", "selected_columns"),
    Input("workcenter_table", "selected_cells"),
    Input("workcenter_table", "active_cell"),
    Input("workcenter_table", "selected_rows"),
    Input("workcenter_table", "selected_columns"),
    Input("capasity_table_workcenter", "selected_cells"),
    Input("capasity_table_workcenter", "active_cell"),
    Input("capasity_table_workcenter", "selected_rows"),
    Input("capasity_table_workcenter", "selected_columns"),
    Input("material_table", "selected_cells"),
    Input("material_table", "active_cell"),
    Input("material_table", "selected_rows"),
    Input("material_table", "selected_columns"),
    State("costcenter_table", "data"),
    State("costcenter_table", "columns"),
    State("capasity_table_costcenter", "data"),
    State("capasity_table_costcenter", "columns"),
    State("workcenter_table", "data"),
    State("workcenter_table", "columns"),
    State("capasity_table_workcenter", "data"),
    State("capasity_table_workcenter", "columns"),
    State("material_table", "data"),
    State("material_table", "columns"),
)
def update_selection_summary(
    cc_sel, cc_active,
    cc_sel_rows, cc_sel_cols,
    cc_kap_sel, cc_kap_active,
    cc_kap_sel_rows, cc_kap_sel_cols,
    wc_sel, wc_active,
    wc_sel_rows, wc_sel_cols,
    wc_kap_sel, wc_kap_active,
    wc_kap_sel_rows, wc_kap_sel_cols,
    mat_sel, mat_active,
    mat_sel_rows, mat_sel_cols,
    cc_data, cc_columns,
    cc_kap_data, cc_kap_columns,
    wc_data, wc_columns,
    wc_kap_data, wc_kap_columns,
    mat_data, mat_columns,
):
    """
    Excel mantığı:
    - selected_cells varken: o hücrelerin toplam/ortalaması
    - selected_cells aynı satırda ve o satırdaki tüm kolonlar seçiliyse: tam satırın toplam/ortalaması
    - selected_cells aynı sütunda ve o sütundaki tüm satırlar seçiliyse: tam sütunun toplam/ortalaması
    - selected_cells yok ama active_cell varsa: aktif hücre
    """
    def _normalize_active(active, cols):
        if not active or not isinstance(active, dict):
            return None
        if active.get("column_id") is not None:
            return active
        col_idx = active.get("column")
        if col_idx is None or not cols:
            return active
        try:
            idx = int(col_idx)
            if 0 <= idx < len(cols):
                cid = cols[idx].get("id")
                if cid is not None:
                    return {**active, "column_id": cid}
        except Exception:
            pass
        return active

    def _cells_from_sel_or_active(label, sel, active, sel_rows, sel_cols, data, cols):
        if data and cols and sel_rows:
            values = []
            col_ids_local = [c.get("id") for c in cols if c and c.get("id") is not None]
            for r in sel_rows:
                if isinstance(r, int) and 0 <= r < len(data):
                    for cid in col_ids_local:
                        values.append(data[r].get(cid))
            return label, values, f"{label} (Satır)"
        if data and cols and sel_cols:
            values = []
            for cid in sel_cols:
                for r in range(len(data)):
                    values.append(data[r].get(cid))
            return label, values, f"{label} (Sütun)"
        if sel and len(sel) > 0:
            return label, sel, data, cols
        a = _normalize_active(active, cols or [])
        if a:
            return label, [a], data, cols
        return None

    table_map = [
        ("costcenter_table", "Yük (Costcenter)", cc_sel, cc_active, cc_sel_rows, cc_sel_cols, cc_data, cc_columns),
        ("capasity_table_costcenter", "Kapasite Süresi (Costcenter)", cc_kap_sel, cc_kap_active, cc_kap_sel_rows, cc_kap_sel_cols, cc_kap_data, cc_kap_columns),
        ("workcenter_table", "Yük (Workcenter)", wc_sel, wc_active, wc_sel_rows, wc_sel_cols, wc_data, wc_columns),
        ("capasity_table_workcenter", "Kapasite Süresi (Workcenter)", wc_kap_sel, wc_kap_active, wc_kap_sel_rows, wc_kap_sel_cols, wc_kap_data, wc_kap_columns),
        ("material_table", "Malzeme Yük", mat_sel, mat_active, mat_sel_rows, mat_sel_cols, mat_data, mat_columns),
    ]

    chosen = None
    trig = ctx.triggered_id if hasattr(ctx, "triggered_id") else None
    trig_s = str(trig) if trig is not None else ""
    if trig:
        for tid, lab, sel, act, sel_rows, sel_cols, dat, coldef in table_map:
            # Dash: triggered_id çoğu sürümde bileşen id'si (örn. costcenter_table), bazen prop ile birlikte
            if trig_s == tid or trig_s.startswith(f"{tid}."):
                chosen = _cells_from_sel_or_active(lab, sel, act, sel_rows, sel_cols, dat, coldef)
                break

    if not chosen:
        for _tid, lab, sel, act, sel_rows, sel_cols, dat, coldef in table_map:
            c = _cells_from_sel_or_active(lab, sel, act, sel_rows, sel_cols, dat, coldef)
            if c:
                chosen = c
                break

    if not chosen:
        return (
            "Seçim: —", "—", "—",
            "Seçim: —", "—", "—",
            "Seçim: —", "—", "—",
            "Seçim: —", "—", "—",
            "Seçim: —", "—", "—",
        )

    if len(chosen) == 3:
        label, values, forced_scope = chosen
        nums = [_try_float(v) for v in values]
        nums = [n for n in nums if n is not None]
        if not nums:
            scope_text = f"{label} - Seçim"
            total_text = "—"
            avg_text = "—"
        else:
            scope_text = forced_scope
            total_text = _format_num(sum(nums))
            avg_text = _format_num(sum(nums) / len(nums))
    else:
        label, selected_cells, data, columns = chosen
        if not data or not columns:
            scope_text = f"{label} - Seçim"
            total_text = "—"
            avg_text = "—"
            values = None
        else:
            values = "continue"

    if values == "continue":
        # DataTable 'columns' = [{'name':..., 'id':...}, ...]
        col_ids = [c.get("id") for c in columns if c and c.get("id") is not None]
        if not col_ids:
            scope_text = f"{label} - Seçim"
            total_text = "—"
            avg_text = "—"
        else:
            col_id_set = set(col_ids)

            # selected_cells içinde column_id eksikse columns ile tamamla
            fixed_cells = []
            for c in selected_cells:
                if not c:
                    continue
                if c.get("column_id") is None and c.get("column") is not None:
                    try:
                        ci = int(c["column"])
                        if 0 <= ci < len(columns):
                            cid = columns[ci].get("id")
                            if cid is not None:
                                c = {**c, "column_id": cid}
                    except Exception:
                        pass
                fixed_cells.append(c)
            selected_cells = fixed_cells

            sel_row_idxs = sorted(set(c.get("row") for c in selected_cells if c and "row" in c))
            sel_col_ids = set(c.get("column_id") for c in selected_cells if c and c.get("column_id") is not None)

            # Tam satır: tek satır + seçilen sütun kümesi tüm tablo sütunlarıyla aynı + hücre sayısı uyumu
            row_mode = False
            if len(sel_row_idxs) == 1 and selected_cells:
                r = sel_row_idxs[0]
                same_row = all(c.get("row") == r for c in selected_cells)
                if same_row and sel_col_ids == col_id_set and len(selected_cells) == len(col_ids):
                    row_mode = True

            # Tam sütun: tek column_id + her satırda bir hücre
            col_mode = False
            if len(sel_col_ids) == 1 and selected_cells:
                cid = next(iter(sel_col_ids))
                rows_in_sel = sorted(set(c.get("row") for c in selected_cells if c and c.get("row") is not None))
                if (
                    all(c.get("column_id") == cid for c in selected_cells)
                    and len(selected_cells) == len(data)
                    and len(rows_in_sel) == len(data)
                    and rows_in_sel == list(range(len(data)))
                ):
                    col_mode = True

            values = []
            if row_mode:
                r = sel_row_idxs[0]
                if 0 <= r < len(data):
                    for cid in col_ids:
                        values.append(data[r].get(cid))
            elif col_mode:
                cid = next(iter(sel_col_ids))
                for r in range(len(data)):
                    values.append(data[r].get(cid))
            else:
                for c in selected_cells:
                    r = c.get("row")
                    cid = c.get("column_id")
                    if r is None or cid is None:
                        continue
                    if 0 <= r < len(data):
                        values.append(data[r].get(cid))

            nums = [_try_float(v) for v in values]
            nums = [n for n in nums if n is not None]
            if not nums:
                scope_text = f"{label} - Seçim"
                total_text = "—"
                avg_text = "—"
            else:
                total = sum(nums)
                avg = total / len(nums)
                scope_text = label
                if row_mode:
                    scope_text = f"{label} (Satır)"
                elif col_mode:
                    scope_text = f"{label} (Sütun)"
                total_text = _format_num(total)
                avg_text = _format_num(avg)

    defaults = {
        "Yük (Costcenter)": ("Seçim: —", "—", "—"),
        "Kapasite Süresi (Costcenter)": ("Seçim: —", "—", "—"),
        "Yük (Workcenter)": ("Seçim: —", "—", "—"),
        "Kapasite Süresi (Workcenter)": ("Seçim: —", "—", "—"),
        "Malzeme Yük": ("Seçim: —", "—", "—"),
    }
    defaults[label] = (scope_text, total_text, avg_text)

    return (
        *defaults["Yük (Costcenter)"],
        *defaults["Kapasite Süresi (Costcenter)"],
        *defaults["Yük (Workcenter)"],
        *defaults["Kapasite Süresi (Workcenter)"],
        *defaults["Malzeme Yük"],
    )


@app.callback(
    Output('filtered_kolon_sum', 'data'),
    Output('filtered_kolon_list', 'data'),
    Output('filtered_kolon_graph', 'data'),
    Output('filtered_kolon_sumb', 'data'),
    Output('costcenter-col-toggle-state', 'data'),
    Input('toggle-columns-costcenter', 'n_clicks'),
    Input('year-selector', 'value'),
    Input('kolon_sum', 'data'),
    Input('kolon_list', 'data'),
    Input('kolon_graph', 'data'),
    Input('kolon_sumb', 'data'),
    State('costcenter-col-toggle-state', 'data'),
)
def toggle_costcenter_columns(n_clicks, selected_years, kolon_sum, kolon_list, kolon_graph, kolon_sumb, show_a_cols):

    if not kolon_sum:
        raise PreventUpdate

    def _split_sql_select_list(select_list: str):
        """
        SQL SELECT listesi virgül ayırır, ama virgül parantez içinde ise bölmez.
        Özellikle format4_graph içindeki DECIMAL(18, 3) gibi durumlar için gerekli.
        """
        if not select_list:
            return []
        parts = []
        buf = []
        depth = 0
        in_square = False
        for ch in select_list:
            if ch == '[':
                in_square = True
                buf.append(ch)
                continue
            if ch == ']':
                in_square = False
                buf.append(ch)
                continue

            if not in_square:
                if ch == '(':
                    depth += 1
                elif ch == ')':
                    depth = max(depth - 1, 0)

            # Top-level (depth==0) virgülde ayır.
            if ch == ',' and depth == 0 and not in_square:
                part = ''.join(buf).strip()
                if part:
                    parts.append(part)
                buf = []
                continue

            buf.append(ch)

        last = ''.join(buf).strip()
        if last:
            parts.append(last)
        return parts

    trigger = ctx.triggered_id if hasattr(ctx, "triggered_id") else None
    if trigger == 'toggle-columns-costcenter':
        new_show = not show_a_cols
    else:
        # yıl seçimi, veri tipi / kolon store güncellemesi — mevcut göster/gizle durumunu koru
        new_show = show_a_cols if show_a_cols is not None else False

    def _filter_sum(s):
        
        if not s:
            return s
        parts = _split_sql_select_list(s)
        if new_show:
            # Hepsini göster — orijinal hali döndür
            return s
        else:
            # A ile biten kolonları filtrele
            filtered = [p for p in parts if not re.search(r'\[\d{4}-\d{2}A\]', p)]
            return ', '.join(filtered)

    def _filter_list(s):
        
        if not s:
            return s
        if isinstance(s, list):
            parts = s
        else:
            parts = [p.strip() for p in s.split(',')]
        if new_show:
            return s
        else:
            filtered = [p for p in parts if not str(p).strip().endswith('A')]
            if isinstance(s, list):
                return filtered
            return ', '.join(filtered)

    filtered_sum = _filter_sum(kolon_sum)
    filtered_list = _filter_list(kolon_list)
    filtered_graph = _filter_sum(kolon_graph) if kolon_graph else kolon_graph
    filtered_sumb = _filter_sum(kolon_sumb) if kolon_sumb else kolon_sumb
    
    if selected_years:
        # multi=True olduğu için selected_years list gelebilir.
        if isinstance(selected_years, list):
            years = [str(y).strip() for y in selected_years if y is not None and str(y).strip()]
        else:
            years = [str(selected_years).strip()]

        def _keep_part(p):
            s = str(p)
            return any(y in s for y in years)

        def _filter_sum_by_year(s):
            if not s:
                return s
            parts = _split_sql_select_list(s)
            filtered = [p for p in parts if _keep_part(p)]
            return ', '.join(filtered)

        def _filter_list_by_year(l):
            if not l:
                return l
            if isinstance(l, str):
                parts = [p.strip() for p in l.split(',')]
                return ', '.join([p for p in parts if _keep_part(p)])
            return [p for p in l if _keep_part(p)]

        filtered_sum = _filter_sum_by_year(filtered_sum)
        filtered_list = _filter_list_by_year(filtered_list)
        filtered_graph = _filter_sum_by_year(filtered_graph) if filtered_graph else filtered_graph
        filtered_sumb = _filter_sum_by_year(filtered_sumb) if filtered_sumb else filtered_sumb

    return filtered_sum, filtered_list, filtered_graph, filtered_sumb, new_show


# ── Year selector visibility & options (show only for Öngörü Miktarı)
@app.callback(
    Output('year-selector', 'options'),
    Output('year-selector', 'value'),
    Output('year-selector', 'style'),
    Output('year-selector-wrap', 'style'),
    Input('data-type-dropdown', 'value'),
)
def update_year_selector_visibility(selected_type):
    if selected_type and 'Öngörü Miktarı' in selected_type:
        start_year = datetime.today().year
        end_year = start_year + 1
        opts = [{'label': str(y), 'value': str(y)} for y in range(start_year, end_year + 1)]
        # Veri Tipi altında düzgün: tam genişlik, üstten boşluk
        year_style = {'display': 'block', 'width': '100%', 'marginTop': '6px'}
        wrap_style = {'display': 'block', 'marginTop': '4px'}
        # Varsayılan: tek yıl (kullanıcı isterse 2. yılı da seçer)
        return opts, [str(start_year)], year_style, wrap_style
    return [], None, {'display': 'none'}, {'display': 'none'}


# ── "Kolonları Göster/Gizle" — workcenter / kapasite / malzeme: buton sadece store'u çevirir;
#    hidden_columns, kolonlar oluşunca veya store değişince aşağıdaki senkron callback'lerle ayarlanır.
@app.callback(
    Output('workcenter-col-toggle-state', 'data'),
    Input('toggle-columns-workcenter', 'n_clicks'),
    State('workcenter-col-toggle-state', 'data'),
    prevent_initial_call=True,
)
def toggle_workcenter_columns(n_clicks, show_a_cols):
    if not n_clicks:
        raise PreventUpdate
    cur = show_a_cols if show_a_cols is not None else False
    return not cur


@app.callback(
    Output('workcenter_table', 'hidden_columns'),
    Input('workcenter_table', 'columns'),
    Input('workcenter-col-toggle-state', 'data'),
)
def sync_workcenter_hidden_columns(columns, show_a_cols):
    if not columns:
        return []
    show = show_a_cols if show_a_cols is not None else False
    return [] if show else [c['id'] for c in columns if str(c.get('id', '')).endswith('A')]


@app.callback(
    Output('cap-costcenter-col-toggle-state', 'data'),
    Input('toggle-columns-capacity-costcenter', 'n_clicks'),
    State('cap-costcenter-col-toggle-state', 'data'),
    prevent_initial_call=True,
)
def toggle_cap_costcenter_columns(n_clicks, show_a_cols):
    if not n_clicks:
        raise PreventUpdate
    cur = show_a_cols if show_a_cols is not None else False
    return not cur


@app.callback(
    Output('capasity_table_costcenter', 'hidden_columns'),
    Input('capasity_table_costcenter', 'columns'),
    Input('cap-costcenter-col-toggle-state', 'data'),
)
def sync_cap_costcenter_hidden_columns(columns, show_a_cols):
    if not columns:
        return []
    show = show_a_cols if show_a_cols is not None else False
    return [] if show else [c['id'] for c in columns if str(c.get('id', '')).endswith('A')]


@app.callback(
    Output('cap-workcenter-col-toggle-state', 'data'),
    Input('toggle-columns-capacity-workcenter', 'n_clicks'),
    State('cap-workcenter-col-toggle-state', 'data'),
    prevent_initial_call=True,
)
def toggle_cap_workcenter_columns(n_clicks, show_a_cols):
    if not n_clicks:
        raise PreventUpdate
    cur = show_a_cols if show_a_cols is not None else False
    return not cur


@app.callback(
    Output('capasity_table_workcenter', 'hidden_columns'),
    Input('capasity_table_workcenter', 'columns'),
    Input('cap-workcenter-col-toggle-state', 'data'),
)
def sync_cap_workcenter_hidden_columns(columns, show_a_cols):
    if not columns:
        return []
    show = show_a_cols if show_a_cols is not None else False
    return [] if show else [c['id'] for c in columns if str(c.get('id', '')).endswith('A')]


@app.callback(
    Output('material-col-toggle-state', 'data'),
    Input('toggle-columns-material', 'n_clicks'),
    State('material-col-toggle-state', 'data'),
    prevent_initial_call=True,
)
def toggle_material_columns(n_clicks, show_a_cols):
    if not n_clicks:
        raise PreventUpdate
    cur = show_a_cols if show_a_cols is not None else False
    return not cur


@app.callback(
    Output('material_table', 'hidden_columns'),
    Input('material_table', 'columns'),
    Input('material-col-toggle-state', 'data'),
)
def sync_material_hidden_columns(columns, show_a_cols):
    if not columns:
        return []
    show = show_a_cols if show_a_cols is not None else False
    if show:
        return []
    # MATERIAL ve DRAWNUM kesinlikle gizlenmesin (fixed columns ile birlikte sabit kalsın).
    return [
        c['id']
        for c in columns
        if str(c.get('id', '')).endswith('A') and str(c.get('id', '')) not in ('MATERIAL', 'DRAWNUM')
    ]


@app.callback(
     Output("costcenter-dropdown", "options"),
     Output("costcenter-dropdown", "value"),
    Output('costcenter_table', 'data'),
    Output('costcenter_table', 'columns'),
     Output('costcenter_table', 'style_data_conditional'),
     Input("table_name", "data"),
     Input("filtered_kolon_sum", "data"),
     Input("kolon_sum", "data"),
     State("costcenter-dropdown", "value"),
     State("filtered_kolon_list", "data"),
     State("kolon_list", "data"),
     State('unit-checkbox', 'value'),
     State('costcenter_table', 'data'),
     State('costcenter_table', 'columns'),
     State('costcenter_table', 'style_data_conditional'),
)
def update_graph(table_name, kolon_sum_filtered, kolon_sum_raw, current_costcenter_value,
                 kolon_list_filtered, kolon_list_raw, selected_units,
                 existing_data, existing_columns, existing_style):
    active_cell = None  # active_cell artık bu callback'te yok
    # prefer filtered list, fallback to raw kolon_sum
    kolon_sum = kolon_sum_filtered or kolon_sum_raw
    kolon_list = kolon_list_filtered or kolon_list_raw
    # debug prints retained for now
    print("aaa")
    print(kolon_sum)
    if isinstance(kolon_sum, str):
        kolon_sum = kolon_sum.split(', ')
    if isinstance(kolon_sum, list):
        kolon_sum = ", ".join(kolon_sum)
    print("bbb")
    print(kolon_sum)

    # table_name henüz set edilmemişse (ilk yükleme/race) sorgu atma; "Invalid object name 'None'" önlenir
    if not table_name or not str(table_name).strip():
        raise PreventUpdate

    # Tek sorgu: costcenter tablosu (GROUP BY STAND). Birim dönüşümü SELECT içinde (SUM/60 veya SUM/510).
    kolon_sum_unit = _select_sum_with_unit(kolon_list, selected_units) or kolon_sum
    sql_query = f"SELECT STAND,{kolon_sum_unit} FROM [{table_name}] GROUP BY STAND ORDER BY STAND"
    df_pivot = ag.run_query(sql_query)

    # Guard: if the query returned nothing, return empty outputs
    if df_pivot is None or (hasattr(df_pivot, "empty") and df_pivot.empty):
        options_list = []
        first_option = None
        return options_list, first_option, [], [], []

    sorted_costcenters = sorted(df_pivot["STAND"].astype(str).unique().tolist())
    options_list = [{"label": wc, "value": wc} for wc in sorted_costcenters]
    first_option = options_list[0]["value"] if options_list else None

    df_pivot = _clean_df_column_names(df_pivot)
    # Seçenek B: Saat/Vardiya da dakika gibi int görünsün.
    df_pivot = df_pivot.round(0)

    # Float dtype kalıp DataTable'da ".0" göstermesin diye sayısal kolonları int'e zorla.
    costcenter_numeric_cols = [c for c in df_pivot.columns if c != "STAND"]
    df_pivot[costcenter_numeric_cols] = df_pivot[costcenter_numeric_cols].round(0).astype("Int64")

    # Keep numeric values as numbers so numeric comparisons work
    data = df_pivot.to_dict('records')  # rowData için uygun format
    # Coerce possible formatted numeric strings back to numbers so style comparisons work
    costcenter_numeric_cols = [c for c in df_pivot.columns if c != "STAND"]
    _coerce_numeric_records(data, costcenter_numeric_cols)
    _blank_zero_cells_in_records(data, costcenter_numeric_cols)

   

    
    # İlk kolon başlığı boş kalmasın (köşe hücresi = sabit kolonun başlangıcı; sağa kaydırmada üstte kalması için)
    column_definitions = []
    zero_first_state = {"available": table_name in ("VLFCAPFINALPIVOT", "VLFCAPFINALSIPARIS"), "first_done": False}
    for col in df_pivot.columns:
        name = _display_col_name(col, zero_first_state)
        column_definitions.append({'name': name, 'id': col, 'hideable': False})

    # normalize kolon_list to a list to avoid TypeError if None
    if kolon_list is None:
        kolon_list = []
    elif isinstance(kolon_list, str):
        kolon_list = kolon_list.split(', ')
    # ensure list type
    if not isinstance(kolon_list, list):
        kolon_list = list(kolon_list)

    style_data_conditional = [
        {'if': {'row_index': 'odd'}, 'backgroundColor': '#eef4ff'},
        {'if': {'row_index': 'even'}, 'backgroundColor': '#ffffff'},
    ]
    for col in kolon_list:
        style_data_conditional.append({
            'if': {
                'filter_query': f'{{{col}}} < 0',
                'column_id': col
            },
            'backgroundColor': 'rgb(180, 0, 0)',
            'color': '#ffffff',
            'fontWeight': '700'
        })
        style_data_conditional.append({
            'if': {
                'filter_query': f'{{{col}}} = 0',
                'column_id': col
            },
            'color': '#bbbbbb'
        })
       






    style_data_conditional.extend(_fixed_pin_style_data_overrides(['STAND']))

    trigger_id = ctx.triggered_id if hasattr(ctx, "triggered_id") else None
    if trigger_id == 'table_name':
        new_costcenter_value = first_option
    else:
        # Kullanıcının seçimini koru
        if current_costcenter_value and current_costcenter_value in [o['value'] for o in options_list]:
            new_costcenter_value = no_update
        else:
            new_costcenter_value = first_option
    # Debugging: print small samples so we can inspect what's returned and why styles may not apply
    try:
        print("DEBUG update_graph: sample data rows:", data[:3])
    except Exception:
        print("DEBUG update_graph: cannot print data sample")
    try:
        print("DEBUG update_graph: style_data_conditional:", style_data_conditional[:8])
    except Exception:
        print("DEBUG update_graph: cannot print style sample")

    return options_list, new_costcenter_value, data, column_definitions, style_data_conditional


@app.callback(
    Output('workcenter-capacity-dropdown', 'options'),
    Output('fig', 'figure'),
    Output('workcenter-dropdown', 'options'),
    Output('workcenter-dropdown', 'value'),
    Input('costcenter-dropdown', 'value'),
    Input("filtered_kolon_sum", "data"),
    Input("kolon_sum", "data"),
    State('unit-checkbox', 'value'),
    State("table_name", "data"),
    State("filtered_kolon_graph", "data"),
    State("kolon_graph", "data"),
    State("capacity_table_name", "data"),
)
def update_capacity_and_workcenter_dropdowns(selected_costcenter, kolon_sum_filtered, kolon_sum_raw, selected_units, table_name, kolon_graph_filtered, kolon_graph_raw, capacity_table_name):
    # CAPGRUP + CAPWORK tek sorguda; fig için ikinci sorgu. Böylece 3 yerine 2 sorgu.
    kolon_sum = kolon_sum_filtered or kolon_sum_raw
    kolon_graph = kolon_graph_filtered or kolon_graph_raw

    if not selected_costcenter or not table_name:
        raise PreventUpdate

    empty_fig = go.Figure()
    empty_cap = [{'label': 'Kapasite Grubu', 'value': 'Kapasite Grubu'}]
    empty_wc = [{'label': 'Hepsi', 'value': 'Hepsi'}]

    if kolon_sum is None:
        kolon_sum = []
    elif isinstance(kolon_sum, str):
        kolon_sum = kolon_sum.split(', ')
    if isinstance(kolon_sum, list):
        kolon_sum = ", ".join(kolon_sum)

    # Tek sorgu: hem CAPGRUP hem CAPWORK listesi (DISTINCT CAPGRUP, CAPWORK)
    distinct_sql = f"SELECT DISTINCT CAPGRUP, CAPWORK FROM [{table_name}] WHERE STAND = '{selected_costcenter}'"
    df_distinct = ag.run_query(distinct_sql)
    if df_distinct is None or df_distinct.empty:
        return empty_cap, empty_fig, empty_wc, 'Hepsi'

    capgrups = sorted(df_distinct["CAPGRUP"].astype(str).dropna().unique().tolist()) if "CAPGRUP" in df_distinct.columns else []
    capworks = sorted(df_distinct["CAPWORK"].astype(str).dropna().unique().tolist()) if "CAPWORK" in df_distinct.columns else []
    options_list_cap = [{'label': 'Kapasite Grubu', 'value': 'Kapasite Grubu'}] + [{'label': w, 'value': w} for w in capgrups]
    workcenters = ['Hepsi'] + capworks
    options_workcenter = [{'label': w, 'value': w} for w in workcenters]

    if not kolon_graph:
        return options_list_cap, empty_fig, options_workcenter, 'Hepsi'

    sum_df_fig_sql = (
        f"SELECT A.STAND, {kolon_graph} "
        f"FROM (SELECT STAND, CAPWORK, {kolon_sum} FROM [{table_name}] WHERE STAND = '{selected_costcenter}' GROUP BY STAND, CAPWORK) A "
        f"LEFT JOIN [{capacity_table_name}] B ON A.CAPWORK = B.WORKCENTER "
        f"GROUP BY A.STAND ORDER BY A.STAND"
    )
    sum_df_fig = ag.run_query(sum_df_fig_sql)
    if sum_df_fig is None or (hasattr(sum_df_fig, "empty") and sum_df_fig.empty):
        return options_list_cap, empty_fig, options_workcenter, 'Hepsi'

    pivot_df_fig = pd.melt(sum_df_fig, id_vars=['STAND'], var_name='current_week', value_name='value_min')



    if len(pivot_df_fig) == 0:
        fig = go.Figure()
    else:
        fig = px.bar(pivot_df_fig, x='current_week', y='value_min')
        fig.update_traces(marker=dict(
            color=['#dc2626' if val > 100 else '#1565c0' for val in pivot_df_fig['value_min']]
        ))
        fig.update_xaxes(type='category')
        
        cap_vals = pivot_df_fig['value_min'].copy()
        reasonable_max = cap_vals[cap_vals <= 500].max() if (cap_vals <= 500).any() else 200
        if reasonable_max != reasonable_max:  # NaN kontrolü
            reasonable_max = 200
        y_max = max(200, float(reasonable_max) * 1.1)

        fig.update_layout(
            paper_bgcolor='#ffffff',
            plot_bgcolor='#f8faff',
            font=dict(color='#1a2332', family='Inter, Segoe UI, sans-serif', size=13),
            xaxis=dict(gridcolor='rgba(21,101,192,0.1)', tickfont=dict(color='#1a2332')),
            yaxis=dict(
                gridcolor='rgba(21,101,192,0.1)',
                tickfont=dict(color='#1a2332'),
                range=[0, y_max],
            ),
            margin=dict(l=40, r=60, t=30, b=60),
            autosize=True,
            shapes=[
                dict(
                    type='line',
                    xref='paper', x0=0, x1=1,
                    yref='y', y0=100, y1=100,
                    line=dict(color='#ff0000', width=3, dash='dot'),
                )
            ],
            annotations=[
                dict(
                    xref='paper', x=1.0,
                    yref='y', y=100,
                    text='<b>⚠ 100%</b>',
                    showarrow=False,
                    font=dict(color='#ff0000', size=13, family='Inter, Segoe UI, sans-serif'),
                    bgcolor='rgba(255,255,255,0.85)',
                    bordercolor='#ff0000',
                    borderwidth=1,
                    borderpad=3,
                    xanchor='left',
                )
            ],
        )
    return options_list_cap, fig, options_workcenter, 'Hepsi'




@app.callback(
    Output('figx', 'figure'),
    Output('workcenter_table', 'data'),
    Output('workcenter_table', 'columns'),
    Output('workcenter_table', 'style_data_conditional'),
    Output('capasity_table_workcenter', 'data'),
    Output('capasity_table_workcenter', 'columns'),
    Output('capasity_table_workcenter', 'style_data_conditional'),
    Output('material_table', 'data'),
    Output('material_table', 'columns'),
    Output('material_table', 'style_data_conditional'),
    Input('workcenter-dropdown', 'value'),
    Input('workcenter-capacity-dropdown', 'value'),
    Input('year-selector', 'value'),
    State("filtered_kolon_sum", "data"),
    State("kolon_sum", "data"),
    State("filtered_kolon_list", "data"),
    State("kolon_list", "data"),
    State("filtered_kolon_sumb", "data"),
    State("kolon_sumb", "data"),
    State("filtered_kolon_graph", "data"),
    State("kolon_graph", "data"),
    State('unit-checkbox', 'value'),
    State("table_name", "data"),
    State("capacity_table_name", "data"),
    State("costcenter-dropdown", "value"),
)
def update_workcenter(selected_workcenter, selected_capgrup,
                      selected_year,
                      kolon_sum_filtered, kolon_sum_raw,
                      kolon_list_filtered, kolon_list_raw,
                      kolon_sumb_filtered, kolon_sumb_raw,
                      kolon_graph_filtered, kolon_graph_raw,
                      selected_units,
                      table_name, capacity_table_name, selected_costcenter):
    # selected_year sadece callback tetiklenmesi için; gerçek filtreleme filtered_kolon_* üzerinden yapılıyor.
    _ = selected_year

    kolon_sum   = kolon_sum_filtered   or kolon_sum_raw
    # Workcenter/Yük/Kapasite tablolarında A kolonları da olsun; Kolonları Gizle/Aç butonu bunları gizleyebilsin
    kolon_list  = kolon_list_raw or kolon_list_filtered
    kolon_sumb  = kolon_sumb_filtered  or kolon_sumb_raw
    kolon_graph = kolon_graph_filtered or kolon_graph_raw

    empty_fig = go.Figure()
    empty_fig.update_layout(paper_bgcolor='#ffffff', plot_bgcolor='#f8faff')
    empty_out = (empty_fig, [], [], [], [], [], [], [], [], [])

    if not table_name or not kolon_sum or not selected_costcenter:
        raise PreventUpdate

    # ── normalize ────────────────────────────────────────────────────────────
    if isinstance(kolon_sum, list):
        kolon_sum = ", ".join(kolon_sum)
    if kolon_list is None:
        kolon_list = []
    elif isinstance(kolon_list, str):
        kolon_list = kolon_list.split(', ')
    if not isinstance(kolon_list, list):
        kolon_list = list(kolon_list)
    kolon_sum_str = kolon_sum

    button_id = ctx.triggered[0]['prop_id'].split('.')[0] if ctx.triggered else 'workcenter-dropdown'

    # ── Workcenter ve malzeme: birim dönüşümü SELECT içinde (SUM/60 veya SUM/510) ─
    sum_select = _select_sum_with_unit(kolon_list, selected_units)
    where_base = f"STAND = '{selected_costcenter}'"
    if selected_capgrup and selected_capgrup != 'Kapasite Grubu':
        where_base += f" AND CAPGRUP = '{selected_capgrup}'"
    if selected_workcenter and selected_workcenter != 'Hepsi':
        where_base += f" AND CAPWORK = '{selected_workcenter}'"

    # Fallback: birimli SELECT üretilemezse ham kolon ifadesi kullan (sonra Python'da bölme yapılmaz)
    if not sum_select:
        sum_select = kolon_sum_str if isinstance(kolon_sum_str, str) else ", ".join(kolon_list or [])

    # ── Workcenter tablosu (SELECT CAPWORK, SUM(...)/60 AS [...] GROUP BY CAPWORK) ─
    try:
        wc_sql = f"SELECT CAPWORK, {sum_select} FROM [{table_name}] WHERE {where_base} GROUP BY CAPWORK ORDER BY CAPWORK"
        wc_df = ag.run_query(wc_sql)
        if wc_df is None or (hasattr(wc_df, 'empty') and wc_df.empty):
            wc_data, wc_cols, wc_style = [], [], []
        else:
            # Seçenek B: Saat/Vardiya da dakika gibi int görünsün.
            wc_df = wc_df.round(0)
            # Verimlilik kolonunu ekle (tek kaynak: kapasite_data.get_verimlilik_df)
        verim_df = kapasite_data.get_verimlilik_df(ag)
        if verim_df is not None:
            wc_df = wc_df.merge(verim_df, left_on="CAPWORK", right_on="WORKCENTER", how="left")
            wc_df = wc_df.drop(columns=["WORKCENTER"], errors="ignore")
            wc_df["Verimlilik"] = pd.to_numeric(wc_df["Verimlilik"], errors="coerce")
        else:
            wc_df["Verimlilik"] = None

        # Float kalıp DataTable'da ".0" göstermesin diye kolonları int'e zorlayalım.
        wc_numeric_cols = [c for c in wc_df.columns if c != "CAPWORK"]
        wc_df[wc_numeric_cols] = wc_df[wc_numeric_cols].round(0).astype("Int64")

        # Keep numeric values as numbers so numeric comparisons work
        wc_data = wc_df.to_dict('records')
        wc_numeric_cols = [c for c in wc_df.columns if c != "CAPWORK"]
        _coerce_numeric_records(wc_data, wc_numeric_cols)
        _blank_zero_cells_in_records(wc_data, wc_numeric_cols)
        zero_first_state = {"available": table_name in ("VLFCAPFINALPIVOT", "VLFCAPFINALSIPARIS"), "first_done": False}
        wc_cols = [{'name': _display_col_name(c, zero_first_state), 'id': c, 'hideable': False} for c in wc_df.columns]
        wc_style = [
            {'if': {'row_index': 'odd'},  'backgroundColor': '#eef4ff'},
            {'if': {'row_index': 'even'}, 'backgroundColor': '#ffffff'},
        ]
        for col in kolon_list:
            if col not in wc_df.columns:
                continue
            # Numeric negative check
            wc_style.append({'if': {'filter_query': f'{{{col}}} < 0', 'column_id': col},
                             'backgroundColor': 'rgb(180, 0, 0)', 'color': '#ffffff', 'fontWeight': '700'})
            wc_style.append({'if': {'filter_query': f'{{{col}}} = 0', 'column_id': col}, 'color': '#bbbbbb'})
        wc_style.extend(_fixed_pin_style_data_overrides(['CAPWORK']))
    except Exception as e:
        print(f"wc_table hatası: {e}")
        wc_data, wc_cols, wc_style = [], [], []

   
    try:
        mat_sql = f"SELECT MATERIAL, DRAWNUM, {sum_select} FROM [{table_name}] WHERE {where_base} GROUP BY MATERIAL, DRAWNUM ORDER BY MATERIAL, DRAWNUM"
        mat_df = ag.run_query(mat_sql)
        if mat_df is None or (hasattr(mat_df, 'empty') and mat_df.empty):
            mat_data, mat_cols, mat_style = [], [], []
        else:
            # Seçenek B: Saat/Vardiya da dakika gibi int görünsün.
            mat_df = mat_df.round(0)
            filtered_cols = ['MATERIAL', 'DRAWNUM'] + [c for c in kolon_list if c in mat_df.columns]
            mat_df = mat_df[[c for c in filtered_cols if c in mat_df.columns]]

            # Float kalıp DataTable'da ".0" göstermesin diye kolonları int'e zorlayalım.
            mat_numeric_cols = [c for c in mat_df.columns if c not in ("MATERIAL", "DRAWNUM")]
            if mat_numeric_cols:
                mat_df[mat_numeric_cols] = mat_df[mat_numeric_cols].round(0).astype("Int64")

            mat_data = mat_df.to_dict('records')
            mat_numeric_cols = [c for c in kolon_list if c in mat_df.columns]
            _coerce_numeric_records(mat_data, mat_numeric_cols)
            _blank_zero_cells_in_records(mat_data, mat_numeric_cols)
            zero_first_state = {"available": table_name in ("VLFCAPFINALPIVOT", "VLFCAPFINALSIPARIS"), "first_done": False}
            mat_cols = [{'name': _display_col_name(c, zero_first_state), 'id': c, 'hideable': False} for c in mat_df.columns]
            mat_style = [
                {'if': {'row_index': 'odd'},  'backgroundColor': '#eef4ff'},
                {'if': {'row_index': 'even'}, 'backgroundColor': '#ffffff'},
            ]
            for col in kolon_list:
                if col not in mat_df.columns:
                    continue
                mat_style.append({'if': {'filter_query': f'{{{col}}} < 0', 'column_id': col},
                                  'backgroundColor': 'rgb(180, 0, 0)', 'color': '#ffffff', 'fontWeight': '700'})
                mat_style.append({'if': {'filter_query': f'{{{col}}} = 0', 'column_id': col}, 'color': '#bbbbbb'})
            mat_style.extend(_fixed_pin_style_data_overrides(['MATERIAL', 'DRAWNUM']))
    except Exception as e:
        print(f"material_table hatası: {e}")
        mat_data, mat_cols, mat_style = [], [], []

    # ── Kapasite Süresi (Workcenter) — STAT pivot ─────────────────────────────
    cap_data, cap_cols_def, cap_wc_style = [], [], []
    if capacity_table_name and kolon_sumb and kolon_list:
        try:
            kolon_sumb_str = kolon_sumb if isinstance(kolon_sumb, str) else (", ".join(kolon_sumb) if isinstance(kolon_sumb, list) else None)
            if not kolon_sumb_str:
                kolon_sumb_str = generate_weekly_columns()['format1']
            ihtiyac_select = _select_sum_with_unit(kolon_list, selected_units) or kolon_sum_str

            # Kapasite İhtiyacı (birim SELECT içinde)
            if not selected_workcenter or selected_workcenter == 'Hepsi':
                if not selected_capgrup or selected_capgrup == 'Kapasite Grubu':
                    ihtiyac_sql = f"SELECT STAND,{ihtiyac_select} FROM {table_name} WHERE STAND = '{selected_costcenter}' GROUP BY STAND ORDER BY STAND"
                    cap_work_sql = f"SELECT {kolon_sumb_str} FROM (SELECT STAND,CAPWORK FROM {table_name} GROUP BY CAPWORK,STAND) A LEFT JOIN {capacity_table_name} B ON A.CAPWORK = B.WORKCENTER WHERE A.STAND = '{selected_costcenter}' GROUP BY A.STAND"
                else:
                    ihtiyac_sql = f"SELECT STAND,{ihtiyac_select} FROM {table_name} WHERE STAND = '{selected_costcenter}' AND CAPGRUP = '{selected_capgrup}' GROUP BY STAND ORDER BY STAND"
                    cap_work_sql = f"SELECT {kolon_sumb_str} FROM (SELECT STAND,CAPWORK,CAPGRUP FROM {table_name} GROUP BY CAPWORK,STAND,CAPGRUP) A LEFT JOIN {capacity_table_name} B ON A.CAPWORK = B.WORKCENTER WHERE A.STAND = '{selected_costcenter}' AND A.CAPGRUP = '{selected_capgrup}' GROUP BY A.STAND,A.CAPGRUP"
            else:
                if not selected_capgrup or selected_capgrup == 'Kapasite Grubu':
                    ihtiyac_sql = f"SELECT CAPWORK,{ihtiyac_select} FROM {table_name} WHERE STAND = '{selected_costcenter}' AND CAPWORK = '{selected_workcenter}' GROUP BY CAPWORK ORDER BY CAPWORK"
                    cap_work_sql = f"SELECT {kolon_sumb_str} FROM (SELECT STAND,CAPWORK FROM {table_name} GROUP BY CAPWORK,STAND) A LEFT JOIN {capacity_table_name} B ON A.CAPWORK = B.WORKCENTER WHERE A.STAND = '{selected_costcenter}' AND A.CAPWORK = '{selected_workcenter}' GROUP BY A.STAND,A.CAPWORK"
                else:
                    ihtiyac_sql = f"SELECT CAPWORK,{ihtiyac_select} FROM {table_name} WHERE STAND = '{selected_costcenter}' AND CAPGRUP = '{selected_capgrup}' AND CAPWORK = '{selected_workcenter}' GROUP BY CAPWORK ORDER BY CAPWORK"
                    cap_work_sql = f"SELECT {kolon_sumb_str} FROM (SELECT STAND,CAPWORK,CAPGRUP FROM {table_name} GROUP BY CAPWORK,STAND,CAPGRUP) A LEFT JOIN {capacity_table_name} B ON A.CAPWORK = B.WORKCENTER WHERE A.STAND = '{selected_costcenter}' AND A.CAPGRUP = '{selected_capgrup}' AND A.CAPWORK = '{selected_workcenter}' GROUP BY A.STAND,A.CAPGRUP"

            sum_df_ihtiyac  = ag.run_query(ihtiyac_sql)
            sum_df_cap_work = ag.run_query(cap_work_sql)

            if (sum_df_ihtiyac is not None and not sum_df_ihtiyac.empty and
                    sum_df_cap_work is not None and not sum_df_cap_work.empty):

                sum_df_ihtiyac['STAT'] = 'Kapasite İhtiyacı'
                # Seçenek B: Saat/Vardiya da dakika gibi int görünsün.
                sum_df_ihtiyac = sum_df_ihtiyac.round(0)
                weeks_cap = [col for col in kolon_list if col in sum_df_ihtiyac.columns]
                if not weeks_cap:
                    weeks_cap = [c for c in sum_df_ihtiyac.columns if c not in (sum_df_ihtiyac.columns[0], 'STAT')]
                filtered_ihtiyac = sum_df_ihtiyac[['STAT'] + weeks_cap].copy()

                if 'hours' in (selected_units or []):
                    sum_df_cap_work.iloc[:, 1:] = sum_df_cap_work.iloc[:, 1:] / 60
                elif 'shifts' in (selected_units or []):
                    sum_df_cap_work.iloc[:, 1:] = sum_df_cap_work.iloc[:, 1:] / 510
                sum_df_cap_work = sum_df_cap_work.round(0)

                toplam_row = {'STAT': 'Toplam Kapasite'}
                toplam_row.update(sum_df_cap_work.iloc[0].to_dict())
                cap_df_wc = pd.concat([filtered_ihtiyac, pd.DataFrame([toplam_row])], ignore_index=True)

                # Kapasite Farkı
                fark_row = {'STAT': 'Kapasite Farkı'}
                for col in weeks_cap:
                    try:
                        fark_row[col] = round(
                            float(cap_df_wc.loc[cap_df_wc['STAT'] == 'Toplam Kapasite', col].iloc[0]) -
                            float(cap_df_wc.loc[cap_df_wc['STAT'] == 'Kapasite İhtiyacı', col].iloc[0]),
                            0)
                    except Exception:
                        fark_row[col] = 0
                cap_df_wc = pd.concat([cap_df_wc, pd.DataFrame([fark_row])], ignore_index=True)

                # Kümülatif Toplam
                numeric_cols_wc = [c for c in cap_df_wc.columns if c != 'STAT']
                cumsum_wc = cap_df_wc.loc[cap_df_wc['STAT'] == 'Kapasite Farkı', numeric_cols_wc].cumsum(axis=1)
                cum_row = {'STAT': 'Kümülatif Toplam'}
                cum_row.update(cumsum_wc.iloc[0].to_dict())
                cum_row = {k: round(v, 0) if isinstance(v, (int, float)) else v for k, v in cum_row.items()}
                cap_df_wc = pd.concat([cap_df_wc, pd.DataFrame([cum_row])], ignore_index=True)

                # Doluluk Oranı (%) — sayısal tutulur, > 100 filter_query için
                doluluk_vals = []
                ui_row = cap_df_wc[cap_df_wc['STAT'] == 'Kapasite İhtiyacı'].iloc[0]
                tk_row = cap_df_wc[cap_df_wc['STAT'] == 'Toplam Kapasite'].iloc[0]
                for col in weeks_cap:
                    try:
                        tk_val = float(tk_row[col])
                        ui_val = float(ui_row[col])
                        # Doluluk Oranı(%) dakikadaki gibi 0 ondalık ve int görünsün
                        doluluk_vals.append(round((ui_val / tk_val) * 100, 0) if tk_val != 0 else 0)
                    except Exception:
                        doluluk_vals.append(0)
                doluluk_df = pd.DataFrame([doluluk_vals], columns=weeks_cap)
                doluluk_df['STAT'] = 'Doluluk Oranı(%)'
                # Saat/Vardiyada da dakika ile aynı görünüm (0 ondalık).
                doluluk_df = doluluk_df.round(0)

                
                cap_df_wc = pd.concat([cap_df_wc, doluluk_df], ignore_index=True)
                for col in weeks_cap:
                    cap_df_wc[col] = pd.to_numeric(cap_df_wc[col], errors='coerce')

                # Seçenek B: Saat/Vardiya da dakika gibi int görünsün.
                cap_df_wc[weeks_cap] = cap_df_wc[weeks_cap].round(0).astype("Int64")

                # Doluluk satırı (STAT = "Doluluk Oranı(%)") her zaman 0 ondalık görünsün
                # (saat/vardiya için yapılan round(1) dolulukta ".0" oluşturuyordu).
                dol_mask = cap_df_wc['STAT'] == 'Doluluk Oranı(%)'
                if dol_mask.any():
                    # float kalmasın: DataTable'da ".0" gibi görünmesin diye int tipine çeviriyoruz
                    cap_df_wc.loc[dol_mask, weeks_cap] = (
                        cap_df_wc.loc[dol_mask, weeks_cap]
                        .round(0)
                        .astype("Int64")
                    )

                # Keep capacity numbers numeric so numeric style rules (>=100, <0) work
                cap_data = cap_df_wc.to_dict('records')
                _coerce_numeric_records(cap_data, weeks_cap)
                _blank_zero_cells_in_records(cap_data, weeks_cap)
                zero_first_state = {"available": table_name in ("VLFCAPFINALPIVOT", "VLFCAPFINALSIPARIS"), "first_done": False}
                cap_cols_def = [{'name': 'STAT', 'id': 'STAT', 'hideable': False}] + [{'name': _display_col_name(c, zero_first_state), 'id': c, 'hideable': False} for c in weeks_cap]
                cap_wc_style = [
                    {'if': {'row_index': 'odd'},  'backgroundColor': '#eef4ff'},
                    {'if': {'row_index': 'even'}, 'backgroundColor': '#ffffff'},
                ]
                for col in weeks_cap:
                    
                    cap_wc_style.append({'if': {'filter_query': f'{{{col}}} < 0', 'column_id': col},
                                         'backgroundColor': 'rgb(180, 0, 0)', 'color': '#ffffff', 'fontWeight': '700'})
                    
                    cap_wc_style.append({'if': {'filter_query': f'{{{col}}} = 0', 'column_id': col}, 'color': '#bbbbbb'})
                    
                    cap_wc_style.append({'if': {'filter_query': f'{{STAT}} = "Doluluk Oranı(%)" && {{{col}}} >= 100', 'column_id': col},
                                         'backgroundColor': 'rgb(180, 0, 0)', 'color': '#ffffff', 'fontWeight': '700'})
                cap_wc_style.extend(_fixed_pin_style_data_overrides(['STAT']))
        except Exception as e:
            print(f"capasity_table_workcenter hatası: {e}")
            cap_data, cap_cols_def, cap_wc_style = [], [], []

   
    figx = empty_fig
    if kolon_graph and capacity_table_name:
        try:
            kolon_graph_str = kolon_graph if isinstance(kolon_graph, str) else ", ".join(kolon_graph)

            if not selected_workcenter or selected_workcenter == 'Hepsi':
                if not selected_capgrup or selected_capgrup == 'Kapasite Grubu':
                    figx_sql = (
                        f"SELECT A.STAND, {kolon_graph_str} "
                        f"FROM (SELECT STAND,CAPWORK,{kolon_sum_str} FROM {table_name} WHERE STAND = '{selected_costcenter}' GROUP BY STAND,CAPWORK) A "
                        f"LEFT JOIN {capacity_table_name} B ON A.CAPWORK = B.WORKCENTER "
                        f"GROUP BY A.STAND ORDER BY A.STAND"
                    )
                    id_var = 'STAND'
                else:
                    figx_sql = (
                        f"SELECT A.CAPGRUP, {kolon_graph_str} "
                        f"FROM (SELECT CAPGRUP,CAPWORK,{kolon_sum_str} FROM {table_name} WHERE STAND = '{selected_costcenter}' AND CAPGRUP = '{selected_capgrup}' GROUP BY CAPGRUP,CAPWORK) A "
                        f"LEFT JOIN {capacity_table_name} B ON A.CAPWORK = B.WORKCENTER "
                        f"GROUP BY A.CAPGRUP ORDER BY A.CAPGRUP"
                    )
                    id_var = 'CAPGRUP'
            else:
                if not selected_capgrup or selected_capgrup == 'Kapasite Grubu':
                    figx_sql = (
                        f"SELECT A.CAPWORK, {kolon_graph_str} "
                        f"FROM (SELECT STAND,CAPWORK,{kolon_sum_str} FROM {table_name} WHERE STAND = '{selected_costcenter}' AND CAPWORK = '{selected_workcenter}' GROUP BY STAND,CAPWORK) A "
                        f"LEFT JOIN {capacity_table_name} B ON A.CAPWORK = B.WORKCENTER "
                        f"GROUP BY A.STAND,A.CAPWORK ORDER BY A.CAPWORK"
                    )
                    id_var = 'CAPWORK'
                else:
                    figx_sql = (
                        f"SELECT A.CAPWORK, {kolon_graph_str} "
                        f"FROM (SELECT STAND,CAPGRUP,CAPWORK,{kolon_sum_str} FROM {table_name} WHERE STAND = '{selected_costcenter}' AND CAPGRUP = '{selected_capgrup}' AND CAPWORK = '{selected_workcenter}' GROUP BY STAND,CAPWORK,CAPGRUP) A "
                        f"LEFT JOIN {capacity_table_name} B ON A.CAPWORK = B.WORKCENTER "
                        f"GROUP BY A.STAND,A.CAPGRUP,A.CAPWORK ORDER BY A.CAPWORK"
                    )
                    id_var = 'CAPWORK'

            df_figx = ag.run_query(figx_sql)

            if df_figx is not None and not (hasattr(df_figx, 'empty') and df_figx.empty) and len(df_figx) > 0:
                pivot_figx = pd.melt(df_figx, id_vars=[id_var], var_name='hafta', value_name='deger')
                pivot_figx['deger'] = pd.to_numeric(pivot_figx['deger'], errors='coerce').fillna(0)

                if len(pivot_figx) > 0:
                    # Kırmızı (>100) / Mavi (<=100) renklendirme
                    bar_colors = ['rgb(180, 0, 0)' if v > 100 else '#1565c0' for v in pivot_figx['deger']]
                    figx = px.bar(pivot_figx, x='hafta', y='deger')
                    figx.update_traces(marker=dict(color=bar_colors))
                    figx.update_xaxes(type='category')
                    # Outlier kontrolü: yüzde değerleri 0-500 arası olmalı
                    cap_vals_x = pivot_figx['deger'].copy()
                    reasonable_max_x = cap_vals_x[cap_vals_x <= 500].max() if (cap_vals_x <= 500).any() else 200
                    if reasonable_max_x != reasonable_max_x:
                        reasonable_max_x = 200
                    y_max_x = max(200, float(reasonable_max_x) * 1.1)

                    figx.update_layout(
                        paper_bgcolor='#ffffff',
                        plot_bgcolor='#f8faff',
                        font=dict(color='#1a2332', family='Inter, Segoe UI, sans-serif', size=13),
                        xaxis=dict(gridcolor='rgba(21,101,192,0.1)', tickfont=dict(color='#1a2332')),
                        yaxis=dict(
                            gridcolor='rgba(21,101,192,0.1)',
                            tickfont=dict(color='#1a2332'),
                            range=[0, y_max_x],
                        ),
                        margin=dict(l=40, r=60, t=20, b=40),
                        autosize=True,
                        shapes=[dict(
                            type='line', xref='paper', x0=0, x1=1,
                            yref='y', y0=100, y1=100,
                            line=dict(color='#ff0000', width=3, dash='dot'),
                        )],
                        annotations=[dict(
                            xref='paper', x=1.0, yref='y', y=100,
                            text='<b>⚠ 100%</b>', showarrow=False,
                            font=dict(color='#ff0000', size=13, family='Inter, Segoe UI, sans-serif'),
                            bgcolor='rgba(255,255,255,0.85)', bordercolor='#ff0000',
                            borderwidth=1, borderpad=3, xanchor='left',
                        )],
                    )
        except Exception as e:
            print(f"figx grafik hatası: {e}")
            figx = empty_fig

    # Debugging: print small samples to server log to inspect returned data and styles
    try:
        print("DEBUG update_workcenter: figx traces:", len(figx.data) if figx and hasattr(figx, "data") else 0)
    except Exception:
        pass
    try:
        print("DEBUG update_workcenter: wc_data sample:", wc_data[:2])
    except Exception:
        pass
    try:
        print("DEBUG update_workcenter: wc_style sample:", wc_style[:6])
    except Exception:
        pass
    try:
        print("DEBUG update_workcenter: cap_data sample:", cap_data[:2])
    except Exception:
        pass
    try:
        print("DEBUG update_workcenter: cap_wc_style sample:", cap_wc_style[:6])
    except Exception:
        pass
    try:
        print("DEBUG update_workcenter: mat_data sample:", mat_data[:2])
    except Exception:
        pass

    return figx, wc_data, wc_cols, wc_style, cap_data, cap_cols_def, cap_wc_style, mat_data, mat_cols, mat_style




# ── Download callbacks ───────────────────────────────────────────────────────
@app.callback(
    Output("download-workcenter", "data"),
    Input("btn-download-workcenter", "n_clicks"),
    State('workcenter-dropdown', 'value'),
    State('workcenter-capacity-dropdown', 'value'),
    State("filtered_kolon_sum", "data"),
    State("kolon_sum", "data"),
    State('table_name', 'data'),
    State("filtered_kolon_list", "data"),
    State("kolon_list", "data"),
    State("unit-checkbox", "value"),
    State("capacity_table_name", "data"),
    State('costcenter-dropdown', 'value'),
    prevent_initial_call=True,
)
def download_workcenter(n_clicks, selected_workcenter, selected_capgrup,
                       kolon_sum_filtered, kolon_sum_raw, table_name,
                       kolon_list_filtered, kolon_list_raw, selected_units,
                       capacity_table_name, selected_costcenter):
    # Build same query as update_workcenter to fetch table data
    kolon_sum = kolon_sum_filtered or kolon_sum_raw
    kolon_list = kolon_list_filtered or kolon_list_raw
    if not table_name or not kolon_sum:
        raise PreventUpdate
    if isinstance(kolon_sum, str):
        kolon_sum_parts = kolon_sum.split(', ')
    else:
        kolon_sum_parts = list(kolon_sum) if kolon_sum else []
    kolon_sum_str = ", ".join(kolon_sum_parts)

    # Determine where clause
    if selected_workcenter:
        wc_where = f"WHERE CAPWORK = '{selected_workcenter}'"
    elif selected_costcenter:
        wc_where = f"WHERE STAND = '{selected_costcenter}'"
    else:
        wc_where = ""

    wc_sql = f"SELECT CAPWORK,{kolon_sum_str} FROM {table_name} {wc_where} GROUP BY CAPWORK ORDER BY CAPWORK"
    df_wc = ag.run_query(wc_sql)
    if df_wc is None or (hasattr(df_wc, 'empty') and df_wc.empty):
        raise PreventUpdate
    # format numbers same as update_workcenter
    try:
        df_wc = df_wc.round(0)
    except Exception:
        pass
    # convert to excel download
    try:
        return _export_styled_excel(df_wc, "workcenter_table.xlsx", table_title="Workcenter Yük Tablosu — Makine Bazlı Analiz")
    except Exception as e:
        print("download_workcenter error:", e)
        raise PreventUpdate



# ── Download: Costcenter Tablosu ────────────────────────────────────────────
@app.callback(
    Output("download-costcenter", "data"),
    Input("btn-download-costcenter", "n_clicks"),
    State("filtered_kolon_sum", "data"),
    State("kolon_sum", "data"),
    State('table_name', 'data'),
    State("filtered_kolon_list", "data"),
    State("kolon_list", "data"),
    State("unit-checkbox", "value"),
    prevent_initial_call=True,
)
def download_costcenter(n_clicks, kolon_sum_filtered, kolon_sum_raw, table_name,
                        kolon_list_filtered, kolon_list_raw, selected_units):
    kolon_sum = kolon_sum_filtered or kolon_sum_raw
    kolon_list = kolon_list_filtered or kolon_list_raw
    if not table_name or not kolon_sum:
        raise PreventUpdate
    if isinstance(kolon_list, str):
        kolon_list = [c.strip() for c in kolon_list.split(',') if c.strip()]
    elif not isinstance(kolon_list, list):
        kolon_list = list(kolon_list) if kolon_list else []
    select_clause = _select_sum_with_unit(kolon_list, selected_units)
    if not select_clause:
        select_clause = kolon_sum if isinstance(kolon_sum, str) else ", ".join(kolon_sum or [])
    sql = f"SELECT STAND,{select_clause} FROM [{table_name}] GROUP BY STAND ORDER BY STAND"
    df = ag.run_query(sql)
    if df is None or (hasattr(df, 'empty') and df.empty):
        raise PreventUpdate
    try:
        # Seçenek B: Saat/Vardiya da dakika gibi int görünsün.
        df = df.round(0)
    except Exception:
        pass
    try:
        return _export_styled_excel(df, "costcenter_yuk_tablosu.xlsx", table_title="Costcenter Yük Tablosu — Haftalık Süre Dağılımı")
    except Exception as e:
        print("download_costcenter error:", e)
        raise PreventUpdate


# ── Download: Kapasite Costcenter Tablosu ───────────────────────────────────
# Canlıyla birebir aynı: tablodaki STAT pivot verisini (capasity_table_costcenter.data) indirir
@app.callback(
    Output("download-costcenter_kapasite", "data"),
    Input("btn-download-costcenter_kapasite", "n_clicks"),
    State('capasity_table_costcenter', 'data'),
    prevent_initial_call=True,
)
def download_costcenter_kapasite(n_clicks, table_data):
    if not table_data:
        raise PreventUpdate
    try:
        df = pd.DataFrame(table_data)
        return _export_styled_excel(df, "costcenter_capacity_data.xlsx", table_title="Kapasite Süresi Tablosu — Costcenter Bazlı")
    except Exception as e:
        print("download_costcenter_kapasite error:", e)
        raise PreventUpdate


# ── Download: Workcenter Kapasite Tablosu ───────────────────────────────────
# Canlıyla birebir aynı: tablodaki STAT pivot verisini (capasity_table_workcenter.data) indirir
@app.callback(
    Output("download-workcenter_kapasite", "data"),
    Input("btn-download-workcenter_kapasite", "n_clicks"),
    State('capasity_table_workcenter', 'data'),
    prevent_initial_call=True,
)
def download_workcenter_kapasite(n_clicks, table_data):
    if not table_data:
        raise PreventUpdate
    try:
        df = pd.DataFrame(table_data)
        return _export_styled_excel(df, "workcenter_capacity_data.xlsx", table_title="Kapasite Süresi Tablosu — Workcenter Bazlı")
    except Exception as e:
        print("download_workcenter_kapasite error:", e)
        raise PreventUpdate


# ── Download: Malzeme Tablosu ────────────────────────────────────────────────
@app.callback(
    Output("download-malzeme", "data"),
    Input("btn-download-malzeme", "n_clicks"),
    State("filtered_kolon_sum", "data"),
    State("kolon_sum", "data"),
    State("filtered_kolon_list", "data"),
    State("kolon_list", "data"),
    State('table_name', 'data'),
    State("unit-checkbox", "value"),
    State('costcenter-dropdown', 'value'),
    prevent_initial_call=True,
)
def download_malzeme(n_clicks, kolon_sum_filtered, kolon_sum_raw, kolon_list_filtered, kolon_list_raw,
                     table_name, selected_units, selected_costcenter):
    kolon_sum = kolon_sum_filtered or kolon_sum_raw
    kolon_list = kolon_list_filtered or kolon_list_raw
    if not table_name or not kolon_sum:
        raise PreventUpdate
    if isinstance(kolon_list, str):
        kolon_list = [c.strip() for c in kolon_list.split(',') if c.strip()]
    elif not isinstance(kolon_list, list):
        kolon_list = list(kolon_list) if kolon_list else []
    select_clause = _select_sum_with_unit(kolon_list, selected_units)
    if not select_clause:
        select_clause = kolon_sum if isinstance(kolon_sum, str) else ", ".join(kolon_sum or [])
    where = f"WHERE STAND = '{selected_costcenter}'" if selected_costcenter else ""
    for id_col in ['MATERIAL', 'CAPWORK', 'STAND']:
        try:
            sql = f"SELECT {id_col},{select_clause} FROM [{table_name}] {where} GROUP BY {id_col} ORDER BY {id_col}"
            df = ag.run_query(sql)
            if df is not None and not (hasattr(df, 'empty') and df.empty):
                break
        except Exception:
            df = None
    if df is None or (hasattr(df, 'empty') and df.empty):
        raise PreventUpdate
    try:
        # Seçenek B: Saat/Vardiya da dakika gibi int görünsün.
        df = df.round(0)
    except Exception:
        pass
    try:
        return _export_styled_excel(df, "malzeme_yuk_tablosu.xlsx", table_title="Malzeme Yük Tablosu — Malzeme Bazlı Analiz")
    except Exception as e:
        print("download_malzeme error:", e)
        raise PreventUpdate


## row highlight callbacks moved into the main table callbacks to avoid duplicate Outputs


# ── Accordion → Panel açma/kapama (Python serverside) ───────────────────────
# active_item değiştiğinde ilgili panel'e kap-panel-visible class'ı ekler;
# çarpı (✕) veya kapasite-close-all tıklanınca paneller kapanır.
@app.callback(
    Output("panel-costcenter",  "className"),
    Output("panel-workcenter",  "className"),
    Output("kapasite-accordion", "active_item"),
    Input("kapasite-accordion", "active_item"),
    Input("kapasite-close-all", "n_clicks"),
    Input("close-panel-costcenter", "n_clicks"),
    Input("close-panel-workcenter", "n_clicks"),
)
def toggle_panels(active_item, close_all_clicks, close_cc_clicks, close_wc_clicks):
    """Accordion item'a göre doğru tam ekran paneli göster; çarpıya basınca kapat."""
    triggered = ctx.triggered_id if hasattr(ctx, "triggered_id") else None
    _open  = "kap-fullscreen-panel kap-panel-visible"
    _close = "kap-fullscreen-panel"
    if triggered in ("kapasite-close-all", "close-panel-costcenter", "close-panel-workcenter"):
        return _close, _close, None
    mapping = {
        "acc-item-1": ("panel-costcenter",  0),
        "acc-item-3": ("panel-workcenter",  1),
    }
    classes = [_close, _close]
    if active_item and active_item in mapping:
        _, idx = mapping[active_item]
        classes[idx] = _open
    return classes[0], classes[1], no_update



# ── Kapasite Süresi (Costcenter) tablosunu doldur — STAT pivot yapısı ────────

@app.callback(
    Output('capasity_table_costcenter', 'data'),
    Output('capasity_table_costcenter', 'columns'),
    Output('capasity_table_costcenter', 'style_data_conditional'),
    Input('costcenter-dropdown', 'value'),
    Input('workcenter-capacity-dropdown', 'value'),
    Input('year-selector', 'value'),
    State("filtered_kolon_sum", "data"),
    State("kolon_sum", "data"),
    State("filtered_kolon_list", "data"),
    State("kolon_list", "data"),
    State("filtered_kolon_sumb", "data"),
    State("kolon_sumb", "data"),
    State('unit-checkbox', 'value'),
    State("table_name", "data"),
    State("capacity_table_name", "data"),
)
def update_costcenter_capacity_table(selected_costcenter, selected_capgrp,
                                      selected_year,
                                      kolon_sum_filtered, kolon_sum_raw,
                                      kolon_list_filtered, kolon_list_raw,
                                      kolon_sumb_filtered, kolon_sumb_raw,
                                      selected_units,
                                      table_name, capacity_table_name):
    # selected_year sadece callback tetiklemek için; gerçek içerik filtered_kolon_* ile geliyor.
    _ = selected_year
    # Kapasite Süresi (Costcenter) tablosu kolonları için:
    # "Kolonları Göster/Gizle" butonu hidden_columns üzerinden A kolonlarını açıp kapatsın.
    # Bu yüzden A kolonları filtered state yüzünden daha baştan filtrelenmesin; raw kolonları kullanıyoruz.
    kolon_sum = kolon_sum_raw or kolon_sum_filtered
    kolon_list = kolon_list_raw or kolon_list_filtered
    kolon_sumb = kolon_sumb_raw or kolon_sumb_filtered

    if not table_name or not kolon_sum or not selected_costcenter:
        raise PreventUpdate

    if isinstance(kolon_sum, list):
        kolon_sum = ", ".join(kolon_sum)
    if isinstance(kolon_list, list):
        kolon_list_str = ", ".join(kolon_list)
    else:
        kolon_list_str = kolon_list or ""
    kolon_list = kolon_list_str.split(', ') if kolon_list_str else []

    try:
        ihtiyac_select = _select_sum_with_unit(kolon_list, selected_units) or kolon_sum
        if not selected_capgrp or selected_capgrp == 'Kapasite Grubu':
            ihtiyac_sql  = f"SELECT STAND,{ihtiyac_select} FROM [{table_name}] WHERE STAND = '{selected_costcenter}' GROUP BY STAND ORDER BY STAND"
            cap_work_sql = f"SELECT {kolon_sumb} FROM (SELECT STAND,CAPWORK FROM [{table_name}] GROUP BY CAPWORK,STAND) A LEFT JOIN [{capacity_table_name}] B ON A.CAPWORK = B.WORKCENTER WHERE A.STAND = '{selected_costcenter}' GROUP BY A.STAND"
        else:
            ihtiyac_sql  = f"SELECT STAND,{ihtiyac_select} FROM [{table_name}] WHERE CAPGRUP = '{selected_capgrp}' AND STAND = '{selected_costcenter}' GROUP BY STAND ORDER BY STAND"
            cap_work_sql = f"SELECT {kolon_sumb} FROM (SELECT STAND,CAPWORK,CAPGRUP FROM [{table_name}] GROUP BY CAPWORK,STAND,CAPGRUP) A LEFT JOIN [{capacity_table_name}] B ON A.CAPWORK = B.WORKCENTER WHERE A.STAND = '{selected_costcenter}' AND A.CAPGRUP = '{selected_capgrp}' GROUP BY A.STAND"

        sum_df          = ag.run_query(ihtiyac_sql)
        sum_df_cap_work = ag.run_query(cap_work_sql)

        if sum_df is None or sum_df.empty or sum_df_cap_work is None or sum_df_cap_work.empty:
            return [], [], []

        # Seçenek B: Saat/Vardiya da dakika gibi int görünsün.
        sum_df = sum_df.round(0)
        sum_df['STAT'] = 'Kapasite İhtiyacı'
        weeks = [col for col in kolon_list if col in sum_df.columns]
        if not weeks:
            weeks = [c for c in sum_df.columns if c not in ('STAND', 'STAT')]
        filtered_sum_df = sum_df[['STAT'] + weeks].copy()

        if 'hours' in (selected_units or []):
            sum_df_cap_work.iloc[:, 1:] = sum_df_cap_work.iloc[:, 1:] / 60
        elif 'shifts' in (selected_units or []):
            sum_df_cap_work.iloc[:, 1:] = sum_df_cap_work.iloc[:, 1:] / 510
        sum_df_cap_work = sum_df_cap_work.round(0)

        toplam_row = {'STAT': 'Toplam Kapasite'}
        toplam_row.update(sum_df_cap_work.iloc[0].to_dict())
        cap_df = pd.concat([filtered_sum_df, pd.DataFrame([toplam_row])], ignore_index=True)

        fark_row = {'STAT': 'Kapasite Farkı'}
        for col in weeks:
            try:
                fark_row[col] = round(
                    float(cap_df.loc[cap_df['STAT'] == 'Toplam Kapasite', col].iloc[0]) -
                    float(cap_df.loc[cap_df['STAT'] == 'Kapasite İhtiyacı', col].iloc[0]),
                    0)
            except Exception:
                fark_row[col] = 0
        cap_df = pd.concat([cap_df, pd.DataFrame([fark_row])], ignore_index=True)

        numeric_cols = [c for c in cap_df.columns if c != 'STAT']
        cumsum = cap_df.loc[cap_df['STAT'] == 'Kapasite Farkı', numeric_cols].cumsum(axis=1)
        cum_row = {'STAT': 'Kümülatif Toplam'}
        cum_row.update(cumsum.iloc[0].to_dict())
        cum_row = {k: round(v, 0) if isinstance(v, (int, float)) else v for k, v in cum_row.items()}
        cap_df = pd.concat([cap_df, pd.DataFrame([cum_row])], ignore_index=True)

        # Doluluk Oranı (%) — sayısal tutulur, > 100 filter_query için
        doluluk_vals = []
        ui_row = cap_df[cap_df['STAT'] == 'Kapasite İhtiyacı'].iloc[0]
        tk_row = cap_df[cap_df['STAT'] == 'Toplam Kapasite'].iloc[0]
        for col in weeks:
            try:
                tk = float(tk_row[col])
                ui = float(ui_row[col])
                # Saat/Vardiya sonuçları dakika ile aynı tipte görünsün: 0 ondalık
                doluluk_vals.append(round((ui / tk) * 100, 0) if tk != 0 else 0)
            except Exception:
                doluluk_vals.append(0)
        doluluk_df = pd.DataFrame([doluluk_vals], columns=weeks)
        doluluk_df['STAT'] = 'Doluluk Oranı(%)'
        # Saat/Vardiyada da dakika gibi 0 ondalık
        doluluk_df = doluluk_df.round(0)

        # Prepare final DataFrame and format for display
        # Seçenek B: Saat/Vardiya da dakika gibi int görünsün.
        cap_df = cap_df.round(0)

        cap_df_prepared = pd.concat([cap_df, doluluk_df], ignore_index=True)

        # Float kalıp DataTable'da ".0" göstermesin diye kolonları int'e zorlayalım.
        cap_df_prepared[weeks] = cap_df_prepared[weeks].round(0).astype("Int64")

        # Doluluk satırı her zaman 0 ondalık görünsün
        dol_mask = cap_df_prepared['STAT'] == 'Doluluk Oranı(%)'
        if dol_mask.any():
            cap_df_prepared.loc[dol_mask, weeks] = (
                cap_df_prepared.loc[dol_mask, weeks]
                .round(0)
                .astype("Int64")
            )
        # Convert to records and coerce formatted strings to numeric types
        cap_records = cap_df_prepared.to_dict('records')
        _coerce_numeric_records(cap_records, weeks)
        _blank_zero_cells_in_records(cap_records, weeks)

        zero_first_state = {"available": table_name in ("VLFCAPFINALPIVOT", "VLFCAPFINALSIPARIS"), "first_done": False}
        cols = [{'name': 'STAT', 'id': 'STAT', 'hideable': False}] + [{'name': _display_col_name(c, zero_first_state), 'id': c, 'hideable': False} for c in weeks]

        sdc = [
            {'if': {'row_index': 'odd'},  'backgroundColor': '#eef4ff'},
            {'if': {'row_index': 'even'}, 'backgroundColor': '#ffffff'},
        ]
        for col in weeks:
            # Numeric comparisons: negative values and zero
            sdc.append({'if': {'filter_query': f'{{{col}}} < 0', 'column_id': col},
                        'backgroundColor': '#dc2626', 'color': '#ffffff', 'fontWeight': '700'})
            sdc.append({'if': {'filter_query': f'{{{col}}} = 0', 'column_id': col}, 'color': '#bbbbbb'})
            # Doluluk Oranı > 100 → kırmızı arka plan
            sdc.append({'if': {'filter_query': f'{{STAT}} = "Doluluk Oranı(%)" && {{{col}}} >= 100', 'column_id': col},
                        'backgroundColor': '#dc2626', 'color': '#ffffff', 'fontWeight': '700'})
        sdc.extend(_fixed_pin_style_data_overrides(['STAT']))

        # Debugging: print sample rows and style rules to server log
        try:
            print("DEBUG update_costcenter_capacity_table: cap_df_prepared sample:", cap_df_prepared.head(3).to_dict('records'))
        except Exception:
            pass
        try:
            print("DEBUG update_costcenter_capacity_table: sdc sample:", sdc[:8])
        except Exception:
            pass

        return cap_records, cols, sdc

    except Exception as e:
        print(f"update_costcenter_capacity_table hatası: {e}")
        return [], [], []


# ── Tab geçişi: Costcenter panel içindeki sekmeler ──────────────────────────
@app.callback(
    Output("tab-panel-yuk",     "style"),
    Output("tab-panel-kap",     "style"),
    Output("tab-panel-grafik",  "style"),
    Output("tab-btn-yuk",       "className"),
    Output("tab-btn-kap",       "className"),
    Output("tab-btn-grafik",    "className"),
    Output("active-tab-store",  "data"),
    Input("tab-btn-yuk",     "n_clicks"),
    Input("tab-btn-kap",     "n_clicks"),
    Input("tab-btn-grafik",  "n_clicks"),
    prevent_initial_call=True,
)
def switch_tab(n_yuk, n_kap, n_grafik):
    """Costcenter panel içindeki 3 sekme arasında geçiş yapar (Yük, Kapasite, Grafik)."""
    triggered = ctx.triggered_id if hasattr(ctx, "triggered_id") else None

    tab_map = {
        "tab-btn-yuk":    "yuk",
        "tab-btn-kap":    "kap",
        "tab-btn-grafik": "grafik",
    }
    active = tab_map.get(triggered, "yuk")

    _base = {
        'width': '100%', 'height': '100%',
        'padding': '12px 24px 16px', 'boxSizing': 'border-box',
        'flexDirection': 'column',
        'flex': '1 1 0%', 'minHeight': '0',
    }
    _show_table  = dict(_base, display='flex', overflow='auto')
    _show_grafik = dict(_base, display='flex', overflow='visible')
    _hide        = dict(_base, display='none')

    styles = {
        "yuk":    _show_table  if active == "yuk"    else _hide,
        "kap":    _show_table  if active == "kap"    else _hide,
        "grafik": _show_grafik if active == "grafik" else _hide,
    }

    def _cls(tab_key):
        return "kap-tab-btn kap-tab-active" if active == tab_key else "kap-tab-btn"

    return (
        styles["yuk"], styles["kap"], styles["grafik"],
        _cls("yuk"), _cls("kap"), _cls("grafik"),
        active,
    )

# ── Tab geçişi: Workcenter panel içindeki sekmeler (Grafik, Yük, Kapasite, Malzeme) ─────
@app.callback(
    Output("wc-tab-panel-grafik",   "style"),
    Output("wc-tab-panel-yuk",      "style"),
    Output("wc-tab-panel-kap",      "style"),
    Output("wc-tab-panel-malzeme",  "style"),
    Output("wc-tab-btn-grafik",     "className"),
    Output("wc-tab-btn-yuk",        "className"),
    Output("wc-tab-btn-kap",        "className"),
    Output("wc-tab-btn-malzeme",    "className"),
    Output("wc-active-tab-store",   "data"),
    Input("wc-tab-btn-grafik",   "n_clicks"),
    Input("wc-tab-btn-yuk",      "n_clicks"),
    Input("wc-tab-btn-kap",      "n_clicks"),
    Input("wc-tab-btn-malzeme",  "n_clicks"),
    prevent_initial_call=True,
)
def switch_wc_tab(n_grafik, n_yuk, n_kap, n_malzeme):
    """Workcenter panel içindeki 4 sekme arasında geçiş yapar (Grafik, Yük, Kapasite, Malzeme)."""
    triggered = ctx.triggered_id if hasattr(ctx, "triggered_id") else None

    tab_map = {
        "wc-tab-btn-grafik":  "grafik",
        "wc-tab-btn-yuk":     "yuk",
        "wc-tab-btn-kap":     "kap",
        "wc-tab-btn-malzeme": "malzeme",
    }
    active = tab_map.get(triggered, "grafik")

    _base = {
        'width': '100%', 'height': '100%',
        'padding': '12px 24px 16px', 'boxSizing': 'border-box',
        'flexDirection': 'column',
        'flex': '1 1 0%', 'minHeight': '0',
    }
    _show_grafik = dict(_base, display='flex', overflow='visible')
    _show_table  = dict(_base, display='flex', overflow='auto')
    _hide        = dict(_base, display='none')

    def _cls(tab_key):
        return "kap-tab-btn kap-tab-active" if active == tab_key else "kap-tab-btn"

    return (
        _show_grafik if active == "grafik"  else _hide,
        _show_table  if active == "yuk"     else _hide,
        _show_table  if active == "kap"     else _hide,
        _show_table  if active == "malzeme" else _hide,
        _cls("grafik"), _cls("yuk"), _cls("kap"), _cls("malzeme"),
        active,
    )