# -*- coding: utf-8 -*-

from datetime import datetime, timedelta
from collections import defaultdict
from contextlib import nullcontext
import io
import zipfile
import base64
import re
import os
import sys
import time
import smtplib
import pandas as pd
from email.message import EmailMessage
import config

from dash import html, dcc, Input, Output, no_update
import dash_bootstrap_components as dbc
from dash.exceptions import PreventUpdate

from app import app
import kapasite_data

_RUN_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "run")
if _RUN_DIR not in sys.path:
    sys.path.append(_RUN_DIR)
from agent import ag  # type: ignore[reportMissingImports]
import agent as run_agent  # type: ignore[reportMissingImports]

# Veri tipleri: (etiket, tablo, kapasite_tablo, haftalık/aylık)
DATA_TYPES = [
    ("İhtiyaç Miktarı", "VLFCAPFINALPIVOT", "VLFVARDIYASURE", "weekly"),
    ("Sipariş Miktarı", "VLFCAPFINALSIPARIS", "VLFVARDIYASURE", "weekly"),
    ("Öngörü Miktarı", "VLFCAPFINALOY", "VLFVARDIYASUREAY", "monthly"),
]


RAPORLAMA_SINGLE_COSTCENTER = None


RAPORLAMA_COSTCENTERS = []


RAPORLAMA_NOTIFY_EMAIL = "dayyildiz@valfsan.com.tr"


RAPORLAMA_PROFILE_TIMING = True


def _normalize_recipients(value):
    if value is None:
        return []
    if isinstance(value, str):
        return [x.strip() for x in value.split(",") if x.strip()]
    try:
        return [str(x).strip() for x in value if str(x).strip()]
    except TypeError:
        text = str(value).strip()
        return [text] if text else []


def _get_notify_recipients():
    to_list = _normalize_recipients(getattr(config, "RAPORLAMA_NOTIFY_TO", None))
    cc_list = _normalize_recipients(getattr(config, "RAPORLAMA_NOTIFY_CC", None))
    bcc_list = _normalize_recipients(getattr(config, "RAPORLAMA_NOTIFY_BCC", None))
    if not to_list:
        to_list = _normalize_recipients(RAPORLAMA_NOTIFY_EMAIL)
    return to_list, cc_list, bcc_list


def _rapor_profile_enabled():
    v = os.environ.get("KAP_RAPOR_PROFILE", "").strip().lower()
    if v in ("0", "false", "no", "off"):
        return False
    if v in ("1", "true", "yes", "on"):
        return True
    return RAPORLAMA_PROFILE_TIMING


def _sql_tables_guess(sql):
    
    if not sql:
        return []
    return re.findall(r"(?:FROM|JOIN)\s+\[([^\]]+)\]", str(sql), flags=re.I)


class _RunQueryProfiler:
    

    def __init__(self, state):
        self.state = state
        self._mod = None
        self._orig_fn = None
        self._orig_agent_rq_fn = None

    def __enter__(self):
        import agent as ra  # type: ignore[reportMissingImports]

        
        if not isinstance(ra.Agent.__dict__.get("run_query"), staticmethod):
            ra.Agent.run_query = staticmethod(ra.run_query)

        self._mod = ra
        self._orig_fn = ra.run_query
        
        orig_agent_attr = ra.Agent.__dict__.get("run_query")
        if isinstance(orig_agent_attr, staticmethod):
            self._orig_agent_rq_fn = orig_agent_attr.__func__
        else:
            self._orig_agent_rq_fn = ra.Agent.run_query
        st = self.state

        def _wrapped(sql):
            t0 = time.perf_counter()
            try:
                return self._orig_fn(sql)
            finally:
                dt = time.perf_counter() - t0
                st["sql_sec"] += dt
                st["sql_calls"] += 1
                tables = _sql_tables_guess(sql)
                key = "+".join(sorted(set(tables))) if tables else "(sorgu)"
                st["by_table"][key] += dt
                preview = (sql or "").replace("\n", " ").strip()
                if len(preview) > 220:
                    preview = preview[:217] + "..."
                st["slow_queries"].append((dt, preview))

        ra.run_query = _wrapped
        ra.Agent.run_query = staticmethod(_wrapped)
        return self

    def __exit__(self, *args):
        if self._mod is not None:
            self._mod.run_query = self._orig_fn
            # Geri yüklemede staticmethod olarak sar, yoksa instance call'da 2 argüman hatası verir.
            self._mod.Agent.run_query = staticmethod(self._orig_agent_rq_fn)
        return False


def _print_rapor_timing_report(zip_wall_sec, prof, per_excel_rows):
    """Konsola özet rapor."""
    lines = [
        "",
        "=" * 72,
        "[Raporlama] SÜRE ÖZETİ (ZIP)",
        "=" * 72,
        f"  ZIP toplam (duvar saati):     {zip_wall_sec:,.2f} s",
        f"  SQL toplam (run_query):       {prof['sql_sec']:,.2f} s  ({prof['sql_calls']} çağrı)",
        f"  ZIP - SQL (Excel+pandas+zip): {max(0.0, zip_wall_sec - prof['sql_sec']):,.2f} s",
        "",
        "  En çok süren SQL tablo grupları (FROM/JOIN [..] özeti, üst 12):",
    ]
    by_t = sorted(prof["by_table"].items(), key=lambda x: -x[1])[:12]
    if not by_t:
        lines.append("    (veri yok)")
    for name, sec in by_t:
        lines.append(f"    {sec:8.2f} s  {name}")
    lines.append("")
    lines.append("  En yavaş tek sorgular (üst 15):")
    slow = sorted(prof["slow_queries"], key=lambda x: -x[0])[:15]
    if not slow:
        lines.append("    (veri yok)")
    for sec, prev in slow:
        lines.append(f"    {sec:8.2f} s  {prev}")
    lines.append("")
    lines.append("  Her Excel dosyası:")
    for row in per_excel_rows:
        lines.append(
            f"    • {row['label']}: duvar {row['wall_sec']:.2f} s | "
            f"SQL {row['sql_sec']:.2f} s | Excel gövde ~{row['non_sql_sec']:.2f} s | "
            f"save {row.get('save_sec', 0):.2f} s | CC={row['cc_count']}"
        )
        seg = row.get("segments") or {}
        if seg:
            lines.append(
                "        aşamalar: "
                + " | ".join(f"{k}={v:.2f}s" for k, v in sorted(seg.items(), key=lambda x: -x[1])[:8])
            )
        cc_top = row.get("cc_detail_top")
        if cc_top:
            lines.append("        en yavaş CC detay sayfası (duvar):")
            for cc, sec in cc_top:
                lines.append(f"          {sec:7.2f} s  {cc}")
    lines.append("=" * 72)
    try:
        print("\n".join(lines))
    except UnicodeEncodeError:
        print("\n".join(lines).encode("ascii", "replace").decode("ascii"))


def _send_report_notification_email(status, detail, created_reports=None, attachment_bytes=None, attachment_name="kapasite_raporlar.zip"):
    
    created_reports = created_reports or []
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    subject = f"Kapasite Raporlama - {status}"
    body = (
        "Kapasite raporlama islemi tamamlandi.\n\n"
        f"Durum: {status}\n"
        f"Tarih: {now_str}\n"
        f"Detay: {detail}\n"
        f"Raporlar: {', '.join(created_reports) if created_reports else '-'}\n"
    )

    to_list, cc_list, bcc_list = _get_notify_recipients()
    recipient_text = ", ".join(to_list + cc_list + bcc_list) if (to_list or cc_list or bcc_list) else "-"

    # 1) SMTP (config.py içindeki MAIL_MODE ayarına göre)
    host = str(getattr(config, "mail_server", "") or "").strip()
    port = int(getattr(config, "mail_port", 25) or 25)
    use_tls = bool(getattr(config, "SMTP_USE_STARTTLS", False))
    require_auth = bool(getattr(config, "SMTP_REQUIRE_AUTH", False))
    user = str(getattr(config, "SMTP_USERNAME", "") or "").strip()
    password = str(getattr(config, "SMTP_PASSWORD", "") or "")
    sender = user or "no-reply@localhost"

    if host:

        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = sender
        if to_list:
            msg["To"] = ", ".join(to_list)
        if cc_list:
            msg["Cc"] = ", ".join(cc_list)
        if bcc_list:
            msg["Bcc"] = ", ".join(bcc_list)
        msg.set_content(body)
        if attachment_bytes:
            msg.add_attachment(
                attachment_bytes,
                maintype="application",
                subtype="zip",
                filename=attachment_name,
            )
        try:
            with smtplib.SMTP(host, port, timeout=12) as server:
                if use_tls:
                    server.starttls()
                if require_auth and user:
                    server.login(user, password)
                server.send_message(msg)
            return True, f"E-posta bildirimi gonderildi ({recipient_text}) [SMTP]."
        except Exception as smtp_err:
            smtp_msg = str(smtp_err)
            if "5.7.139" in smtp_msg or "SmtpClientAuthentication is disabled" in smtp_msg:
                return (
                    False,
                    "SMTP reddedildi: Microsoft 365 tarafinda SMTP AUTH kapali (5.7.139). "
                    "IT yoneticisi tenant veya posta kutusu icin SMTP AUTH acmadan sifresiz gonderim mumkun degil.",
                )
    else:
        smtp_msg = "SMTP ayari bulunamadi (config.mail_server yok)."

    # 2) Outlook fallback (opsiyonel)
    try:
        import win32com.client  # type: ignore
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
        except Exception:
            from win32com.client import gencache  # type: ignore
            outlook = gencache.EnsureDispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        if to_list:
            mail.To = ";".join(to_list)
        if cc_list:
            mail.CC = ";".join(cc_list)
        mail.Subject = subject
        mail.Body = body
        if attachment_bytes:
            import tempfile
            with tempfile.NamedTemporaryFile(delete=False, suffix=".zip") as tmp:
                tmp.write(attachment_bytes)
                tmp_path = tmp.name
            try:
                mail.Attachments.Add(tmp_path)
                mail.Send()
            finally:
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass
        else:
            mail.Send()
        return True, f"E-posta bildirimi gonderildi ({recipient_text}) [Outlook]."
    except Exception as outlook_err:
        return (
            False,
            "E-posta gonderilemedi. "
            f"SMTP: {smtp_msg} | Outlook: {outlook_err}",
        )



def _build_styles():
    """openpyxl stil nesnelerini tek sözlükte döndürür."""
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    thin  = Side(style="thin",   color="B0BEC5")
    med   = Side(style="medium", color="37474F")

    def _border(l=thin, r=thin, t=thin, b=thin):
        return Border(left=l, right=r, top=t, bottom=b)

    s = {}

   
    s["cc_title_font"]  = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    s["cc_title_fill"]  = PatternFill("solid", fgColor="0D2B5E")
    s["cc_title_align"] = Alignment(horizontal="left", vertical="center", indent=1)
    s["cc_title_border"]= _border(l=med, r=med, t=med, b=med)

    
    s["sec_font"]  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    s["sec_fill"]  = PatternFill("solid", fgColor="1565C0")
    s["sec_align"] = Alignment(horizontal="left", vertical="center", indent=1)
    s["sec_border"]= _border(l=med, r=med, t=med, b=med)

   
    s["hdr_font"]  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    s["hdr_fill"]  = PatternFill("solid", fgColor="1976D2")
    s["hdr_align"] = Alignment(horizontal="center", vertical="center", wrap_text=True)
    s["hdr_border"]= _border(l=thin, r=thin, t=med, b=med)

    
    s["date_hdr_font"]  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    s["date_hdr_fill"]  = PatternFill("solid", fgColor="00695C")
    s["date_hdr_align"] = Alignment(horizontal="center", vertical="center", wrap_text=True)
    s["date_hdr_border"]= _border(l=thin, r=thin, t=med, b=med)

   
    s["first_col_font"]  = Font(name="Arial", bold=True, color="0D2B5E", size=9)
    s["first_col_fill"]  = PatternFill("solid", fgColor="E3F2FD")
    s["first_col_align"] = Alignment(horizontal="left", vertical="center", indent=1)
    s["first_col_border"]= _border(l=thin, r=med, t=thin, b=thin)

    
    s["data_font"]  = Font(name="Arial", size=9, color="212121")
    s["data_align"] = Alignment(horizontal="right", vertical="center")
    s["data_border"]= _border()

    
    s["zebra_fill"] = PatternFill("solid", fgColor="F5F5F5")

    
    s["doluluk_font"] = Font(name="Arial", bold=True, color="1A237E", size=9)
    s["doluluk_fill"] = PatternFill("solid", fgColor="E8EAF6")

    
    s["red_fill"] = PatternFill("solid", fgColor="DC2626")
    s["red_font"] = Font(name="Arial", bold=True, color="FFFFFF", size=9)

    
    s["sorunlu_cc_fill"] = PatternFill("solid", fgColor="E8EAF6")
    s["sorunlu_cc_font"] = Font(name="Arial", size=10, color="1A237E")
    s["sorunlu_deger_fill"] = PatternFill("solid", fgColor="B71C1C")
    s["sorunlu_deger_font"] = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    s["sorunlu_kap_fark_fill"] = PatternFill("solid", fgColor="FFE0B2")
    s["sorunlu_kum_fill"] = PatternFill("solid", fgColor="FFF9C4")
    s["sorunlu_doluluk_fill"] = PatternFill("solid", fgColor="F8BBD9")
    s["sorunlu_section_fill"] = PatternFill("solid", fgColor="37474F")
    s["sorunlu_section_font"] = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    s["sorunlu_cc_big_font"] = Font(name="Arial", bold=True, color="FFFFFF", size=15)
    s["sorunlu_cc_big_fill"] = PatternFill("solid", fgColor="0D47A1")
    s["sorunlu_sheet_title_font"] = Font(name="Arial", bold=True, color="FFFFFF", size=16)
    s["sorunlu_sheet_title_fill"] = PatternFill("solid", fgColor="B71C1C")
    s["sorunlu_hdr_font"] = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    s["sorunlu_hdr_fill"] = PatternFill("solid", fgColor="455A64")
    s["sorunlu_empty_font"] = Font(name="Arial", italic=True, color="78909C", size=10)

    
    cc_header_colors = ["0D47A1", "1565C0", "00695C", "2E7D32", "4A148C", "BF360C", "E65100", "1B5E20"]
    s["cc_title_fills"] = [PatternFill("solid", fgColor=c) for c in cc_header_colors]
    section_colors = ["37474F", "455A64", "546E7A", "607D8B", "78909C", "5C6BC0", "00838F", "6A1B9A"]
    s["sorunlu_section_fills"] = [PatternFill("solid", fgColor=c) for c in section_colors]
    s["sec_fills"] = [PatternFill("solid", fgColor=c) for c in section_colors]

    # Kapalı bloklarda görünen mini özet: sol başlıktan sakin; kart zemin; vurgu sadece oran satırı / son hücre
    thin_line = Side(style="thin", color="90A4AE")
    s["mini_card_fill"] = PatternFill("solid", fgColor="F2F5F8")
    s["mini_card_border_side"] = Side(style="thin", color="78909C")
    s["mini_card_border"] = Border(
        left=s["mini_card_border_side"],
        right=s["mini_card_border_side"],
        top=s["mini_card_border_side"],
        bottom=s["mini_card_border_side"],
    )
    s["mini_gutter_border"] = Border(right=Side(style="thin", color="B0BEC5"))
    s["mini_title_font"] = Font(name="Arial", bold=True, color="546E7A", size=9)
    s["mini_title_fill"] = PatternFill("solid", fgColor="E8ECEF")
    s["mini_hdr_font"] = Font(name="Arial", bold=True, color="455A64", size=9)
    s["mini_hdr_fill"] = PatternFill("solid", fgColor="DCDFE3")
    s["mini_hdr_border"] = Border(
        left=thin_line, right=thin_line, top=thin_line, bottom=thin_line
    )
    s["mini_label_font"] = Font(name="Arial", bold=True, color="37474F", size=9)
    s["mini_label_fill"] = PatternFill("solid", fgColor="ECEFF1")
    s["mini_data_font"] = Font(name="Arial", size=9, color="263238")
    s["mini_ratio_font"] = Font(name="Arial", bold=True, color="1A237E", size=9)
    s["mini_ratio_row_fill"] = PatternFill("solid", fgColor="FFF9C4")
    s["mini_ratio_row_label_font"] = Font(name="Arial", bold=True, color="5D4037", size=9)
    s["mini_ratio_row_value_font"] = Font(name="Arial", bold=True, color="4E342E", size=9)
    s["mini_ratio_last_cell_fill"] = PatternFill("solid", fgColor="6A1B9A")
    s["mini_ratio_last_cell_font"] = Font(name="Arial", bold=True, color="FFFFFF", size=10)

    s["detail_anchor_font"] = Font(name="Arial", bold=True, color="0D47A1", size=10)
    s["detail_anchor_fill"] = PatternFill("solid", fgColor="E1F5FE")

    return s


def _sanitize_sheet_name(name, max_len=31):
    """Excel sheet adı: \ / * ? : [ ] kullanılamaz; en fazla 31 karakter."""
    if not name:
        return "Sheet"
    s = str(name).strip()
    for c in r'\/:*?[]':
        s = s.replace(c, "_")
    return s[:max_len] if len(s) > max_len else s


def _unique_sheet_name(base_name, existing_names, max_len=31):
    """Mevcut adlarla çakışmayan sheet adı üretir (gerekirse sonuna _2, _3 ekler)."""
    base = _sanitize_sheet_name(base_name, max_len)
    if base not in existing_names:
        return base
    existing = set(existing_names)
    for i in range(2, 100):
        candidate = f"{base[:max_len - len(str(i)) - 1]}_{i}"
        if candidate not in existing:
            return candidate
    return base + "_1"


def _is_date_col(col_name):
    """Kolon adı tarih/hafta/ay biçimindeyse True."""
    c = str(col_name).strip()
    if re.match(r"^\d{4}[-/W]\d{2}", c):
        return True
    if re.match(r"^W\d{1,2}$", c, re.IGNORECASE):
        return True
    if c == "0":
        return True
    return False


def _auto_col_widths(ws, first_col_width=30, date_col_width=10, other_col_width=14, max_w=40):
    
    from openpyxl.utils import get_column_letter
    for i, col_cells in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(i)
        if i == 1:
            ws.column_dimensions[col_letter].width = first_col_width
            continue
        header_val = str(col_cells[0].value or "")
        if _is_date_col(header_val):
            ws.column_dimensions[col_letter].width = date_col_width
        else:
            max_len = max((len(str(cell.value or "")) for cell in col_cells), default=0)
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, other_col_width), max_w)



def _reorder_malzeme_columns_for_report(df, week_cols=None):
    
    if df is None or df.empty:
        return df
    end_cols = [c for c in ["MACHINE", "BASEQUAN", "MTUNIT"] if c in df.columns]
    if not end_cols:
        return df
    # MATERIAL → MACHINE hemen önüne; diğer kolonlar (DRAWNUM, hafta vb.) solda kalsın
    other = [c for c in df.columns if c not in ("MACHINE", "BASEQUAN", "MTUNIT", "MATERIAL")]
    material_col = ["MATERIAL"] if "MATERIAL" in df.columns else []
    return df[other + material_col + end_cols]


def _add_cap_alignment_column(cap_df):
    
    return cap_df


def _apply_export_columns(df, hidden_columns=None, table_columns=None, table_name=None):
    
    if df is None or df.empty:
        return df
    df = df.copy()
    hidden_set = set(hidden_columns) if hidden_columns is not None else set()
    
    if hidden_columns is None and not hidden_set:
        hidden_set = {c for c in df.columns if isinstance(c, str) and len(c) >= 6 and c.endswith("A")}
    if hidden_set:
        df = df[[c for c in df.columns if c not in hidden_set]]
    if table_columns:
        rename_map = {c.get("id"): c.get("name") for c in table_columns if c.get("id") and c.get("name") and c["id"] in df.columns}
        if rename_map:
            df = df.rename(columns=rename_map)
    elif table_name in ("VLFCAPFINALPIVOT", "VLFCAPFINALSIPARIS"):
        for c in df.columns:
            if re.match(r"^\d{4}-\d{2}$", str(c)):
                df = df.rename(columns={c: "0"})
                break
    return df


def _format_df_turkish_thousands(df):
    
    if df is None or df.empty:
        return df
    df = df.copy()
    for col in df.columns[1:]:
        if col == "Verimlilik":
            continue
        def _fmt(x):
            if x is None or (isinstance(x, float) and pd.isna(x)):
                return x
            try:
                return f"{float(x):,.0f}".replace(",", ".")
            except Exception:
                return x
        df[col] = df[col].apply(_fmt)
    return df


def _parse_cell_to_float(val):
    
    if val is None:
        return None
    if isinstance(val, bool):
        return None
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    try:
        if pd.api.types.is_number(val):
            return float(val)
    except (TypeError, ValueError):
        pass
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def _dash_like_capacity_excel_number(num_val):
    """Kapasite sayısı: dakika=0, saat=1, vardiya=2; A kolonları her zaman int, 0 -> boş."""
    if num_val is None:
        return None
    try:
        if pd.isna(num_val):
            return None
    except (TypeError, ValueError):
        pass
    x = float(num_val)
    if pd.isna(x):
        return None
    iv = int(round(x, 0))
    return None if iv == 0 else iv


def _unit_decimals(selected_units):
    if selected_units and "hours" in selected_units:
        return 1
    if selected_units and "shifts" in selected_units:
        return 2
    return 0


def _format_capacity_number(num_val, col_name, selected_units):
    if num_val is None:
        return None
    try:
        if pd.isna(num_val):
            return None
    except (TypeError, ValueError):
        pass
    x = float(num_val)
    if pd.isna(x):
        return None
    su = selected_units or []
    if str(col_name).strip().endswith("A"):
        y = round(x, 0)
        if y == 0:
            return None
        return int(y)
    if "hours" in su or "shifts" in su:
        if x == 0:
            return None
        return round(x, kapasite_data.DISPLAY_MAX_DECIMALS_NON_A)
    y = round(x, 0)
    if y == 0:
        return None
    return int(y)


def _excel_number_format_for_capacity_col(col_name, selected_units):
    
    if str(col_name).strip().endswith("A"):
        return "0"
    su = selected_units or []
    if "hours" in su or "shifts" in su:
        return "0." + ("#" * kapasite_data.DISPLAY_MAX_DECIMALS_NON_A)
    return "0"


def _excel_safe_value(val):
    """openpyxl uyumu: pandas NA/NaN -> None."""
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    return val


def _display_blank_if_zero(val):
    """Rapor hücresinde 0 yerine boş değer göster."""
    num = _parse_cell_to_float(val)
    if num is not None and num == 0:
        return None
    return val


def _write_sorunlu_sheet_title(ws, row, s, num_cols=9):
    
    from openpyxl.styles import Alignment
    cell = ws.cell(row=row, column=1, value="Sorunlu Kapasite — Detay için başlığa tıklayın")
    cell.font = s["sorunlu_sheet_title_font"]
    cell.fill = s["sorunlu_sheet_title_fill"]
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = s["sec_border"]
    if num_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    ws.row_dimensions[row].height = 32
    return row + 3


def _write_sorunlu_section_title(ws, row, title, s, section_index=0, num_cols=1, start_col=1):
    
    from openpyxl.styles import Alignment
    cell = ws.cell(row=row, column=start_col, value=title)
    cell.font = s["sorunlu_section_font"]
    fills = s.get("sorunlu_section_fills", [s["sorunlu_section_fill"]])
    cell.fill = fills[section_index % len(fills)]
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=2, wrap_text=True)
    cell.border = s["sec_border"]
    if num_cols > 1:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + num_cols - 1)
    ws.row_dimensions[row].height = 26
    return row + 2


def _write_sorunlu_cc_big_title(ws, row, costcenter_name, s, cc_index=0, num_cols=1):
    
    from openpyxl.styles import Alignment
    cell = ws.cell(row=row, column=1, value=costcenter_name)
    cell.font = s["sorunlu_cc_big_font"]
    fills = s.get("cc_title_fills", [s["sorunlu_cc_big_fill"]])
    cell.fill = fills[cc_index % len(fills)]
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    cell.border = s["sec_border"]
    if num_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    ws.row_dimensions[row].height = 30
    return row + 2



SORUNLU_DISPLAY_COLS = ["Costcenter", "Satır Tipi", "Hafta/Ay", "Değer"]


def _write_sorunlu_table_to_sheet(ws, start_row, df, s, start_col=1):
    
    from openpyxl.styles import Alignment
    if df is None or df.empty:
        return start_row + 1
    cols = SORUNLU_DISPLAY_COLS
    
    for c_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=start_row, column=start_col + c_idx - 1, value=col_name)
        cell.font = s.get("sorunlu_hdr_font", s["hdr_font"])
        cell.fill = s.get("sorunlu_hdr_fill", s["hdr_fill"])
        cell.alignment = s["hdr_align"]
        cell.border = s["hdr_border"]
    ws.row_dimensions[start_row].height = 24
    row = start_row + 1
 
    for _, r in df.iterrows():
        satir_tipi = str(r.get("Satır Tipi", "") or "").strip()
        if satir_tipi == "Kapasite Farkı":
            tip_fill = s["sorunlu_kap_fark_fill"]
        elif satir_tipi == "Kümülatif Toplam":
            tip_fill = s["sorunlu_kum_fill"]
        elif satir_tipi == "Doluluk Oranı(%)":
            tip_fill = s["sorunlu_doluluk_fill"]
        else:
            tip_fill = s["zebra_fill"]
        for c_idx, col_name in enumerate(cols, start=1):
            val = r.get(col_name)
            if c_idx >= 3:
                val = _display_blank_if_zero(val)
            cell = ws.cell(row=row, column=start_col + c_idx - 1, value=_excel_safe_value(val))
            cell.alignment = Alignment(horizontal="right" if c_idx >= 3 else "left", vertical="center")
            cell.border = s["data_border"]
            if c_idx == 1:
                cell.font = s["sorunlu_cc_font"]
                cell.fill = tip_fill
            elif c_idx == 2:
                cell.font = s["data_font"]
                cell.fill = tip_fill
            elif c_idx == 4:
                cell.font = s["sorunlu_deger_font"]
                cell.fill = s["sorunlu_deger_fill"]
            else:
                cell.font = s["data_font"]
                cell.fill = s["zebra_fill"]
        ws.row_dimensions[row].height = 20
        row += 1
    return row + 2


def _write_sorunlu_empty_cell(ws, row, col, s):
    """Sorunlu Kapasite'de 'Sorunlu nokta yok.' mesajı — italic gri."""
    cell = ws.cell(row=row, column=col, value="Sorunlu nokta yok.")
    cell.font = s.get("sorunlu_empty_font", s["data_font"])
    cell.border = s["data_border"]


def _merge_section_title_row(ws, row, s, end_col=3):
    """Hepsi / Workcenter satırı: A–(end_col) birleştir, sağa mini tablo için yer açılır."""
    from openpyxl.styles import Alignment
    cell = ws.cell(row=row, column=1)
    if cell.value is None:
        return
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell.alignment = s["sec_align"]


def _write_detail_block_anchor_row(ws, row, text, s, merge_end_col=3):
    
    from openpyxl.styles import Alignment
    c = ws.cell(row=row, column=1, value=text)
    c.font = s["detail_anchor_font"]
    c.fill = s["detail_anchor_fill"]
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2, wrap_text=True)
    c.border = s["data_border"]
    if merge_end_col > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_end_col)
    ws.row_dimensions[row].height = 22


def _extract_sorunlu_from_cap_df(cap_df, costcenter, workcenter="Tümü"):
    
    if cap_df is None or cap_df.empty or "STAT" not in cap_df.columns:
        return []
    
    week_cols = [c for c in cap_df.columns if c != "STAT" and str(c).strip() != "0" and _is_date_col(c)]
    out = []
    for _, row in cap_df.iterrows():
        stat_str = str(row.get("STAT") or row.iloc[0]).strip()
        if stat_str != "Kümülatif Toplam":
            continue
        for col in week_cols:
            val = row.get(col)
            num = _parse_cell_to_float(val)
            if num is None:
                continue
            if num < 0:
                out.append({
                    "Costcenter": costcenter,
                    "Workcenter": workcenter,
                    "Satır Tipi": stat_str,
                    "Hafta/Ay": col,
                    "Değer": int(round(float(num), 0)),
                })
    return out


def _extract_kumulatif_row_from_cap_df(cap_df, costcenter, workcenter="Tümü", selected_units=None):
    
    if cap_df is None or cap_df.empty or "STAT" not in cap_df.columns:
        return None
    
    week_cols = [c for c in cap_df.columns if c != "STAT"]
    for _, row in cap_df.iterrows():
        stat_str = str(row.get("STAT") or "").strip()
        if stat_str != "Kümülatif Toplam":
            continue
        rec = {"Costcenter": costcenter, "Workcenter": workcenter, "STAT": stat_str}
        for col in week_cols:
            val = row.get(col)
            num = _parse_cell_to_float(val)
            if num is not None:
                rec[col] = _format_capacity_number(num, col, selected_units)
            else:
                rec[col] = val
        return rec
    return None


def _build_cumulative_ratio_from_cap_df(cap_df):
    
    if cap_df is None or cap_df.empty or "STAT" not in cap_df.columns:
        return None, None, None, None

    week_cols = [c for c in cap_df.columns if c != "STAT" and _is_date_col(c) and str(c).strip() != "0"]
    if not week_cols:
        return None, None, None, None

    row_need = None
    row_total = None
    for _, r in cap_df.iterrows():
        stat = str(r.get("STAT") or "").strip()
        if stat == "Kapasite İhtiyacı":
            row_need = r
        elif stat == "Toplam Kapasite":
            row_total = r

    if row_need is None or row_total is None:
        return None, None, None, None

    cum_need = []
    cum_total = []
    cum_ratio = []
    # "0" kolonu olan veri tiplerinde, bu değer sadece Küm. Kapasite İhtiyacı
    # başlangıç birikimine dahil edilir (görünür dönem başlıkları yine week_cols kalır).
    base_need = 0.0
    if "0" in cap_df.columns:
        n0 = _parse_cell_to_float(row_need.get("0"))
        base_need = n0 if n0 is not None else 0.0

    running_need = base_need
    running_total = 0.0
    for col in week_cols:
        n_need = _parse_cell_to_float(row_need.get(col))
        n_total = _parse_cell_to_float(row_total.get(col))
        running_need += (n_need if n_need is not None else 0.0)
        running_total += (n_total if n_total is not None else 0.0)
        cum_need.append(running_need)
        cum_total.append(running_total)
        cum_ratio.append((running_need / running_total * 100.0) if running_total else None)

    return week_cols, cum_need, cum_total, cum_ratio


def _prefix_col_count_before_weeks(cap_df):
    
    if cap_df is None or cap_df.empty:
        return 1
    for i, c in enumerate(cap_df.columns):
        cs = str(c).strip() if c is not None else ""
        if cs != "STAT" and _is_date_col(c) and cs != "0":
            return max(1, i)
    return 1


def _apply_rect_outer_border(ws, r1, r2, c1, c2, outer_side):
    
    from openpyxl.styles import Border

    for r in range(r1, r2 + 1):
        for cc in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=cc)
            prev = cell.border
            cell.border = Border(
                left=outer_side if cc == c1 else prev.left,
                right=outer_side if cc == c2 else prev.right,
                top=outer_side if r == r1 else prev.top,
                bottom=outer_side if r == r2 else prev.bottom,
            )


def _write_cumulative_mini_table(
    ws,
    header_row,
    cap_df,
    s,
    title_text,
    selected_units=None,
    merge_end_col=3,
    right_edge_col=18,
):
    
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    cols, cum_need, cum_total, cum_ratio = _build_cumulative_ratio_from_cap_df(cap_df)
    if not cols:
        return header_row + 2

    prefix_cols = _prefix_col_count_before_weeks(cap_df)

    
    week_positions = {}
    for idx, c in enumerate(cap_df.columns, start=1):
        cs = str(c).strip() if c is not None else ""
        if cs != "STAT" and _is_date_col(c) and cs != "0":
            week_positions[cs] = idx

    
    if len(week_positions) != len(cols):
        first_data_col = 1 + prefix_cols
        data_positions = [first_data_col + i for i in range(len(cols))]
    else:
        data_positions = [week_positions[str(c).strip()] for c in cols]
        first_data_col = min(data_positions)

    label_col = max(1, first_data_col - prefix_cols)
    last_data_col = max(data_positions) if data_positions else first_data_col

    
    start_row = header_row + 1

    def _outline0(r):
        try:
            ws.row_dimensions[r].outlineLevel = 0
        except Exception:
            pass

    # Başlık
    t_cell = ws.cell(row=start_row, column=label_col, value=title_text)
    t_cell.font = s["mini_title_font"]
    t_cell.fill = s["mini_title_fill"]
    t_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    t_cell.border = s["mini_card_border"]
    ws.merge_cells(start_row=start_row, start_column=label_col, end_row=start_row, end_column=last_data_col)
    ws.row_dimensions[start_row].height = 20
    _outline0(start_row)

    # Ay başlıkları (düşük kontrast şerit)
    h_row = start_row + 1
    h_label = ws.cell(row=h_row, column=label_col, value="Dönem")
    h_label.font = s["mini_hdr_font"]
    h_label.fill = s["mini_hdr_fill"]
    h_label.alignment = Alignment(horizontal="center", vertical="center")
    h_label.border = s["mini_hdr_border"]
    if prefix_cols > 1:
        ws.merge_cells(
            start_row=h_row,
            start_column=label_col,
            end_row=h_row,
            end_column=label_col + prefix_cols - 1,
        )
    for i, c in enumerate(cols):
        hc = ws.cell(row=h_row, column=data_positions[i], value=str(c))
        hc.font = s["mini_hdr_font"]
        hc.fill = s["mini_hdr_fill"]
        hc.alignment = Alignment(horizontal="center", vertical="center")
        hc.border = s["mini_hdr_border"]
    ws.row_dimensions[h_row].height = 19
    _outline0(h_row)

    rows_def = [
        ("Küm. Kapasite İhtiyacı", cum_need, False),
        ("Küm. Toplam Kapasite", cum_total, False),
        ("Küm. Oran (%)", cum_ratio, True),
    ]
    cur = h_row + 1
    for label, values, is_ratio in rows_def:
        lc = ws.cell(row=cur, column=label_col, value=label)
        if is_ratio:
            lc.font = s["mini_ratio_row_label_font"]
            lc.fill = s["mini_ratio_row_fill"]
        else:
            lc.font = s["mini_label_font"]
            lc.fill = s["mini_card_fill"]
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lc.border = s["data_border"]
        if prefix_cols > 1:
            ws.merge_cells(
                start_row=cur,
                start_column=label_col,
                end_row=cur,
                end_column=label_col + prefix_cols - 1,
            )
        for i, v in enumerate(values):
            cc = ws.cell(row=cur, column=data_positions[i])
            cc.border = s["data_border"]
            _rv = None
            if v is None:
                cc.value = None
            elif is_ratio:
                _rv = _parse_cell_to_float(v)
                if _rv is None:
                    cc.value = None
                else:
                    p_dec = _unit_decimals(selected_units)
                    p_val = round(_rv, p_dec)
                    cc.value = f"%{int(p_val)}%" if p_dec == 0 else f"%{p_val:.{p_dec}f}%"
            else:
                _nv = _parse_cell_to_float(v)
                cc.value = _format_capacity_number(_nv, cols[i], selected_units)
                if cc.value is not None:
                    cc.number_format = _excel_number_format_for_capacity_col(cols[i], selected_units)
            cc.alignment = Alignment(horizontal="right", vertical="center")
            if is_ratio:
                last = i == len(values) - 1
                if last:
                    cc.fill = s["mini_ratio_last_cell_fill"]
                    cc.font = s["mini_ratio_last_cell_font"]
                elif _rv is not None and _rv > 100:
                    cc.fill = s["red_fill"]
                    cc.font = s["red_font"]
                else:
                    cc.fill = s["mini_ratio_row_fill"]
                    cc.font = s["mini_ratio_row_value_font"]
            else:
                cc.fill = s["mini_card_fill"]
                cc.font = s["mini_data_font"]
        ws.row_dimensions[cur].height = 17
        _outline0(cur)
        cur += 1

    end_row = cur - 1
    _apply_rect_outer_border(ws, start_row, end_row, label_col, last_data_col, s["mini_card_border_side"])

    for pc in range(prefix_cols):
        label_letter = get_column_letter(label_col + pc)
        ld = ws.column_dimensions.get(label_letter)
        lw = ld.width if ld is not None and ld.width is not None else 0
        min_w = 24.0 if pc == 0 else 10.0
        ws.column_dimensions[label_letter].width = max(lw, min_w) if lw else min_w
    for ccix in range(first_data_col, last_data_col + 1):
        L = get_column_letter(ccix)
        d = ws.column_dimensions.get(L)
        w = d.width if d is not None and d.width is not None else 0
        ws.column_dimensions[L].width = max(w, 10.0) if w else 10.0

    return cur + 1


def _write_kumulatif_table_to_sheet(ws, start_row, df, s, selected_units=None):
    
    from openpyxl.styles import Alignment, PatternFill
    _white_fill = PatternFill("solid", fgColor="FFFFFF")
    if df is None or df.empty:
        return start_row + 1
    cols = list(df.columns)
    # Başlık satırı
    for c_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=start_row, column=c_idx, value=col_name)
        cell.font = s.get("sorunlu_hdr_font", s["hdr_font"])
        cell.fill = s.get("sorunlu_hdr_fill", s["hdr_fill"])
        cell.alignment = s["hdr_align"]
        cell.border = s["hdr_border"]
    ws.row_dimensions[start_row].height = 24
    row = start_row + 1
    for row_idx, r in enumerate(df.to_dict("records")):
        for c_idx, col_name in enumerate(cols, start=1):
            val = r.get(col_name)
            cell = ws.cell(row=row, column=c_idx, value=_excel_safe_value(val))
            cell.border = s["data_border"]
            if col_name in ("Workcenter", "Costcenter"):
                cell.font = s["data_font"]
                cell.fill = s["sorunlu_kum_fill"]
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                num_val = _parse_cell_to_float(val)
                if num_val is not None:
                    cell.value = _format_capacity_number(num_val, col_name, selected_units)
                    cell.number_format = _excel_number_format_for_capacity_col(col_name, selected_units)
                if num_val is not None and num_val < 0:
                    cell.fill = s["sorunlu_deger_fill"]
                    cell.font = s["sorunlu_deger_font"]
                else:
                    cell.font = s["data_font"]
                    cell.fill = s["zebra_fill"] if (row_idx % 2 == 1) else _white_fill
        ws.row_dimensions[row].height = 20
        row += 1
    return row + 2



def _write_table_to_sheet(ws, start_row, df, s, title=None, is_section_title=False, cc_index=None, section_index=None, selected_units=None):
    
    def _title_fill():
        if is_section_title and cc_index is not None and "cc_title_fills" in s:
            return s["cc_title_fills"][cc_index % len(s["cc_title_fills"])]
        if not is_section_title and section_index is not None and "sec_fills" in s:
            return s["sec_fills"][section_index % len(s["sec_fills"])]
        return s["cc_title_fill"] if is_section_title else s["sec_fill"]

    if df is None or df.empty:
        if title:
            row = start_row
            cell = ws.cell(row=row, column=1, value=title)
            cell.font   = s["cc_title_font"]  if is_section_title else s["sec_font"]
            cell.fill   = _title_fill()
            cell.alignment = s["cc_title_align"] if is_section_title else s["sec_align"]
            cell.border = s["cc_title_border"] if is_section_title else s["sec_border"]
            ws.row_dimensions[row].height = 22
            return row + 2
        return start_row + 1

    
    df = df.copy()
    is_capacity_table = len(df.columns) > 0 and df.columns[0] == "STAT"
    num_cols = len(df.columns)
    row = start_row

    if title:
        cell = ws.cell(row=row, column=1, value=title)
        cell.font      = s["cc_title_font"]  if is_section_title else s["sec_font"]
        cell.fill      = _title_fill()
        cell.alignment = s["cc_title_align"] if is_section_title else s["sec_align"]
        cell.border    = s["cc_title_border"] if is_section_title else s["sec_border"]
        if num_cols > 1:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        ws.row_dimensions[row].height = 22
        row += 1

   
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        if _is_date_col(col_name) and col_idx > 1:
            cell.font      = s["date_hdr_font"]
            cell.fill      = s["date_hdr_fill"]
            cell.alignment = s["date_hdr_align"]
            cell.border    = s["date_hdr_border"]
        else:
            cell.font      = s["hdr_font"]
            cell.fill      = s["hdr_fill"]
            cell.alignment = s["hdr_align"]
            cell.border    = s["hdr_border"]
    ws.row_dimensions[row].height = 32  
    row += 1

   
    for row_idx, r in enumerate(df.itertuples(index=False)):
        first_col_val  = r[0] if len(r) else None
        is_doluluk_row = str(first_col_val).strip() == "Doluluk Oranı(%)"
        is_even_row    = (row_idx % 2 == 1)

        for c_idx, val in enumerate(r, start=1):
            num_val = None
            
            write_val = val
            if is_capacity_table and c_idx > 1:
                num_val = _parse_cell_to_float(val)
                if num_val is not None:
                     
                    write_val = None if num_val == 0 else num_val
                else:
                    write_val = _excel_safe_value(val)
            elif c_idx > 1:
                col_name = df.columns[c_idx - 1]
                if _is_date_col(col_name):
                    num_val = _parse_cell_to_float(val)
                    write_val = (
                        _format_capacity_number(num_val, col_name, selected_units)
                        if num_val is not None
                        else None
                    )
                else:
                    write_val = _display_blank_if_zero(val)
            cell = ws.cell(row=row, column=c_idx, value=_excel_safe_value(write_val))
            if c_idx > 1 and write_val is not None:
                col_name = df.columns[c_idx - 1]
                if is_capacity_table or _is_date_col(col_name):
                    cell.number_format = _excel_number_format_for_capacity_col(col_name, selected_units)

            # Temel stil
            if c_idx == 1:
                cell.font      = s["first_col_font"]
                cell.fill      = s["first_col_fill"]
                cell.alignment = s["first_col_align"]
                cell.border    = s["first_col_border"]
            elif is_doluluk_row:
                cell.font      = s["doluluk_font"]
                cell.fill      = s["doluluk_fill"]
                cell.alignment = s["data_align"]
                cell.border    = s["data_border"]
            elif is_even_row:
                cell.font      = s["data_font"]
                cell.fill      = s["zebra_fill"]
                cell.alignment = s["data_align"]
                cell.border    = s["data_border"]
            else:
                cell.font      = s["data_font"]
                cell.alignment = s["data_align"]
                cell.border    = s["data_border"]

            
            if is_capacity_table and c_idx > 1 and num_val is not None and (
                    num_val < 0 or (is_doluluk_row and num_val > 100)):
                cell.fill = s["red_fill"]
                cell.font = s["red_font"]

        ws.row_dimensions[row].height = 17
        row += 1

    return row + 2  



def _build_rapor_excel(costcenters, table_name, capacity_table_name, columns_dict, selected_units,
                       hidden_columns=None, table_columns=None, excel_profile=None):
    
    if not costcenters or ag is None:
        return None
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return None

    s  = _build_styles()
    wb = Workbook()

    kolon_sum  = columns_dict["format2"]
    kolon_sumb = columns_dict["format1"]
    week_cols  = kapasite_data.get_kolon_list_from_format3(columns_dict.get("format3", ""))
    selected_units_key = tuple(selected_units or [])

    
    cap_cc_cache = {}
    cap_wc_cache = {}
    wc_list_cache = {}

    def _get_cap_cc_cached(cc):
        key = (
            table_name,
            capacity_table_name,
            cc,
            selected_units_key,
            kolon_sum,
            kolon_sumb,
            tuple(week_cols),
        )
        if key not in cap_cc_cache:
            cap_cc_cache[key] = kapasite_data.build_capacity_table_for_cc(
                ag, table_name, capacity_table_name, cc,
                kolon_sum, kolon_sumb, week_cols, selected_units
            )
        cached_df = cap_cc_cache.get(key)
        if cached_df is None:
            return None
        
        return cached_df.copy(deep=True)

    def _get_workcenters_cached(cc):
        key = (table_name, cc)
        if key not in wc_list_cache:
            wc_list_cache[key] = kapasite_data.get_workcenters_for_cc(ag, table_name, cc) or []
        
        return list(wc_list_cache[key])

    def _get_cap_wc_cached(cc, wc):
        key = (
            table_name,
            capacity_table_name,
            cc,
            wc,
            selected_units_key,
            kolon_sum,
            kolon_sumb,
            tuple(week_cols),
        )
        if key not in cap_wc_cache:
            cap_wc_cache[key] = kapasite_data.build_capacity_table_for_cc_workcenter(
                ag, table_name, capacity_table_name, cc, wc,
                kolon_sum, kolon_sumb, week_cols, selected_units
            )
        cached_df = cap_wc_cache.get(key)
        if cached_df is None:
            return None
        return cached_df.copy(deep=True)

    seg = cc_detail = None
    tmark = time.perf_counter()
    if excel_profile is not None:
        excel_profile.setdefault("segments", defaultdict(float))
        excel_profile.setdefault("cc_detail_sec", defaultdict(float))
        seg = excel_profile["segments"]
        cc_detail = excel_profile["cc_detail_sec"]

    # ─────────────────────────────────────────────────────────────
    # Sheet 1: 1. Accordion — Cost Center Kapasite Süresi
    # ─────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Costcenter Kapasite"
    ws1.sheet_properties.tabColor = "1565C0"
    row1 = 1
    
    kumulatif_rows = []
    for idx, cc in enumerate(costcenters):
        cap_df = _get_cap_cc_cached(cc)
        cap_df = _apply_export_columns(cap_df, hidden_columns, table_columns, table_name)
        rec = _extract_kumulatif_row_from_cap_df(cap_df, cc, workcenter="Tümü", selected_units=selected_units)
        if rec:
            kumulatif_rows.append(rec)
        row1 = _write_table_to_sheet(
            ws1, row1, cap_df, s,
            title=f"◈  {cc}  —  1. Accordion Kapasite Süresi",
            is_section_title=True,
            cc_index=idx,
            selected_units=selected_units,
        )

    if row1 == 1:
        ws1.cell(row=1, column=1, value="Cost center bulunamadı veya veri yok.")

    _auto_col_widths(ws1)

    if seg is not None:
        seg["sheet1_costcenter_kapasite"] += time.perf_counter() - tmark
        tmark = time.perf_counter()

    
    for cc in costcenters:
        workcenters = _get_workcenters_cached(cc)
        for wc in workcenters:
            cap_wc_df = _get_cap_wc_cached(cc, wc)
            cap_wc_df = _apply_export_columns(cap_wc_df, hidden_columns, table_columns, table_name)
            rec = _extract_kumulatif_row_from_cap_df(cap_wc_df, cc, workcenter=wc, selected_units=selected_units)
            if rec:
                kumulatif_rows.append(rec)

    if seg is not None:
        seg["kumulatif_makine_satirlari"] += time.perf_counter() - tmark
        tmark = time.perf_counter()

    
    used_for_sheet = [ws1.title]
    cc_to_sheet = {}
    for cc in costcenters:
        sn = _unique_sheet_name(cc, used_for_sheet)
        cc_to_sheet[cc] = sn
        used_for_sheet.append(sn)

    ws_sorunlu = wb.create_sheet("Sorunlu Kapasite", 1)
    ws_sorunlu.sheet_properties.tabColor = "B71C1C"
    row_sorunlu = 1
    num_sorunlu_cols = 9  # başlık merge genişliği

    row_sorunlu = _write_sorunlu_sheet_title(ws_sorunlu, 1, s, num_cols=num_sorunlu_cols)

    # İki tablo: 1) Tümü olanlar — ilk kolon Costcenter adı (Azot vb.); 2) Makine bazlı — 0 kolonu dahil, sadece kırmızı (negatif) olan satırlar.
    sorunlu_hyperlink_cells = []
    if kumulatif_rows:
        try:
            week_cols_order = [c for c in kumulatif_rows[0].keys() if c not in ("Costcenter", "Workcenter", "STAT")]

            tumu_rows = [r for r in kumulatif_rows if r.get("Workcenter") == "Tümü"]
            diger_rows = [r for r in kumulatif_rows if r.get("Workcenter") != "Tümü"]

            
            def _row_has_negative_except_zero(rec):
                for col in week_cols_order:
                    if col == "0":
                        continue
                    v = _parse_cell_to_float(rec.get(col))
                    if v is not None and v < 0:
                        return True
                return False
            diger_rows = [r for r in diger_rows if _row_has_negative_except_zero(r)]

            display_cols_tumu = ["Costcenter"] + week_cols_order   
            display_cols_diger = ["Workcenter"] + week_cols_order   

            
            row_sorunlu = _write_sorunlu_section_title(
                ws_sorunlu, row_sorunlu, "1. Kümülatif Toplam — Tümü (Costcenter toplamları)", s,
                section_index=0, num_cols=len(display_cols_tumu), start_col=1
            )
            if tumu_rows:
                df_tumu = pd.DataFrame(tumu_rows)[display_cols_tumu]
                data_start_1 = row_sorunlu + 1
                row_sorunlu = _write_kumulatif_table_to_sheet(ws_sorunlu, row_sorunlu, df_tumu, s, selected_units=selected_units)
                for i, r in enumerate(tumu_rows):
                    sorunlu_hyperlink_cells.append((data_start_1 + i, 1, r.get("Costcenter"), r.get("Workcenter")))
            else:
                _write_sorunlu_empty_cell(ws_sorunlu, row_sorunlu, 1, s)
                row_sorunlu += 4

            
            row_sorunlu = _write_sorunlu_section_title(
                ws_sorunlu, row_sorunlu, "2. Kümülatif Toplam — Makine bazlı (sadece negatif içeren satırlar)", s,
                section_index=1, num_cols=len(display_cols_diger), start_col=1
            )
            if diger_rows:
                df_diger = pd.DataFrame(diger_rows)[display_cols_diger]
                data_start_2 = row_sorunlu + 1
                row_sorunlu = _write_kumulatif_table_to_sheet(ws_sorunlu, row_sorunlu, df_diger, s, selected_units=selected_units)
                for i, r in enumerate(diger_rows):
                    sorunlu_hyperlink_cells.append((data_start_2 + i, 1, r.get("Costcenter"), r.get("Workcenter")))
            else:
                _write_sorunlu_empty_cell(ws_sorunlu, row_sorunlu, 1, s)
        except Exception as ex:
            _write_sorunlu_empty_cell(ws_sorunlu, row_sorunlu, 1, s)
            print(f"[Raporlama] Sorunlu Kapasite tablosu yazılırken hata: {ex}")
    else:
        _write_sorunlu_empty_cell(ws_sorunlu, row_sorunlu, 1, s)

    _auto_col_widths(ws_sorunlu, first_col_width=22, other_col_width=14, date_col_width=12)

    if seg is not None:
        seg["sorunlu_sayfa"] += time.perf_counter() - tmark
        tmark = time.perf_counter()

    cc_section_rows = {}
    used_sheet_names = [ws1.title, ws_sorunlu.title]
    for idx, cc in enumerate(costcenters):
        t_cc0 = time.perf_counter()
        sheet_name = _unique_sheet_name(cc, used_sheet_names)
        used_sheet_names.append(sheet_name)
        ws_cc = wb.create_sheet(sheet_name, 2 + idx)
        ws_cc.sheet_properties.tabColor = "00695C"
        row_cc = 1
        cc_section_rows[sheet_name] = {}

        # CC başlığı (her costcenter farklı renk)
        row_cc = _write_table_to_sheet(
            ws_cc, row_cc, pd.DataFrame(), s,
            title=f"◈  {cc}",
            is_section_title=True,
            cc_index=idx,
            selected_units=selected_units,
        )
        # 1) Blok "Hepsi" — tek grup: başlık (Hepsi) + altında Kapasite + Malzeme. Açılışta kapalı, + ile açılır. Outline sembolleri (1,2,3,4) kapalı.
        row_hepsi = row_cc
        row_cc = _write_table_to_sheet(
            ws_cc, row_cc, pd.DataFrame(), s,
            title="  Hepsi",
            is_section_title=False,
            section_index=0,
            selected_units=selected_units,
        )
        cc_section_rows[sheet_name]["Hepsi"] = row_hepsi
        _merge_section_title_row(ws_cc, row_hepsi, s, end_col=3)
        cap_hepsi = _get_cap_cc_cached(cc)
        cap_hepsi = _add_cap_alignment_column(cap_hepsi)
        cap_hepsi = _apply_export_columns(cap_hepsi, hidden_columns, table_columns, table_name)
        # Başlık satırının sağında: D gutter, özet kartı sağa yaslı; Kapasite bundan SONRA yazılır.
        row_cc = _write_cumulative_mini_table(
            ws_cc,
            header_row=row_hepsi,
            cap_df=cap_hepsi,
            s=s,
            title_text="Kümülatif doluluk özeti (Hepsi)",
            selected_units=selected_units,
        )
        detail_start = row_cc
        _write_detail_block_anchor_row(
            ws_cc, detail_start, "  Hepsi — Kapasite Süresi · Malzeme", s, merge_end_col=3
        )
        row_cc = detail_start + 1
        row_cc = _write_table_to_sheet(
            ws_cc, row_cc, cap_hepsi, s,
            title="    Kapasite Süresi",
            is_section_title=False,
            selected_units=selected_units,
        )
        malz_hepsi = kapasite_data.build_malzeme_table_for_cc(
            ag,
            table_name,
            cc,
            kolon_sum,
            selected_units,
            for_report=True,
            kolon_list=week_cols,
        )
        malz_hepsi = _reorder_malzeme_columns_for_report(malz_hepsi, week_cols)
        malz_hepsi = _apply_export_columns(malz_hepsi, hidden_columns, table_columns, table_name)
        row_cc = _write_table_to_sheet(
            ws_cc, row_cc, malz_hepsi, s,
            title="    Malzeme Tablosu",
            is_section_title=False,
            selected_units=selected_units,
        )
        # Başlık satırı özet satırıdır; summaryBelow=False ile + başlığın yanında görünür.
        ws_cc.row_dimensions[row_hepsi].outlineLevel = 0
        ws_cc.row_dimensions[row_hepsi].collapsed = True
        for r in range(row_hepsi + 1, row_cc):
            ws_cc.row_dimensions[r].outlineLevel = 1
            ws_cc.row_dimensions[r].hidden = True

        # 2) Her workcenter için ayrı blok: tek grup (WC başlığı + Kapasite + Malzeme). Açılışta kapalı, + ile açılır.
        workcenters = _get_workcenters_cached(cc)
        for wc_idx, wc in enumerate(workcenters):
            row_wc = row_cc
            row_cc = _write_table_to_sheet(
                ws_cc, row_cc, pd.DataFrame(), s,
                title=f"  {wc}",
                is_section_title=False,
                section_index=wc_idx + 1,
                selected_units=selected_units,
            )
            cc_section_rows[sheet_name][wc] = row_wc
            _merge_section_title_row(ws_cc, row_wc, s, end_col=3)
            cap_wc = _get_cap_wc_cached(cc, wc)
            cap_wc = _add_cap_alignment_column(cap_wc)
            cap_wc = _apply_export_columns(cap_wc, hidden_columns, table_columns, table_name)
            row_cc = _write_cumulative_mini_table(
                ws_cc,
                header_row=row_wc,
                cap_df=cap_wc,
                s=s,
                title_text=f"Kümülatif doluluk özeti — {wc}",
                selected_units=selected_units,
            )
            detail_start = row_cc
            _write_detail_block_anchor_row(
                ws_cc, detail_start, f"  {wc} — Kapasite Süresi · Malzeme", s, merge_end_col=3
            )
            row_cc = detail_start + 1
            row_cc = _write_table_to_sheet(
                ws_cc, row_cc, cap_wc, s,
                title="    Kapasite Süresi",
                is_section_title=False,
                selected_units=selected_units,
            )
            malz_wc = kapasite_data.build_malzeme_table_for_cc_workcenter(
                ag,
                table_name,
                cc,
                wc,
                kolon_sum,
                selected_units,
                for_report=True,
                kolon_list=week_cols,
            )
            malz_wc = _reorder_malzeme_columns_for_report(malz_wc, week_cols)
            malz_wc = _apply_export_columns(malz_wc, hidden_columns, table_columns, table_name)
            row_cc = _write_table_to_sheet(
                ws_cc, row_cc, malz_wc, s,
                title="    Malzeme Tablosu",
                is_section_title=False,
                selected_units=selected_units,
            )
            # Başlık satırı özet satırıdır; summaryBelow=False ile + başlığın yanında görünür.
            ws_cc.row_dimensions[row_wc].outlineLevel = 0
            ws_cc.row_dimensions[row_wc].collapsed = True
            for r in range(row_wc + 1, row_cc):
                ws_cc.row_dimensions[r].outlineLevel = 1
                ws_cc.row_dimensions[r].hidden = True

        # Grup sembollerini göster: satırlar başlangıçta kapalı gelir, kullanıcı + / - ile açıp kapatır.
        ws_cc.sheet_view.showOutlineSymbols = True
        try:
            # Bazı Excel sürümlerinde sadece sheet_view yetmez; outlinePr da açık olmalı.
            ws_cc.sheet_properties.outlinePr.showOutlineSymbols = True
            # Özet satırı üstte: + / - işareti bölüm başlığı satırına hizalanır.
            ws_cc.sheet_properties.outlinePr.summaryBelow = False
            ws_cc.sheet_properties.outlinePr.summaryRight = True
            ws_cc.sheet_properties.outlinePr.applyStyles = True
        except Exception:
            pass
        _auto_col_widths(ws_cc)

        if cc_detail is not None:
            cc_detail[cc] += time.perf_counter() - t_cc0

    if seg is not None:
        seg["cc_detay_sayfalari"] += time.perf_counter() - tmark
        tmark = time.perf_counter()

    # Sorunlu Kapasite sayfasındaki Costcenter hücresine tıklanınca ilgili detay sayfasına git
    from openpyxl.styles import Font
    for (row, col, cc, workcenter) in sorunlu_hyperlink_cells:
        sheet_name = cc_to_sheet.get(cc)
        if not sheet_name:
            continue
        target_row = 1
        section_key = "Hepsi" if (workcenter == "Tümü" or not workcenter) else workcenter
        if sheet_name in cc_section_rows:
            target_row = cc_section_rows[sheet_name].get(section_key, 1)
        # Excel dahili link: #'Sayfa Adı'!A1
        safe_name = sheet_name.replace("'", "''")
        cell = ws_sorunlu.cell(row=row, column=col)
        cell.hyperlink = f"#'{safe_name}'!A{target_row}"
        old_font = cell.font
        cell.font = Font(
            name=old_font.name if old_font else "Arial",
            bold=old_font.bold if old_font else False,
            size=old_font.size if old_font else 11,
            color="0563C1",
            underline="single",
        )

    if seg is not None:
        seg["hyperlinkler"] += time.perf_counter() - tmark
        tmark = time.perf_counter()

    buffer = io.BytesIO()
    t_save0 = time.perf_counter()
    wb.save(buffer)
    if seg is not None:
        seg["workbook_save"] += time.perf_counter() - t_save0

    buffer.seek(0)
    return buffer.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# LAYOUT
# ─────────────────────────────────────────────────────────────────────────────
# Dashboard ile aynı callback ID'leri (gizli): /raporlama açıkken renderer hatası olmasın.
layout = dbc.Container([
    html.Div(
        [
            dcc.Dropdown(
                id='data-type-dropdown',
                options=["İhtiyaç Miktarı", "Sipariş Miktarı", "Öngörü Miktarı"],
                value="İhtiyaç Miktarı",
                style={'display': 'none'},
            ),
            html.Div(id='veri-tipi-loading-hint', style={'display': 'none'}),
            dcc.Interval(id='veri-tipi-load-hide', interval=2000, n_intervals=0, disabled=True),
        ],
        style={'display': 'none'},
        className='kap-raporlama-dash-stubs',
    ),
    html.Div([
        html.Div([
            html.Span("◈", className="kap-header-icon"),
            html.Div([
                html.H1("KAPASİTE RAPORLAMA", className="kap-header-title"),
                html.P(
                    "İhtiyaç, Sipariş ve Öngörü veri tipleri için ayrı Excel dosyaları — tek ZIP olarak iner. Her tablo bloğunun başında cost center adı yazılır.",
                    className="kap-header-sub",
                ),
                html.P(
                    "Excel'de rapor açıldığında sadece bölüm başlıkları (Hepsi, iş merkezi adları) görünür. Sol taraftaki + işaretine tıklayarak ilgili Malzeme veya Kapasite Süresi tablosunu açar, − ile kapatırsınız.",
                    className="kap-header-sub",
                    style={"fontSize": "0.9em", "opacity": 0.85, "marginTop": "4px"},
                ),
            ]),
        ], className="kap-header-inner"),
    ], className="kap-page-header"),

    dbc.Row([
        dbc.Col([
            html.Label(["Veri Tipi"], className="kap-label"),
            html.Div(
                "İhtiyaç · Sipariş · Öngörü (hepsi ZIP içinde)",
                style={
                    "padding": "8px 12px",
                    "backgroundColor": "#e3f2fd",
                    "borderRadius": "8px",
                    "fontWeight": "600",
                },
            ),
        ], md=4),
        dbc.Col([
            html.Label(["Zaman Birimi"], className="kap-label"),
            html.Div(
                "Saat (sat)",
                style={
                    "padding": "8px 12px",
                    "backgroundColor": "#e8f5e9",
                    "borderRadius": "8px",
                    "fontWeight": "600",
                },
            ),
        ], md=4),
    ], className="mb-4"),

    dbc.Row([
        dbc.Col([
            html.Button(
                [html.Span("⬇ ", style={"marginRight": "6px"}), "Raporu Oluştur ve İndir"],
                id="raporlama-btn-generate",
                n_clicks=0,
                disabled=False,
                className="kap-btn kap-btn-primary",
            ),
            dcc.Download(id="raporlama-download"),
        ], md=12),
    ]),

    html.Div(id="raporlama-status", className="mt-3"),
], fluid=True, className="kap-control-panel")



def build_raporlama_zip(ag_instance, hidden_columns=None, table_columns=None):
    
    if ag_instance is None:
        return None, []
    # Dash'ta "Saat" seçili kapasite tablolarıyla birebir hizalı üretim (ZIP job varsayılanı).
    selected_units = ["hours"]
    zip_buffer = io.BytesIO()
    created = []
    do_prof = _rapor_profile_enabled()
    prof = (
        {
            "sql_sec": 0.0,
            "sql_calls": 0,
            "by_table": defaultdict(float),
            "slow_queries": [],
        }
        if do_prof
        else None
    )
    per_excel_rows = []
    zip_t0 = time.perf_counter()
    prof_ctx = _RunQueryProfiler(prof) if prof is not None else nullcontext()
    db_ctx = run_agent.use_connection() if hasattr(run_agent, "use_connection") else nullcontext()
    with prof_ctx, db_ctx:
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for label, table_name, capacity_table_name, period in DATA_TYPES:
                try:
                    df_cc = ag_instance.run_query(f"SELECT DISTINCT STAND FROM [{table_name}] ORDER BY STAND")
                    if df_cc is None or df_cc.empty:
                        continue
                    all_cc_list = df_cc["STAND"].tolist()
                    # Single costcenter override (deneme)
                    if RAPORLAMA_SINGLE_COSTCENTER is not None:
                        if RAPORLAMA_SINGLE_COSTCENTER == "__AUTO__":
                            costcenters = [all_cc_list[0]] if all_cc_list else []
                        else:
                            costcenters = (
                                [RAPORLAMA_SINGLE_COSTCENTER]
                                if RAPORLAMA_SINGLE_COSTCENTER in set(all_cc_list)
                                else []
                            )
                    else:
                        
                        if RAPORLAMA_COSTCENTERS:
                            all_cc = set(all_cc_list)
                            costcenters = [cc for cc in RAPORLAMA_COSTCENTERS if cc in all_cc]
                        else:
                            costcenters = all_cc_list
                    if not costcenters:
                        continue
                except Exception as e:
                    import traceback
                    print(f"[Raporlama] Sorgu hatası ({table_name}): {e}")
                    traceback.print_exc()
                    continue
                if period == "weekly":
                    columns_dict = kapasite_data.generate_weekly_columns()
                else:
                    columns_dict = kapasite_data.generate_monthly_columns_filtered(
                        ag_instance, table_name, capacity_table_name
                    )
                sql_before_excel = prof["sql_sec"] if prof is not None else 0.0
                t_excel_wall0 = time.perf_counter()
                excel_profile = (
                    {"segments": defaultdict(float), "cc_detail_sec": defaultdict(float)}
                    if do_prof
                    else None
                )
                excel_bytes = _build_rapor_excel(
                    costcenters, table_name, capacity_table_name, columns_dict, selected_units,
                    hidden_columns=hidden_columns, table_columns=table_columns,
                    excel_profile=excel_profile,
                )
                t_excel_wall = time.perf_counter() - t_excel_wall0
                if excel_bytes:
                    safe_name = (label
                                 .replace(" ", "_")
                                 .replace("ı", "i").replace("İ", "I")
                                 .replace("ö", "o").replace("Ö", "O")
                                 .replace("ü", "u").replace("Ü", "U")
                                 .replace("ş", "s").replace("Ş", "S")
                                 .replace("ç", "c").replace("Ç", "C")
                                 .replace("ğ", "g").replace("Ğ", "G"))
                    zf.writestr(f"kapasite_rapor_{safe_name}.xlsx", excel_bytes)
                    created.append(f"{label} ({len(costcenters)} CC)")
                    if prof is not None and excel_profile is not None:
                        sql_in_excel = prof["sql_sec"] - sql_before_excel
                        segs = dict(excel_profile["segments"])
                        save_sec = float(segs.get("workbook_save", 0.0))
                        cc_top = sorted(
                            excel_profile["cc_detail_sec"].items(),
                            key=lambda x: -x[1],
                        )[:10]
                        per_excel_rows.append({
                            "label": label,
                            "wall_sec": t_excel_wall,
                            "sql_sec": sql_in_excel,
                            "non_sql_sec": max(0.0, t_excel_wall - sql_in_excel),
                            "save_sec": save_sec,
                            "cc_count": len(costcenters),
                            "segments": segs,
                            "cc_detail_top": cc_top,
                        })
    zip_wall_sec = time.perf_counter() - zip_t0
    if prof is not None:
        _print_rapor_timing_report(zip_wall_sec, prof, per_excel_rows)
    zip_buffer.seek(0)
    return zip_buffer.getvalue() if created else None, created


# ─────────────────────────────────────────────────────────────────────────────
# CALLBACK: Rapor oluştur ve indir
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output("raporlama-download", "data"),
    Output("raporlama-status", "children"),
    Input("raporlama-btn-generate", "n_clicks"),
    prevent_initial_call=True,
)
def raporlama_generate(n_clicks):
    if not n_clicks:
        raise PreventUpdate
    if ag is None:
        email_sent, email_msg = _send_report_notification_email(
            status="BASARISIZ",
            detail="Veritabani baglantisi yok.",
            created_reports=[],
        )
        email_color = "success" if email_sent else "warning"
        return no_update, html.Div([
            dbc.Alert("Veritabanı bağlantısı yok. Lütfen bağlantıyı kontrol edin.", color="danger"),
            dbc.Alert(email_msg, color=email_color),
        ])
    try:
        zip_bytes, created = build_raporlama_zip(ag)
    except Exception as e:
        email_sent, email_msg = _send_report_notification_email(
            status="BASARISIZ",
            detail=f"Rapor olusturma hatasi: {e!r}",
            created_reports=[],
        )
        email_color = "success" if email_sent else "warning"
        return no_update, dbc.Alert(
            [
                f"Rapor oluşturulurken hata: Veritabanı bağlantısı kurulamıyor veya sorgu hatası. ({e!r})",
                dbc.Alert(email_msg, color=email_color, className="mt-2"),
            ],
            color="danger",
        )
    if not zip_bytes:
        email_sent, email_msg = _send_report_notification_email(
            status="BASARISIZ",
            detail="Hicbir veri tipi icin rapor olusturulamadi.",
            created_reports=[],
        )
        email_color = "success" if email_sent else "warning"
        return no_update, html.Div([
            dbc.Alert("Hiçbir veri tipi için rapor oluşturulamadı. Veritabanında veri olmayabilir.", color="warning"),
            dbc.Alert(email_msg, color=email_color),
        ])

    email_sent, email_msg = _send_report_notification_email(
        status="BASARILI",
        detail="ZIP dosyasi basariyla olusturuldu ve indirildi.",
        created_reports=created,
        attachment_bytes=zip_bytes,
        attachment_name="kapasite_raporlar.zip",
    )
    email_color = "success" if email_sent else "warning"
    return (
        dict(
            content=base64.b64encode(zip_bytes).decode("ascii"),
            filename="kapasite_raporlar.zip",
            base64=True,
            type="application/zip",
        ),
        html.Div([
            dbc.Alert(f"Raporlar oluşturuldu: {', '.join(created)}. ZIP indiriliyor.", color="success"),
            dbc.Alert(email_msg, color=email_color),
        ]),
    )